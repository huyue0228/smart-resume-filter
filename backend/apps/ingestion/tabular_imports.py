"""四类表格导入的标准模板与表头契约。"""

import re
from collections import Counter
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


@dataclass(frozen=True)
class TabularImportSchema:
    key: str
    label: str
    filename: str
    sheet_name: str
    headers: tuple[str, ...]


IMPORT_TABLE_SCHEMAS = {
    "resume_list": TabularImportSchema(
        key="resume_list",
        label="简历信息列表",
        filename="简历信息列表标准模板.xlsx",
        sheet_name="简历信息列表",
        headers=(
            "招聘主体",
            "所属机构",
            "姓名",
            "应聘ID",
            "手机号",
            "性别",
            "对外职位名称",
            "学历",
            "第一学历毕业院校",
            "最高学历毕业院校",
            "最高学历专业",
            "应聘状态",
            "户口所在地",
            "应聘日期",
        ),
    ),
    "jobs": TabularImportSchema(
        key="jobs",
        label="岗位",
        filename="岗位标准模板.xlsx",
        sheet_name="职位清单",
        headers=(
            "招聘主体",
            "一层部门",
            "二层部门",
            "岗位类别",
            "对外发布名称",
            "是否对外发布",
            "职位名称",
            "岗位族",
            "工作地点",
            "学历",
            "工作职责",
            "需求专业",
            "HC",
        ),
    ),
    "schools": TabularImportSchema(
        key="schools",
        label="院校分类",
        filename="院校分类标准模板.xlsx",
        sheet_name="院校分类",
        headers=("学校", "院校标签"),
    ),
    "contacts": TabularImportSchema(
        key="contacts",
        label="部门接口人",
        filename="部门接口人标准模板.xlsx",
        sheet_name="部门接口人",
        headers=(
            "一层部门",
            "二层部门",
            "三级部门",
            "姓名",
            "工号",
            "邮箱",
            "接口人层级",
            "可转派",
            "是否启用",
        ),
    ),
}


def get_import_table_schema(schema_key):
    try:
        return IMPORT_TABLE_SCHEMAS[schema_key]
    except KeyError as exc:
        raise ValueError("未知导入模板类型") from exc


def validate_table_headers(table, schema_key):
    """严格校验标准模板表头，不猜测或映射客户自定义标签。"""
    schema = get_import_table_schema(schema_key)
    actual_headers = [str(column) for column in table.columns]
    expected_headers = list(schema.headers)
    actual_counts = Counter(actual_headers)
    duplicates = {header for header, count in actual_counts.items() if count > 1}

    # pandas 会把重复表头改写为“字段.1”，还原该信息以返回明确的重复字段。
    pandas_duplicate_headers = set()
    for header in actual_headers:
        match = re.fullmatch(r"(.+)\.(\d+)", header)
        if match and match.group(1) in actual_headers:
            pandas_duplicate_headers.add(match.group(1))
    duplicates.update(pandas_duplicate_headers)

    missing = [header for header in expected_headers if header not in actual_counts]
    unknown = []
    for header in actual_headers:
        match = re.fullmatch(r"(.+)\.(\d+)", header)
        if match and match.group(1) in pandas_duplicate_headers:
            continue
        if header not in expected_headers and header not in unknown:
            unknown.append(header)

    issues = []
    if missing:
        issues.append(f"缺少字段【{'、'.join(missing)}】")
    if unknown:
        issues.append(f"未知字段【{'、'.join(unknown)}】")
    if duplicates:
        ordered_duplicates = [
            header for header in expected_headers if header in duplicates
        ]
        ordered_duplicates.extend(
            header
            for header in actual_headers
            if header in duplicates and header not in ordered_duplicates
        )
        issues.append(f"重复字段【{'、'.join(ordered_duplicates)}】")
    if issues:
        raise ValueError(
            f"{schema.label}表头不符合标准模板：{'；'.join(issues)}。"
            "请下载最新版标准模板后重新填写"
        )


def build_import_template_workbook(schema_key):
    """根据同一份表头契约生成可填写的标准 XLSX 模板。"""
    schema = get_import_table_schema(schema_key)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = schema.sheet_name
    sheet.append(list(schema.headers))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for index, cell in enumerate(sheet[1], start=1):
        cell.font = Font(bold=True)
        cell.fill = header_fill
        sheet.column_dimensions[get_column_letter(index)].width = min(
            36, max(12, len(str(cell.value)) * 2 + 4)
        )

    list_options = {
        "是否对外发布": "是,否",
        "性别": "男,女",
        "接口人层级": "二级接口人,三级接口人",
        "可转派": "是,否",
        "是否启用": "是,否",
    }
    for index, header in enumerate(schema.headers, start=1):
        options = list_options.get(header)
        if not options:
            continue
        validation = DataValidation(type="list", formula1=f'"{options}"')
        sheet.add_data_validation(validation)
        column = get_column_letter(index)
        validation.add(f"{column}2:{column}5000")

    instructions = workbook.create_sheet("填写说明")
    instructions.append(["模板版本", "1"])
    instructions.append(["模板类型", schema.label])
    instructions.append(
        ["填写要求", "请勿修改、删除、重复或新增第一行表头；无值字段保留列并留空。"]
    )
    instructions.append(
        ["数据位置", f"请在“{schema.sheet_name}”工作表第二行起填写数据。"]
    )
    instructions.column_dimensions["A"].width = 16
    instructions.column_dimensions["B"].width = 80
    return workbook
