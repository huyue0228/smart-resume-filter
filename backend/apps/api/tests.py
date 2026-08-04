from datetime import datetime, timedelta
from io import BytesIO
import importlib
import os
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch
from types import SimpleNamespace
import zipfile
from urllib.parse import quote

from django.contrib.auth.models import Group, Permission
from django.apps import apps as django_apps
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.utils import timezone
import pandas as pd
from openpyxl import load_workbook
from rest_framework.authtoken.models import Token
from rest_framework.test import APIClient

from apps.accounts.models import User
from apps.accounts.permissions import ensure_rbac_defaults, permission_codename
from apps.accounts.protected_users import (
    PROTECTED_ADMIN_EMAIL,
    PROTECTED_ADMIN_USERNAME,
)
from apps.core import models as m
from apps.ingestion.tabular_imports import (
    build_import_template_workbook,
    get_import_table_schema,
)
from apps.pipeline import ai_config
from apps.pipeline.ai.service import AIServiceError
from apps.pipeline.services import classify_school


def rest_framework_test_settings():
    return {
        "DEFAULT_PERMISSION_CLASSES": ["rest_framework.permissions.IsAuthenticated"],
        "DEFAULT_AUTHENTICATION_CLASSES": [
            "rest_framework.authentication.TokenAuthentication"
        ],
        "DEFAULT_PAGINATION_CLASS": "rest_framework.pagination.PageNumberPagination",
        "PAGE_SIZE": 20,
    }


def standard_import_template_bytes(template_type):
    output = BytesIO()
    build_import_template_workbook(template_type).save(output)
    return output.getvalue()


@override_settings(
    REST_FRAMEWORK={
        "DEFAULT_PERMISSION_CLASSES": ["rest_framework.permissions.IsAuthenticated"],
        "DEFAULT_AUTHENTICATION_CLASSES": [
            "rest_framework.authentication.TokenAuthentication"
        ],
        "DEFAULT_PAGINATION_CLASS": "rest_framework.pagination.PageNumberPagination",
        "PAGE_SIZE": 20,
    }
)
class AgentDispatchDecisionApiTests(TestCase):
    def setUp(self):
        ai_config.save_ai_connection_config(
            {
                "api_style": "responses",
                "model_name": "gpt-test",
                "base_url": "https://model.internal/v1",
                "api_key": "test-key",
            }
        )
        ai_config.mark_ai_connection_tested()
        self.client = APIClient()
        ensure_rbac_defaults()
        self.user = User.objects.create_user(
            username="hr", password="pass", role=User.ROLE_HR
        )
        self.user.groups.add(Group.objects.get(name="HR"))
        self.client.force_authenticate(self.user)
        department = m.Department.objects.create(name="技术部", level=2)
        m.Contact.objects.create(
            name="二级接口人",
            employee_no="L2001",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
            is_active=True,
        )
        self.candidate = m.Candidate.objects.create(
            identity_hash="candidate-1",
            name="张三",
            phone="13800000000",
            first_degree_platform="平台A",
            highest_degree_platform="平台A",
        )
        self.resume = m.Resume.objects.create(
            candidate=self.candidate,
            apply_id="A1001",
            position_name="后端工程师",
            volunteer_rank=1,
        )
        m.Job.objects.create(
            department=department,
            public_name="后端工程师",
            position_name="后端工程师",
            category="技术类",
            responsibilities="负责后端服务开发、维护和性能优化。",
            headcount=1,
        )
        self.workflow = m.CandidateWorkflow.objects.create(
            candidate=self.candidate,
            status=m.CandidateWorkflow.STATUS_ARCHIVED,
            current_resume=self.resume,
            current_rank=1,
            dispatch_strategy="ai",
            archive_reason=m.CandidateWorkflow.ARCHIVE_AGENT_NO_RECOMMENDATION,
        )
        self.decision = m.AgentDispatchDecision.objects.create(
            workflow=self.workflow,
            resume=self.resume,
            recommendation=None,
            confidence_score=None,
            error_code="pdf_missing",
            error_message="缺少 PDF 简历文件",
            prompt_version="demo-v1",
            decision_version="demo-v1",
        )

    def test_specialist_audit_fields_are_not_exposed_by_decision_api(self):
        self.decision.ai_specialist_match = True
        self.decision.ai_specialist_confidence = 0.96
        self.decision.ai_specialist_evidence = ["内部证据"]
        self.decision.special_route_applied = True
        self.decision.special_route_config_snapshot = {"version": "internal"}
        self.decision.save()

        response = self.client.get(f"/api/agent-decisions/{self.decision.id}/")

        self.assertEqual(response.status_code, 200)
        for field in {
            "ai_specialist_match",
            "ai_specialist_confidence",
            "ai_specialist_evidence",
            "special_route_applied",
            "special_route_config_snapshot",
        }:
            self.assertNotIn(field, response.data)

    def test_special_route_attempt_is_presented_as_normal_ai_assignment(self):
        contact = m.Contact.objects.get(employee_no="L2001")
        attempt = m.AssignmentAttempt.objects.create(
            workflow=self.workflow,
            resume=self.resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_AI,
            status=m.AssignmentAttempt.STATUS_ASSIGNED_L3,
            department=contact.department,
            contact=contact,
            agent_decision=self.decision,
            match_mode="ai",
            match_reason="AI 专项强制分配：内部审计信息",
            route_code="ai_special_route",
            special_route_confidence=0.96,
            special_route_evidence=["内部证据"],
            special_route_config_snapshot={"version": "internal"},
        )

        response = self.client.get(f"/api/workflow-attempts/{attempt.id}/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["match_reason"], "AI 自动分配")
        for field in {
            "route_code",
            "special_route_confidence",
            "special_route_evidence",
            "special_route_config_snapshot",
        }:
            self.assertNotIn(field, response.data)

    def test_retry_records_new_failed_decision_when_pdf_is_still_missing(self):
        response = self.client.post(f"/api/agent-decisions/{self.decision.id}/retry/")

        self.assertEqual(response.status_code, 202)
        self.assertFalse(m.AssignmentAttempt.objects.exists())
        self.assertEqual(m.AgentDispatchDecision.objects.count(), 2)
        new_decision = m.AgentDispatchDecision.objects.order_by("-id").first()
        self.assertNotEqual(new_decision.id, self.decision.id)
        self.assertIsNone(new_decision.recommendation)
        self.assertIsNone(new_decision.confidence_score)
        self.assertEqual(new_decision.error_code, "pdf_missing")
        run = m.ProcessingRun.objects.get()
        self.assertEqual(response.data["run"]["id"], run.id)
        self.assertEqual(run.mode, "ai")
        self.assertEqual(run.scope["source"], "ai_retry")
        self.assertEqual(run.scope["retry_decision_id"], self.decision.id)
        self.assertEqual(new_decision.processing_run_id, run.id)

    def test_retry_is_disabled_when_ai_connection_is_not_ready(self):
        ai_config.invalidate_ai_connection_test()

        response = self.client.post(f"/api/agent-decisions/{self.decision.id}/retry/")

        self.assertEqual(response.status_code, 400)
        self.assertIn("模型连接尚未测试成功", response.data["detail"])
        self.assertEqual(m.AgentDispatchDecision.objects.count(), 1)

    def test_retry_rejects_high_confidence_dispatch_decision(self):
        self.decision.recommendation = m.AgentDispatchDecision.RECOMMEND_DISPATCH
        self.decision.confidence_score = 0.88
        self.decision.error_code = ""
        self.decision.error_message = ""
        self.decision.save(
            update_fields=[
                "recommendation",
                "confidence_score",
                "error_code",
                "error_message",
            ]
        )

        response = self.client.post(f"/api/agent-decisions/{self.decision.id}/retry/")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(m.AgentDispatchDecision.objects.count(), 1)

    def test_retry_task_keeps_the_original_decision_resume(self):
        m.Resume.objects.create(
            candidate=self.candidate,
            apply_id="A0000",
            position_name="后端工程师",
            volunteer_rank=0,
        )

        with patch(
            "apps.pipeline.services.allocate.ai_service.screen_resume",
            side_effect=AIServiceError("pdf_missing", "缺少 PDF 简历文件"),
        ) as screen_resume:
            response = self.client.post(
                f"/api/agent-decisions/{self.decision.id}/retry/"
            )

        self.assertEqual(response.status_code, 202)
        self.assertEqual(screen_resume.call_args.args[0].id, self.resume.id)

    def test_retry_cancels_old_active_ai_attempt_before_creating_new_one(self):
        self.resume.resume_file = "张三（A1001）.pdf"
        self.resume.save(update_fields=["resume_file"])
        self.candidate.highest_major = "计算机科学与技术"
        self.candidate.save(update_fields=["highest_major"])
        old_attempt = m.AssignmentAttempt.objects.create(
            workflow=self.workflow,
            resume=self.resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_AI,
            status=m.AssignmentAttempt.STATUS_PENDING_REVIEW,
            contact=m.Contact.objects.get(employee_no="L2001"),
            department=m.Department.objects.get(name="技术部"),
            agent_decision=self.decision,
            confidence_score=0.55,
            review_required=True,
            match_mode="ai",
        )

        profile = m.ResumeProfile.objects.create(
            resume=self.resume, parse_status="parsed", raw_text="简历正文"
        )
        job = m.Job.objects.get()
        m.JobMajor.objects.create(job=job, major="电气工程")
        contact = m.Contact.objects.get(employee_no="L2001")
        decision_output = SimpleNamespace(
            recommendation="dispatch",
            summary="结构化 AI 建议",
            reason="证据充分",
            evidence=["项目经历"],
            risks=[],
        )
        result = SimpleNamespace(
            profile=profile,
            output=SimpleNamespace(
                decision=decision_output,
                profile=SimpleNamespace(risk_flags=[]),
            ),
            job=job,
            department=contact.department,
            contact=contact,
            confidence=0.78,
            score_breakdown={"major_match": 0.78},
        )
        with patch(
            "apps.pipeline.services.allocate.ai_service.screen_resume",
            return_value=result,
        ):
            response = self.client.post(f"/api/agent-decisions/{self.decision.id}/retry/")

        self.assertEqual(response.status_code, 202)
        old_attempt.refresh_from_db()
        self.assertEqual(old_attempt.status, m.AssignmentAttempt.STATUS_CANCELLED)
        self.assertEqual(old_attempt.cancel_reason, m.AssignmentAttempt.CANCEL_RERUN)
        self.assertEqual(
            m.AssignmentAttempt.objects.filter(
                status__in=[
                    m.AssignmentAttempt.STATUS_PENDING_REVIEW,
                    m.AssignmentAttempt.STATUS_PENDING_DISPATCH,
                ]
            ).count(),
            1,
        )
        item = m.ProcessingRunScopeItem.objects.get()
        self.assertEqual(item.reason_code, "ai_dispatched")
        self.assertNotEqual(item.reason_code, "major_not_matched")


@override_settings(REST_FRAMEWORK=rest_framework_test_settings())
class PipelineRunApiTests(TestCase):
    def setUp(self):
        ensure_rbac_defaults()
        self.client = APIClient()
        self.user = User.objects.create_user(
            username="hr-pipeline", password="pass", role=User.ROLE_HR
        )
        self.user.groups.add(Group.objects.get(name="HR"))
        self.client.force_authenticate(self.user)

    def test_pipeline_run_forwards_scope_to_runner(self):
        run = m.ProcessingRun.objects.create(
            step="step2",
            mode="rule",
            status="success",
            message="ok",
        )
        scope = {
            "system_statuses": ["screening_passed", "screening_rejected"],
            "candidate_filters": {"system_status": "screening_passed,screening_rejected"},
        }

        with patch("apps.api.views.runner.create_run", return_value=run) as mock_create, patch(
            "apps.api.views.execute_runs_sequence_task.delay",
            return_value=SimpleNamespace(id="task-123"),
        ):
            response = self.client.post(
                "/api/pipeline/run/",
                {"step": "step2", "mode": "rule", "scope": scope},
                format="json",
            )

        self.assertEqual(response.status_code, 200)
        mock_create.assert_called_once_with(
            "step2", mode="rule", scope=scope, created_by=self.user
        )
        self.assertEqual(response.data["processing_runs"][0]["message"], "ok")

    def test_pipeline_run_rejects_removed_and_unknown_system_statuses(self):
        for status_code in ["classified", "allocated", "unknown"]:
            response = self.client.post(
                "/api/pipeline/run/",
                {
                    "step": "step2",
                    "mode": "rule",
                    "scope": {"system_statuses": [status_code]},
                },
                format="json",
            )
            self.assertEqual(response.status_code, 400)
            self.assertIn(status_code, response.data["detail"])

    def test_pipeline_run_accepts_selected_force_reprocess_scope(self):
        run = m.ProcessingRun.objects.create(
            step="step2",
            mode="rule",
            status="success",
            message="ok",
        )
        scope = {"candidate_ids": [3, 5], "force_reprocess": True}

        with patch(
            "apps.api.views.runner.create_run", return_value=run
        ) as mock_create, patch(
            "apps.api.views.execute_runs_sequence_task.delay",
            return_value=SimpleNamespace(id="task-force"),
        ):
            response = self.client.post(
                "/api/pipeline/run/",
                {"step": "step2", "mode": "rule", "scope": scope},
                format="json",
            )

        self.assertEqual(response.status_code, 200)
        mock_create.assert_called_once_with(
            "step2", mode="rule", scope=scope, created_by=self.user
        )

    def test_pipeline_run_rejects_invalid_force_reprocess_scopes(self):
        invalid_requests = [
            (
                {"step": "step1", "scope": {"candidate_ids": [1], "force_reprocess": True}},
                "step2",
            ),
            ({"step": "step2", "scope": {"force_reprocess": True}}, "candidate_ids"),
            (
                {"step": "step2", "scope": {"candidate_ids": [], "force_reprocess": True}},
                "candidate_ids",
            ),
            (
                {
                    "step": "step2",
                    "scope": {"candidate_ids": [True], "force_reprocess": True},
                },
                "candidate_ids",
            ),
            (
                {
                    "step": "step2",
                    "scope": {
                        "candidate_ids": [1],
                        "force_reprocess": True,
                        "system_statuses": [],
                    },
                },
                "不得与",
            ),
            (
                {
                    "step": "step2",
                    "scope": {
                        "candidate_ids": [1],
                        "force_reprocess": True,
                        "candidate_filters": {},
                    },
                },
                "不得与",
            ),
            (
                {
                    "step": "step2",
                    "scope": {"candidate_ids": [1], "force_reprocess": "true"},
                },
                "布尔值",
            ),
            (
                {
                    "step": "step2",
                    "scope": {"candidate_ids": [1], "force_reprocess": False},
                },
                "true",
            ),
        ]

        for payload, expected_detail in invalid_requests:
            with self.subTest(payload=payload):
                response = self.client.post(
                    "/api/pipeline/run/", payload, format="json"
                )
                self.assertEqual(response.status_code, 400)
                self.assertIn(expected_detail, response.data["detail"])

        self.assertFalse(m.ProcessingRun.objects.exists())

    def test_pipeline_run_rejects_caller_mode_override(self):
        response = self.client.post(
            "/api/pipeline/run/",
            {"step": "step2", "modes": ["rule", "ai"]},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("不接受 modes", response.data["detail"])

    def test_pipeline_run_rejects_ai_when_connection_is_not_ready(self):
        with patch(
            "apps.pipeline.ai_config.is_ai_connection_tested",
            return_value=False,
        ):
            response = self.client.post(
                "/api/pipeline/run/",
                {"step": "step2", "mode": "ai"},
                format="json",
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn("尚未测试成功", response.data["detail"])

    def test_pipeline_run_rejects_explicit_empty_modes(self):
        response = self.client.post(
            "/api/pipeline/run/",
            {"step": "step2", "modes": []},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("不接受 modes", response.data["detail"])

    def test_pipeline_run_requires_one_explicit_mode(self):
        response = self.client.post(
            "/api/pipeline/run/", {"step": "step2"}, format="json"
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("rule 或 ai", response.data["detail"])

    def test_pipeline_run_rejects_reserved_ai_retry_scope(self):
        response = self.client.post(
            "/api/pipeline/run/",
            {
                "step": "step2",
                "scope": {
                    "source": "ai_retry",
                    "retry_decision_id": 1,
                    "retry_resume_id": 1,
                },
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("简历详情", response.data["detail"])
        self.assertFalse(m.ProcessingRun.objects.exists())

    def test_processing_run_list_includes_elapsed_seconds(self):
        now = timezone.now()
        running = m.ProcessingRun.objects.create(
            step="step2", mode="rule", status="running"
        )
        finished = m.ProcessingRun.objects.create(
            step="step2", mode="ai", status="success"
        )
        m.ProcessingRun.objects.filter(pk=running.pk).update(
            created_at=now - timedelta(seconds=75)
        )
        m.ProcessingRun.objects.filter(pk=finished.pk).update(
            created_at=now - timedelta(seconds=130),
            finished_at=now - timedelta(seconds=65),
        )

        with patch("apps.api.serializers.timezone.now", return_value=now):
            response = self.client.get("/api/pipeline/runs/")

        self.assertEqual(response.status_code, 200)
        runs = {item["id"]: item for item in response.data["results"]}
        self.assertEqual(runs[running.id]["elapsed_seconds"], 75)
        self.assertEqual(runs[finished.id]["elapsed_seconds"], 65)

    def test_pending_processing_run_can_be_cancelled(self):
        run = m.ProcessingRun.objects.create(step="step2", mode="ai", status="pending")
        stage = m.ProcessingRunStage.objects.create(
            run=run, sequence=1, step="step2", label="院校分类与学历/院校准入"
        )

        response = self.client.post(f"/api/pipeline/runs/{run.id}/cancel/")

        self.assertEqual(response.status_code, 200)
        run.refresh_from_db()
        stage.refresh_from_db()
        self.assertEqual(run.status, "cancelled")
        self.assertEqual(run.cancelled_by, self.user)
        self.assertTrue(run.cancel_requested_at)
        self.assertTrue(run.cancelled_at)
        self.assertTrue(run.finished_at)
        self.assertEqual(stage.status, "cancelled")

    def test_finished_processing_run_cannot_be_cancelled(self):
        run = m.ProcessingRun.objects.create(
            step="step2", mode="rule", status="success"
        )

        response = self.client.post(f"/api/pipeline/runs/{run.id}/cancel/")

        self.assertEqual(response.status_code, 409)
        run.refresh_from_db()
        self.assertEqual(run.status, "success")


@override_settings(
    REST_FRAMEWORK={
        "DEFAULT_PERMISSION_CLASSES": ["rest_framework.permissions.IsAuthenticated"],
        "DEFAULT_AUTHENTICATION_CLASSES": [
            "rest_framework.authentication.TokenAuthentication"
        ],
        "DEFAULT_PAGINATION_CLASS": "rest_framework.pagination.PageNumberPagination",
        "PAGE_SIZE": 20,
    }
)
class RbacApiTests(TestCase):
    def setUp(self):
        ensure_rbac_defaults()
        self.client = APIClient()
        self.hr = User.objects.create_user(
            username="hr",
            email="hr@example.com",
            password="pass",
            role=User.ROLE_HR,
        )
        self.hr.groups.add(Group.objects.get(name="HR"))
        self.admin = User.objects.create_user(
            username="admin", password="pass", role=User.ROLE_ADMIN
        )
        self.admin.groups.add(Group.objects.get(name="管理员"))

        self.dept_a = m.Department.objects.create(name="技术二部", level=2)
        self.dept_b = m.Department.objects.create(name="产品二部", level=2)
        self.secondary_a = m.Contact.objects.create(
            name="技术二级接口人",
            employee_no="S-A",
            department=self.dept_a,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        self.secondary_b = m.Contact.objects.create(
            name="产品二级接口人",
            employee_no="S-B",
            department=self.dept_b,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        self.tertiary_a = m.Contact.objects.create(
            name="技术三级接口人",
            employee_no="T-A",
            department=m.Department.objects.create(
                name="技术三级组", level=3, parent=self.dept_a
            ),
            contact_level=m.Contact.LEVEL_TERTIARY,
        )
        self.secondary_user = User.objects.create_user(
            username="secondary-a",
            password="pass",
            role=User.ROLE_SECONDARY_CONTACT,
            contact=self.secondary_a,
        )
        self.secondary_user.groups.add(Group.objects.get(name="二级接口人"))
        self.tertiary_user = User.objects.create_user(
            username="tertiary-a",
            password="pass",
            role=User.ROLE_TERTIARY_CONTACT,
            contact=self.tertiary_a,
        )
        self.tertiary_user.groups.add(Group.objects.get(name="三级接口人"))

        self.attempt_a = self._attempt(
            "candidate-a", "张三", "A1001", self.dept_a, self.secondary_a
        )
        self.attempt_b = self._attempt(
            "candidate-b", "李四", "B1001", self.dept_b, self.secondary_b
        )
        self.attempt_a.sub_contact = self.tertiary_a
        self.attempt_a.sub_department = self.tertiary_a.department
        self.attempt_a.status = m.AssignmentAttempt.STATUS_ASSIGNED_L3
        self.attempt_a.save(update_fields=["sub_contact", "sub_department", "status"])

    def _attempt(self, identity, name, apply_id, department, contact):
        candidate = m.Candidate.objects.create(
            identity_hash=identity,
            name=name,
            phone="13800000000",
        )
        resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id=apply_id,
            position_name="后端工程师",
            volunteer_rank=1,
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=resume,
            current_rank=1,
        )
        return m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
            department=department,
            contact=contact,
        )

    def test_local_password_login_route_is_removed(self):
        response = self.client.post(
            "/api/auth/login/", {"username": "hr", "password": "pass"}, format="json"
        )

        self.assertEqual(response.status_code, 404)

    def test_secondary_contact_only_sees_own_received_attempts(self):
        self.client.force_authenticate(self.secondary_user)

        response = self.client.get("/api/workflow-attempts/")

        self.assertEqual(response.status_code, 200)
        ids = [item["id"] for item in response.data["results"]]
        self.assertEqual(ids, [self.attempt_a.id])
        self.assertEqual(response.data["results"][0]["volunteer_rank"], 1)
        self.assertEqual(response.data["results"][0]["apply_id"], "A1001")
        self.assertEqual(response.data["results"][0]["position_name"], "后端工程师")

    def test_secondary_contact_does_not_see_hr_pending_attempts(self):
        pending_candidate = m.Candidate.objects.create(
            identity_hash="candidate-pending",
            name="王五",
            phone="13700000000",
        )
        pending_resume = m.Resume.objects.create(
            candidate=pending_candidate,
            apply_id="P1001",
            position_name="算法工程师",
            volunteer_rank=1,
        )
        pending_workflow = m.CandidateWorkflow.objects.create(
            candidate=pending_candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=pending_resume,
            current_rank=1,
        )
        m.AssignmentAttempt.objects.create(
            workflow=pending_workflow,
            resume=pending_resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_AI,
            status=m.AssignmentAttempt.STATUS_PENDING_DISPATCH,
            department=self.dept_a,
            contact=self.secondary_a,
        )
        m.AssignmentAttempt.objects.create(
            workflow=pending_workflow,
            resume=pending_resume,
            attempt_no=2,
            source=m.AssignmentAttempt.SOURCE_AI,
            status=m.AssignmentAttempt.STATUS_PENDING_REVIEW,
            department=self.dept_a,
            contact=self.secondary_a,
        )
        self.client.force_authenticate(self.secondary_user)

        response = self.client.get("/api/workflow-attempts/")

        self.assertEqual(response.status_code, 200)
        ids = [item["id"] for item in response.data["results"]]
        self.assertEqual(ids, [self.attempt_a.id])

    def test_tertiary_contact_only_sees_assigned_attempts(self):
        self.client.force_authenticate(self.tertiary_user)

        response = self.client.get("/api/workflow-attempts/")

        self.assertEqual(response.status_code, 200)
        ids = [item["id"] for item in response.data["results"]]
        self.assertEqual(ids, [self.attempt_a.id])

    def _grant_secondary_feedback(self):
        group = Group.objects.get(name="二级接口人")
        group.permissions.add(
            Permission.objects.get(
                codename=permission_codename("attempt.feedback")
            )
        )

    def test_secondary_contact_with_permission_can_feedback_before_transfer(self):
        self.attempt_a.status = m.AssignmentAttempt.STATUS_DISPATCHED_L2
        self.attempt_a.sub_contact = None
        self.attempt_a.sub_department = None
        self.attempt_a.save(
            update_fields=["status", "sub_contact", "sub_department"]
        )
        self._grant_secondary_feedback()
        self.client.force_authenticate(self.secondary_user)

        response = self.client.post(
            f"/api/workflow-attempts/{self.attempt_a.id}/feedback/",
            {"result": "passed", "note": "二级确认通过"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.attempt_a.refresh_from_db()
        self.assertEqual(self.attempt_a.status, m.AssignmentAttempt.STATUS_PASSED)
        self.assertEqual(self.attempt_a.feedback_note, "二级确认通过")

    def test_secondary_contact_cannot_feedback_after_transfer(self):
        self._grant_secondary_feedback()
        self.client.force_authenticate(self.secondary_user)

        response = self.client.post(
            f"/api/workflow-attempts/{self.attempt_a.id}/feedback/",
            {"result": "rejected", "note": "不应允许"},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("绑定", response.data["detail"])

    def test_tertiary_contact_can_feedback_after_transfer_and_cannot_repeat(self):
        self.client.force_authenticate(self.tertiary_user)

        first_response = self.client.post(
            f"/api/workflow-attempts/{self.attempt_a.id}/feedback/",
            {"result": "passed", "note": "三级确认"},
            format="json",
        )
        second_response = self.client.post(
            f"/api/workflow-attempts/{self.attempt_a.id}/feedback/",
            {"result": "rejected", "note": "重复"},
            format="json",
        )

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(second_response.status_code, 400)
        self.assertIn("反馈已提交", second_response.data["detail"])

    def test_secondary_contact_without_permission_cannot_feedback(self):
        self.attempt_a.status = m.AssignmentAttempt.STATUS_DISPATCHED_L2
        self.attempt_a.sub_contact = None
        self.attempt_a.sub_department = None
        self.attempt_a.save(
            update_fields=["status", "sub_contact", "sub_department"]
        )
        self.client.force_authenticate(self.secondary_user)

        response = self.client.post(
            f"/api/workflow-attempts/{self.attempt_a.id}/feedback/",
            {"result": "passed"},
            format="json",
        )

        self.assertEqual(response.status_code, 403)

    def test_feedback_permission_does_not_bypass_contact_binding_or_data_scope(self):
        self.client.force_authenticate(self.admin)
        admin_response = self.client.post(
            f"/api/workflow-attempts/{self.attempt_a.id}/feedback/",
            {"result": "passed"},
            format="json",
        )
        self._grant_secondary_feedback()
        self.client.force_authenticate(self.secondary_user)
        other_response = self.client.post(
            f"/api/workflow-attempts/{self.attempt_b.id}/feedback/",
            {"result": "passed"},
            format="json",
        )

        self.assertEqual(admin_response.status_code, 400)
        self.assertIn("绑定", admin_response.data["detail"])
        self.assertEqual(other_response.status_code, 404)

    def test_secondary_contact_uses_resume_library_with_scoped_safe_fields(self):
        self.client.force_authenticate(self.secondary_user)

        response = self.client.get("/api/candidates/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual([item["id"] for item in response.data["results"]], [self.attempt_a.workflow.candidate_id])
        row = response.data["results"][0]
        self.assertEqual(row["phone"], "")
        self.assertIsNone(row["preview_resume"])
        self.assertEqual([item["id"] for item in row["resumes"]], [self.attempt_a.resume_id])
        self.assertEqual([item["id"] for item in row["attempts"]], [self.attempt_a.id])
        self.assertEqual(row["current_attempt"]["id"], self.attempt_a.id)
        self.assertNotIn("agent_decision", row["attempts"][0])
        self.assertNotIn("agent_decision_summary", row["current_attempt"])

    def test_tertiary_contact_uses_resume_library_with_only_own_assignment(self):
        self.client.force_authenticate(self.tertiary_user)

        response = self.client.get("/api/candidates/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual([item["id"] for item in response.data["results"]], [self.attempt_a.workflow.candidate_id])
        row = response.data["results"][0]
        self.assertEqual(row["phone"], "")
        self.assertEqual(row["current_attempt"]["sub_contact"], self.tertiary_a.id)
        self.assertEqual(len(row["resumes"]), 1)
        self.assertEqual(len(row["attempts"]), 1)

    def test_hr_bulk_dispatches_pending_attempt_from_resume_library(self):
        attempt = self._attempt(
            "candidate-pending-dispatch",
            "王五",
            "C1001",
            self.dept_a,
            self.secondary_a,
        )
        attempt.status = m.AssignmentAttempt.STATUS_PENDING_DISPATCH
        attempt.save(update_fields=["status"])
        self.client.force_authenticate(self.hr)

        response = self.client.post(
            "/api/candidates/bulk-dispatch/",
            {"candidate_ids": [attempt.workflow.candidate_id]},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["dispatched"], 1)
        self.assertEqual(response.data["skipped"], 0)
        attempt.refresh_from_db()
        self.assertEqual(attempt.status, m.AssignmentAttempt.STATUS_DISPATCHED_L2)

    def test_bulk_dispatch_validates_scope_and_reports_partial_success(self):
        pending = self._attempt(
            "candidate-bulk-pending", "批量待下发", "BD1001", self.dept_a,
            self.secondary_a,
        )
        pending.status = m.AssignmentAttempt.STATUS_PENDING_DISPATCH
        pending.save(update_fields=["status"])
        self.client.force_authenticate(self.hr)

        response = self.client.post(
            "/api/candidates/bulk-dispatch/",
            {
                "candidate_ids": [
                    pending.workflow.candidate_id,
                    self.attempt_a.workflow.candidate_id,
                ]
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["total"], 2)
        self.assertEqual(response.data["eligible"], 1)
        self.assertEqual(response.data["dispatched"], 1)
        self.assertEqual(response.data["skipped"], 1)
        self.assertEqual(response.data["failed"], 0)

        invalid_bodies = [
            {},
            {"candidate_ids": []},
            {"candidate_ids": [pending.workflow.candidate_id], "candidate_filters": {}},
            {"candidate_filters": {}},
            {"candidate_filters": {"name": "张三"}},
            {"candidate_filters": {"system_statuses": ["allocated"]}},
        ]
        for body in invalid_bodies:
            invalid = self.client.post(
                "/api/candidates/bulk-dispatch/", body, format="json"
            )
            self.assertEqual(invalid.status_code, 400, body)

        ai_pending = self._attempt(
            "candidate-bulk-ai", "AI 待下发", "BD1002", self.dept_a,
            self.secondary_a,
        )
        ai_pending.status = m.AssignmentAttempt.STATUS_PENDING_DISPATCH
        ai_pending.source = m.AssignmentAttempt.SOURCE_AI
        ai_pending.save(update_fields=["status", "source"])
        rule_pending = self._attempt(
            "candidate-bulk-rule", "Rule 待下发", "BD1003", self.dept_a,
            self.secondary_a,
        )
        rule_pending.status = m.AssignmentAttempt.STATUS_PENDING_DISPATCH
        rule_pending.save(update_fields=["status"])

        filtered = self.client.post(
            "/api/candidates/bulk-dispatch/",
            {
                "candidate_filters": {
                    "system_statuses": ["pending_dispatch"],
                    "allocation_source": ["ai"],
                }
            },
            format="json",
        )

        self.assertEqual(filtered.status_code, 200)
        self.assertEqual(filtered.data["total"], 1)
        self.assertEqual(filtered.data["dispatched"], 1)
        ai_pending.refresh_from_db()
        rule_pending.refresh_from_db()
        self.assertEqual(ai_pending.status, m.AssignmentAttempt.STATUS_DISPATCHED_L2)
        self.assertEqual(rule_pending.status, m.AssignmentAttempt.STATUS_PENDING_DISPATCH)

    def test_secondary_contact_loads_only_eligible_sub_contacts_for_own_attempt(self):
        other_tertiary = m.Contact.objects.create(
            name="产品三级接口人",
            employee_no="T-B",
            department=m.Department.objects.create(
                name="产品三级组", level=3, parent=self.dept_b
            ),
            contact_level=m.Contact.LEVEL_TERTIARY,
        )
        self.client.force_authenticate(self.secondary_user)

        response = self.client.get(
            f"/api/workflow-attempts/{self.attempt_a.id}/eligible-sub-contacts/"
        )
        forbidden_response = self.client.get(
            f"/api/workflow-attempts/{self.attempt_b.id}/eligible-sub-contacts/"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual([item["id"] for item in response.data], [self.tertiary_a.id])
        self.assertNotIn(other_tertiary.id, [item["id"] for item in response.data])
        self.assertEqual(forbidden_response.status_code, 404)

    def test_rule_and_ai_attempt_lists_return_assignment_reason(self):
        self.attempt_a.match_reason = "规则：院校准入、专业匹配"
        self.attempt_a.save(update_fields=["match_reason"])
        self.attempt_b.source = m.AssignmentAttempt.SOURCE_AI
        self.attempt_b.match_reason = "AI：简历能力与岗位要求匹配"
        self.attempt_b.save(update_fields=["source", "match_reason"])
        self.client.force_authenticate(self.hr)

        rule_response = self.client.get("/api/workflow-attempts/", {"source": "rule"})
        ai_response = self.client.get("/api/workflow-attempts/", {"source": "ai"})

        self.assertEqual(rule_response.status_code, 200)
        self.assertEqual(ai_response.status_code, 200)
        self.assertEqual(rule_response.data["results"][0]["match_reason"], "规则：院校准入、专业匹配")
        self.assertEqual(ai_response.data["results"][0]["match_reason"], "AI：简历能力与岗位要求匹配")

    def test_workflow_and_attempt_header_filters_match_visible_columns(self):
        self.attempt_a.match_reason = "院校准入且专业匹配"
        self.attempt_a.save(update_fields=["match_reason"])
        workflow = self.attempt_a.workflow
        workflow.dispatch_strategy = "rule"
        workflow.save(update_fields=["dispatch_strategy"])
        self.client.force_authenticate(self.hr)

        workflow_response = self.client.get(
            "/api/workflows/",
            {
                "candidate_name": "张",
                "current_rank": "1",
                "current_apply_id": "A1001",
                "current_position_name": "后端",
                "dispatch_strategy": "rule",
            },
        )
        attempt_response = self.client.get(
            "/api/workflow-attempts/",
            {
                "candidate_name": "张",
                "volunteer_rank": "1",
                "apply_id": "A1001",
                "position_name": "后端",
                "department_name": "技术",
                "contact_name": "技术二级",
                "sub_contact_name": "技术三级",
                "match_reason": "专业匹配",
            },
        )

        self.assertEqual(workflow_response.status_code, 200)
        self.assertEqual([item["id"] for item in workflow_response.data["results"]], [workflow.id])
        self.assertEqual(attempt_response.status_code, 200)
        self.assertEqual([item["id"] for item in attempt_response.data["results"]], [self.attempt_a.id])

    def test_user_and_role_header_filters(self):
        self.client.force_authenticate(self.admin)

        user_response = self.client.get(
            "/api/users/",
            {
                "username": "hr",
                "email": "hr@",
                "roles": "HR",
                "is_active": "true",
            },
        )
        role_response = self.client.get("/api/roles/", {"name": "管理"})

        self.assertEqual(user_response.status_code, 200)
        self.assertEqual([item["username"] for item in user_response.data["results"]], ["hr"])
        self.assertEqual(role_response.status_code, 200)
        self.assertEqual([item["name"] for item in role_response.data["results"]], ["管理员"])

    def test_contact_cannot_access_settings(self):
        self.client.force_authenticate(self.secondary_user)

        response = self.client.get("/api/configs/")

        self.assertEqual(response.status_code, 403)

    def test_department_manager_can_update_welink_notification_setting(self):
        manager = User.objects.create_user(username="department-manager", password="pass")
        manager.user_permissions.add(
            Permission.objects.get(codename=permission_codename("department.manage"))
        )
        self.client.force_authenticate(manager)

        response = self.client.patch(
            "/api/configs/welink_enabled/", {"value": True}, format="json"
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["value"])
        self.assertTrue(m.Config.objects.get(key="welink_enabled").value)

        forbidden = self.client.patch(
            "/api/configs/job_hc_coefficient/", {"value": 2}, format="json"
        )
        self.assertEqual(forbidden.status_code, 403)

    def test_config_manager_updates_bounded_job_hc_coefficient(self):
        self.client.force_authenticate(self.admin)

        response = self.client.patch(
            "/api/configs/job_hc_coefficient/", {"value": 3}, format="json"
        )
        invalid_low = self.client.patch(
            "/api/configs/job_hc_coefficient/", {"value": 0}, format="json"
        )
        invalid_high = self.client.patch(
            "/api/configs/job_hc_coefficient/", {"value": 101}, format="json"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["value"], 3)
        self.assertEqual(invalid_low.status_code, 400)
        self.assertEqual(invalid_high.status_code, 400)

    def test_resume_import_permission_can_read_allocation_mode_without_pipeline_run(self):
        importer = User.objects.create_user(username="import-only", password="pass")
        importer.user_permissions.add(
            Permission.objects.get(codename=permission_codename("resume.import"))
        )
        self.client.force_authenticate(importer)

        response = self.client.get("/api/allocation-mode/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["default_mode"], "rule")
        self.assertEqual(response.data["available_modes"], ["rule"])

    def test_admin_can_update_known_ai_setting(self):
        self.client.force_authenticate(self.admin)

        response = self.client.patch(
            "/api/ai-connection/settings/ai_dispatch_threshold/",
            {"value": 0.82},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["key"], "ai_dispatch_threshold")
        self.assertEqual(response.data["value"], 0.82)
        self.assertEqual(m.Config.objects.get(key="ai_dispatch_threshold").value, 0.82)

    def test_admin_can_enable_valid_ai_special_route_config(self):
        self.client.force_authenticate(self.admin)
        for key, value in [
            ("ai_special_route_secondary_contact_id", self.secondary_a.id),
            ("ai_special_route_tertiary_contact_id", self.tertiary_a.id),
            ("ai_special_route_threshold", 0.9),
            ("ai_special_route_enabled", True),
        ]:
            response = self.client.patch(
                f"/api/ai-connection/settings/{key}/", {"value": value}, format="json"
            )
            self.assertEqual(response.status_code, 200, (key, response.data))

        config = ai_config.get_ai_special_route_config(validate=True)
        self.assertTrue(config.enabled)
        self.assertEqual(config.secondary_contact_id, self.secondary_a.id)
        self.assertEqual(config.tertiary_contact_id, self.tertiary_a.id)

    def test_special_route_rejects_contacts_outside_same_department_tree(self):
        wrong_tertiary = m.Contact.objects.create(
            name="产品三级接口人",
            employee_no="T-B",
            department=m.Department.objects.create(
                name="产品三级组", level=3, parent=self.dept_b
            ),
            contact_level=m.Contact.LEVEL_TERTIARY,
        )
        self.client.force_authenticate(self.admin)
        for key, value in [
            ("ai_special_route_secondary_contact_id", self.secondary_a.id),
            ("ai_special_route_tertiary_contact_id", wrong_tertiary.id),
        ]:
            response = self.client.patch(
                f"/api/ai-connection/settings/{key}/", {"value": value}, format="json"
            )
            self.assertEqual(response.status_code, 200)

        response = self.client.patch(
            "/api/ai-connection/settings/ai_special_route_enabled/",
            {"value": True},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("上下级部门", response.data["detail"])

    def test_ai_mode_becomes_available_after_successful_connection_test(self):
        self.client.force_authenticate(self.admin)

        unavailable_response = self.client.get("/api/allocation-mode/")
        self.assertEqual(
            unavailable_response.data,
            {"default_mode": "rule", "available_modes": ["rule"], "ai_ready": False},
        )

        ai_config.save_ai_connection_config(
            {
                "api_style": "responses",
                "model_name": "gpt-test",
                "base_url": "https://model.internal/v1",
                "api_key": "test-key",
            }
        )
        with patch(
            "apps.api.views.ai_service.test_model_connection",
            return_value={
                "model_name": "gpt-test",
                "api_style": "responses",
                "base_url": "https://model.internal/v1",
            },
        ):
            test_response = self.client.post("/api/ai-connection/test/")
        mode_response = self.client.get("/api/allocation-mode/")

        self.assertEqual(test_response.status_code, 200)
        self.assertTrue(test_response.data["ok"])
        self.assertEqual(
            mode_response.data,
            {
                "default_mode": "rule",
                "available_modes": ["rule", "ai"],
                "ai_ready": True,
            },
        )
        self.assertEqual(
            self.client.patch(
                "/api/configs/ai_enabled/", {"value": True}, format="json"
            ).status_code,
            404,
        )

    def test_saving_ai_connection_invalidates_test_and_removes_legacy_switch(self):
        self.client.force_authenticate(self.admin)
        ai_config.save_ai_connection_config(
            {
                "api_style": "responses",
                "model_name": "gpt-test",
                "base_url": "https://model.internal/v1",
                "api_key": "test-key",
            }
        )
        ai_config.mark_ai_connection_tested()
        m.Config.objects.update_or_create(key="ai_enabled", defaults={"value": True})

        response = self.client.patch(
            "/api/ai-connection/",
            {"model_name": "gpt-test-v2"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data["test_passed"])
        self.assertFalse(m.Config.objects.filter(key="ai_enabled").exists())
        self.assertEqual(ai_config.available_allocation_modes(), ["rule"])

    def test_admin_config_api_excludes_ai_connection_settings(self):
        self.client.force_authenticate(self.admin)

        response = self.client.get("/api/configs/")

        self.assertEqual(response.status_code, 200)
        keys = {item["key"] for item in response.data}
        self.assertNotIn("ai_dispatch_threshold", keys)
        self.assertNotIn("ai_enabled", keys)
        self.assertNotIn("AI_MODEL_NAME", keys)
        self.assertNotIn("AI_API_KEY_ENV", keys)
        self.assertNotIn("AI_BASE_URL_ENV", keys)
        self.assertFalse(
            {"api_key", "api_key_env", "base_url", "base_url_env"}
            & {field for item in response.data for field in item}
        )

        update_response = self.client.patch(
            "/api/configs/AI_MODEL_NAME/",
            {"value": "gpt-test"},
            format="json",
        )
        self.assertEqual(update_response.status_code, 404)

        ai_settings_response = self.client.get("/api/ai-connection/settings/")
        self.assertEqual(ai_settings_response.status_code, 200)
        ai_keys = {item["key"] for item in ai_settings_response.data["settings"]}
        self.assertIn("ai_dispatch_threshold", ai_keys)
        self.assertIn("ai_special_route_enabled", ai_keys)
        self.assertNotIn("ai_enabled", ai_keys)

    def test_only_users_with_ai_connection_permission_can_save_and_view_redacted_connection(self):
        self.client.force_authenticate(self.hr)
        self.assertEqual(self.client.get("/api/ai-connection/").status_code, 403)
        self.assertEqual(self.client.get("/api/ai-connection/settings/").status_code, 403)

        self.client.force_authenticate(self.admin)
        payload = {
            "api_style": "chat_json",
            "model_name": "deepseek-v4-pro",
            "base_url": "https://api.deepseek.com",
            "api_key": "sk-test-secret",
        }
        response = self.client.patch("/api/ai-connection/", payload, format="json")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["api_key_configured"])
        self.assertNotIn("api_key", response.data)
        self.assertNotIn("profile", response.data)
        self.assertNotIn("profiles", response.data)
        self.assertNotIn("sk-test-secret", str(response.data))
        stored = m.Config.objects.get(key="ai_connection_api_key").value
        self.assertNotEqual(stored, "sk-test-secret")
        model_config = ai_config.get_ai_model_config()
        self.assertEqual(model_config.model_name, "deepseek-v4-pro")
        self.assertEqual(model_config.api_key, "sk-test-secret")

    def test_admin_can_discover_models_and_hr_cannot(self):
        self.client.force_authenticate(self.hr)
        denied = self.client.post(
            "/api/ai-connection/models/",
            {"base_url": "https://model.internal/v1"},
            format="json",
        )
        self.assertEqual(denied.status_code, 403)

        self.client.force_authenticate(self.admin)
        with patch(
            "apps.api.views.ai_service.list_available_models",
            return_value=["deepseek-v4", "glm-4.7"],
        ) as discover:
            response = self.client.post(
                "/api/ai-connection/models/",
                {
                    "base_url": "https://model.internal/v1",
                    "api_key": "temporary-token",
                },
                format="json",
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["models"], ["deepseek-v4", "glm-4.7"])
        discover.assert_called_once_with(
            base_url="https://model.internal/v1", api_key="temporary-token"
        )
        self.assertNotIn("temporary-token", str(response.data))

    def test_ai_connection_allows_empty_access_token(self):
        self.client.force_authenticate(self.admin)

        response = self.client.patch(
            "/api/ai-connection/",
            {
                "api_style": "chat_json",
                "model_name": "glm-4.7",
                "base_url": "https://model.internal/v1",
                "clear_api_key": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data["api_key_configured"])
        self.assertEqual(ai_config.get_ai_model_config().api_key, "")

    def test_changing_base_url_without_new_token_clears_saved_token(self):
        self.client.force_authenticate(self.admin)
        first = self.client.patch(
            "/api/ai-connection/",
            {
                "api_style": "chat_json",
                "model_name": "deepseek-v4",
                "base_url": "https://model.internal/v1",
                "api_key": "bound-secret",
            },
            format="json",
        )

        changed = self.client.patch(
            "/api/ai-connection/",
            {
                "base_url": "https://another-model.internal/v1",
                "api_key": "",
            },
            format="json",
        )

        self.assertEqual(first.status_code, 200)
        self.assertTrue(first.data["api_key_configured"])
        self.assertEqual(changed.status_code, 200)
        self.assertFalse(changed.data["api_key_configured"])
        self.assertFalse(
            m.Config.objects.filter(key="ai_connection_api_key").exists()
        )

    def test_role_granted_ai_connection_permission_can_manage_connection(self):
        permission = Permission.objects.get(
            codename=permission_codename("settings.manage_ai_connection")
        )
        ai_operator = Group.objects.create(name="AI 模型管理员")
        ai_operator.permissions.add(permission)
        user = User.objects.create_user(
            username="ai-operator", password="pass", role=User.ROLE_HR
        )
        user.groups.add(ai_operator)
        self.client.force_authenticate(user)

        response = self.client.get("/api/ai-connection/")

        self.assertEqual(response.status_code, 200)
        update_response = self.client.patch(
            "/api/ai-connection/",
            {
                "api_style": "chat_json",
                "model_name": "delegated-model",
                "base_url": "https://model.internal/v1",
                "api_key": "delegated-test-key",
            },
            format="json",
        )
        self.assertEqual(update_response.status_code, 200)
        with patch(
            "apps.api.views.ai_service.test_model_connection",
            return_value={
                "model_name": "delegated-model",
                "api_style": "chat_json",
                "base_url": "https://api.deepseek.com",
            },
        ):
            test_response = self.client.post("/api/ai-connection/test/")

        self.assertEqual(test_response.status_code, 200)
        self.assertTrue(test_response.data["ok"])

    def test_admin_can_test_ai_connection(self):
        self.client.force_authenticate(self.admin)
        config_response = self.client.patch(
            "/api/ai-connection/",
            {
                "api_style": "responses",
                "model_name": "gpt-test",
                "base_url": "https://model.internal/v1",
                "api_key": "test-key",
            },
            format="json",
        )
        self.assertEqual(config_response.status_code, 200)
        with patch(
            "apps.api.views.ai_service.test_model_connection",
            return_value={
                "model_name": "deepseek-v4-pro",
                "api_style": "chat_json",
                "base_url": "https://api.deepseek.com",
            },
        ):
            response = self.client.post("/api/ai-connection/test/")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["ok"])
        self.assertNotIn("api_key", response.data)

    def test_environment_key_never_enables_ai_connection(self):
        self.client.force_authenticate(self.admin)
        with patch.dict("os.environ", {"DEEPSEEK_API_KEY": "env-secret"}, clear=False):
            response = self.client.get("/api/ai-connection/")

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data["api_key_configured"])
        self.assertFalse(ai_config.is_ai_available())

    def test_permissions_endpoint_returns_tree_for_admin(self):
        self.client.force_authenticate(self.admin)

        response = self.client.get("/api/permissions/")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data)
        codes = [
            permission["code"]
            for module in response.data
            for permission in module["children"]
        ]
        self.assertIn("settings.manage_permissions", codes)
        self.assertIn("settings.manage_ai_connection", codes)


class SchoolRuleConfigApiTests(TestCase):
    def setUp(self):
        ensure_rbac_defaults()
        self.client = APIClient()
        self.admin = User.objects.create_user(
            username="admin-school-rule", password="pass", role=User.ROLE_ADMIN
        )
        self.admin.groups.add(Group.objects.get(name="管理员"))
        self.client.force_authenticate(self.admin)

    def test_school_tag_crud_endpoint_creates_tag_dictionary_item(self):
        response = self.client.post(
            "/api/school-tags/",
            {
                "code": "A",
                "name": "平台A",
                "is_default": False,
                "is_active": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["code"], "A")
        self.assertEqual(response.data["name"], "平台A")
        self.assertTrue(m.SchoolTag.objects.filter(code="A", name="平台A").exists())

    def test_school_tag_rule_accepts_first_and_highest_tag_ids(self):
        first_tag = m.SchoolTag.objects.create(code="A", name="平台A")
        highest_tag = m.SchoolTag.objects.create(code="B", name="平台B")

        response = self.client.post(
            "/api/school-tag-rules/",
            {
                "name": "重点院校组合",
                "priority": 10,
                "is_active": True,
                "first_degree_tag_ids": [first_tag.id],
                "highest_degree_tag_ids": [highest_tag.id],
                "allowed_highest_educations": ["bachelor", "master"],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["first_degree_tags"][0]["id"], first_tag.id)
        self.assertEqual(response.data["highest_degree_tags"][0]["id"], highest_tag.id)
        self.assertEqual(
            response.data["allowed_highest_educations"], ["bachelor", "master"]
        )
        self.assertTrue(
            m.SchoolTagRuleTag.objects.filter(
                rule_id=response.data["id"],
                school_tag=first_tag,
                degree_type=m.SchoolTagRuleTag.DEGREE_FIRST,
            ).exists()
        )
        self.assertEqual(
            set(
                m.SchoolTagRuleEducation.objects.filter(
                    rule_id=response.data["id"]
                ).values_list("education", flat=True)
            ),
            {"bachelor", "master"},
        )

    def test_school_tag_rule_omitted_education_list_means_unrestricted(self):
        first_tag = m.SchoolTag.objects.create(code="FIRST", name="第一学历")
        highest_tag = m.SchoolTag.objects.create(code="HIGH", name="最高学历")

        response = self.client.post(
            "/api/school-tag-rules/",
            {
                "name": "兼容旧规则",
                "is_active": True,
                "first_degree_tag_ids": [first_tag.id],
                "highest_degree_tag_ids": [highest_tag.id],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["allowed_highest_educations"], [])

    def test_active_school_tag_rule_requires_first_and_highest_tags(self):
        first_tag = m.SchoolTag.objects.create(code="A", name="平台A")

        response = self.client.post(
            "/api/school-tag-rules/",
            {
                "name": "缺少最高学历标签",
                "priority": 10,
                "is_active": True,
                "first_degree_tag_ids": [first_tag.id],
                "highest_degree_tag_ids": [],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("highest_degree_tag_ids", response.data)

    def test_school_tag_rule_model_no_longer_keeps_legacy_json_fields(self):
        field_names = {field.name for field in m.SchoolTagRule._meta.get_fields()}

        self.assertNotIn("first_degree_tags", field_names)
        self.assertNotIn("highest_degree_tags", field_names)

    def test_school_tag_and_rule_header_filters(self):
        default_tag = m.SchoolTag.objects.create(
            code="NON_TARGET", name="非目标院校", is_default=True, is_active=True
        )
        other_tag = m.SchoolTag.objects.create(
            code="TARGET", name="目标院校", is_default=False, is_active=True
        )
        rule = m.SchoolTagRule.objects.create(name="默认准入规则", priority=8, is_active=True)
        m.SchoolTagRuleTag.objects.create(
            rule=rule,
            school_tag=default_tag,
            degree_type=m.SchoolTagRuleTag.DEGREE_FIRST,
        )
        m.SchoolTagRuleTag.objects.create(
            rule=rule,
            school_tag=other_tag,
            degree_type=m.SchoolTagRuleTag.DEGREE_HIGHEST,
        )

        tag_response = self.client.get(
            "/api/school-tags/", {"code": "NON", "is_default": "true"}
        )
        rule_response = self.client.get(
            "/api/school-tag-rules/", {"name": "默认", "priority": "8", "is_active": "true"}
        )

        self.assertEqual(tag_response.status_code, 200)
        self.assertEqual([item["id"] for item in tag_response.data["results"]], [default_tag.id])
        self.assertEqual(rule_response.status_code, 200)
        self.assertEqual([item["id"] for item in rule_response.data["results"]], [rule.id])


class ListFilteringPaginationApiTests(TestCase):
    def setUp(self):
        ensure_rbac_defaults()
        self.client = APIClient()
        self.admin = User.objects.create_user(
            username="admin-list", password="pass", role=User.ROLE_ADMIN
        )
        self.admin.groups.add(Group.objects.get(name="管理员"))
        self.client.force_authenticate(self.admin)

    def test_page_size_query_param_controls_candidate_page_length(self):
        for index in range(5):
            m.Candidate.objects.create(
                identity_hash=f"candidate-page-{index}",
                name=f"候选人{index}",
                phone=f"1380000000{index}",
            )

        response = self.client.get("/api/candidates/", {"page_size": 2})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 5)
        self.assertEqual(len(response.data["results"]), 2)

    def test_candidate_list_filters_current_apply_date_with_inclusive_and_one_sided_ranges(self):
        candidates = {}
        for suffix, apply_date in [
            ("start", "2026-07-01"),
            ("middle", "2026-07-15"),
            ("end", "2026-07-31"),
            ("empty", None),
        ]:
            candidate = m.Candidate.objects.create(
                identity_hash=f"candidate-apply-date-{suffix}",
                name=f"投递日期候选人-{suffix}",
                phone=f"1386000000{len(candidates)}",
            )
            resume = m.Resume.objects.create(
                candidate=candidate,
                apply_id=f"APPLY-DATE-{suffix.upper()}",
                position_name="后端工程师",
                volunteer_rank=1,
                apply_date=apply_date,
            )
            m.CandidateWorkflow.objects.create(
                candidate=candidate,
                status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
                current_resume=resume,
                current_rank=1,
            )
            candidates[suffix] = candidate

        inclusive = self.client.get(
            "/api/candidates/",
            {
                "current_apply_date_from": "2026-07-01",
                "current_apply_date_to": "2026-07-31",
            },
        )
        from_only = self.client.get(
            "/api/candidates/", {"current_apply_date_from": "2026-07-15"}
        )
        to_only = self.client.get(
            "/api/candidates/", {"current_apply_date_to": "2026-07-15"}
        )
        exact = self.client.get(
            "/api/candidates/",
            {
                "current_apply_date_from": "2026-07-15",
                "current_apply_date_to": "2026-07-15",
            },
        )
        unfiltered = self.client.get("/api/candidates/")

        self.assertEqual(inclusive.status_code, 200)
        self.assertEqual(
            {item["id"] for item in inclusive.data["results"]},
            {candidates["start"].id, candidates["middle"].id, candidates["end"].id},
        )
        self.assertEqual(
            {item["id"] for item in from_only.data["results"]},
            {candidates["middle"].id, candidates["end"].id},
        )
        self.assertEqual(
            {item["id"] for item in to_only.data["results"]},
            {candidates["start"].id, candidates["middle"].id},
        )
        self.assertEqual(exact.data["results"][0]["current_apply_date"], "2026-07-15")
        empty_row = next(
            item
            for item in unfiltered.data["results"]
            if item["id"] == candidates["empty"].id
        )
        self.assertIsNone(empty_row["current_apply_date"])

    def test_candidate_list_rejects_invalid_current_apply_date_ranges(self):
        invalid_params = [
            {"current_apply_date_from": "2026-7-01"},
            {"current_apply_date_to": "2026-02-30"},
            {
                "current_apply_date_from": "2026-07-31",
                "current_apply_date_to": "2026-07-01",
            },
        ]

        for params in invalid_params:
            with self.subTest(params=params):
                response = self.client.get("/api/candidates/", params)
                self.assertEqual(response.status_code, 400)

    def test_candidate_apply_date_filter_uses_current_resume_and_no_workflow_fallback(self):
        workflow_candidate = m.Candidate.objects.create(
            identity_hash="candidate-apply-date-workflow",
            name="工作流投递日期候选人",
            phone="13860000100",
        )
        m.Resume.objects.create(
            candidate=workflow_candidate,
            apply_id="APPLY-DATE-HISTORY",
            volunteer_rank=1,
            apply_date="2026-06-01",
        )
        current_resume = m.Resume.objects.create(
            candidate=workflow_candidate,
            apply_id="APPLY-DATE-CURRENT",
            volunteer_rank=2,
            apply_date="2026-07-20",
        )
        m.CandidateWorkflow.objects.create(
            candidate=workflow_candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=current_resume,
            current_rank=2,
        )
        fallback_candidate = m.Candidate.objects.create(
            identity_hash="candidate-apply-date-fallback",
            name="无工作流投递日期候选人",
            phone="13860000101",
        )
        m.Resume.objects.create(
            candidate=fallback_candidate,
            apply_id="APPLY-DATE-FALLBACK-SECOND",
            volunteer_rank=2,
            apply_date="2026-07-25",
        )
        m.Resume.objects.create(
            candidate=fallback_candidate,
            apply_id="APPLY-DATE-FALLBACK-FIRST",
            volunteer_rank=1,
            apply_date="2026-06-01",
        )

        history_date = self.client.get(
            "/api/candidates/",
            {
                "current_apply_date_from": "2026-06-01",
                "current_apply_date_to": "2026-06-01",
            },
        )
        current_date = self.client.get(
            "/api/candidates/",
            {
                "current_apply_date_from": "2026-07-20",
                "current_apply_date_to": "2026-07-20",
            },
        )

        self.assertEqual(
            [item["id"] for item in history_date.data["results"]],
            [fallback_candidate.id],
        )
        self.assertEqual(
            [item["id"] for item in current_date.data["results"]],
            [workflow_candidate.id],
        )
        self.assertEqual(
            history_date.data["results"][0]["current_apply_id"],
            "APPLY-DATE-FALLBACK-FIRST",
        )

    def test_contact_candidate_apply_date_uses_own_visible_attempt_resume(self):
        own_department = m.Department.objects.create(name="日期范围本人二部", level=2)
        other_department = m.Department.objects.create(name="日期范围其他二部", level=2)
        own_contact = m.Contact.objects.create(
            name="日期范围本人接口人",
            employee_no="APPLY-DATE-OWN",
            department=own_department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        other_contact = m.Contact.objects.create(
            name="日期范围其他接口人",
            employee_no="APPLY-DATE-OTHER",
            department=other_department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-contact-apply-date",
            name="接口人投递日期候选人",
            phone="13860000200",
        )
        visible_resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="APPLY-DATE-VISIBLE",
            volunteer_rank=1,
            apply_date="2026-07-05",
        )
        hidden_resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="APPLY-DATE-HIDDEN",
            volunteer_rank=2,
            apply_date="2026-07-20",
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=hidden_resume,
            current_rank=2,
        )
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=visible_resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
            department=own_department,
            contact=own_contact,
        )
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=hidden_resume,
            attempt_no=2,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
            department=other_department,
            contact=other_contact,
        )
        contact_user = User.objects.create_user(
            username="APPLY-DATE-OWN",
            password="pass",
            role=User.ROLE_SECONDARY_CONTACT,
            contact=own_contact,
        )
        contact_user.groups.add(Group.objects.get(name="二级接口人"))
        self.client.force_authenticate(contact_user)

        visible_date = self.client.get(
            "/api/candidates/",
            {
                "current_apply_date_from": "2026-07-05",
                "current_apply_date_to": "2026-07-05",
            },
        )
        hidden_date = self.client.get(
            "/api/candidates/",
            {
                "current_apply_date_from": "2026-07-20",
                "current_apply_date_to": "2026-07-20",
            },
        )

        self.assertEqual(visible_date.status_code, 200)
        self.assertEqual(visible_date.data["count"], 1)
        self.assertEqual(
            visible_date.data["results"][0]["current_apply_date"], "2026-07-05"
        )
        self.assertEqual(
            visible_date.data["results"][0]["current_apply_id"],
            "APPLY-DATE-VISIBLE",
        )
        self.assertEqual(hidden_date.data["count"], 0)

    def test_candidate_allocation_source_covers_auto_archives_but_not_unstarted_workflows(self):
        rule_candidate = m.Candidate.objects.create(
            identity_hash="candidate-source-rule",
            name="规则归档候选人",
            phone="13810000001",
        )
        ai_candidate = m.Candidate.objects.create(
            identity_hash="candidate-source-ai",
            name="AI归档候选人",
            phone="13810000002",
        )
        untouched_candidate = m.Candidate.objects.create(
            identity_hash="candidate-source-untouched",
            name="未处理候选人",
            phone="13810000003",
        )
        m.CandidateWorkflow.objects.create(
            candidate=rule_candidate,
            status=m.CandidateWorkflow.STATUS_ARCHIVED,
            dispatch_strategy=m.AssignmentAttempt.SOURCE_RULE,
            archive_reason=m.CandidateWorkflow.ARCHIVE_JOB_NOT_MATCHED,
            started_at=timezone.now(),
        )
        m.CandidateWorkflow.objects.create(
            candidate=ai_candidate,
            status=m.CandidateWorkflow.STATUS_ARCHIVED,
            dispatch_strategy=m.AssignmentAttempt.SOURCE_AI,
            archive_reason=m.CandidateWorkflow.ARCHIVE_AGENT_NO_RECOMMENDATION,
            started_at=timezone.now(),
        )
        m.CandidateWorkflow.objects.create(
            candidate=untouched_candidate,
            status=m.CandidateWorkflow.STATUS_PENDING,
            dispatch_strategy=m.AssignmentAttempt.SOURCE_AI,
        )

        response = self.client.get("/api/candidates/")

        self.assertEqual(response.status_code, 200)
        sources = {item["id"]: item["allocation_source"] for item in response.data["results"]}
        self.assertEqual(sources[rule_candidate.id], m.AssignmentAttempt.SOURCE_RULE)
        self.assertEqual(sources[ai_candidate.id], m.AssignmentAttempt.SOURCE_AI)
        self.assertEqual(sources[untouched_candidate.id], "")

        rule_response = self.client.get(
            "/api/candidates/", {"allocation_source": m.AssignmentAttempt.SOURCE_RULE}
        )
        ai_response = self.client.get(
            "/api/candidates/", {"allocation_source": m.AssignmentAttempt.SOURCE_AI}
        )

        self.assertEqual([item["id"] for item in rule_response.data["results"]], [rule_candidate.id])
        self.assertEqual([item["id"] for item in ai_response.data["results"]], [ai_candidate.id])

    def test_candidate_list_filters_by_processing_run_result(self):
        run = m.ProcessingRun.objects.create(step="step2", mode="ai")

        def create_result_candidate(
            code,
            scope_status,
            recommendation=None,
            *,
            result_type="completed",
            reason_code="",
        ):
            candidate = m.Candidate.objects.create(
                identity_hash=f"candidate-run-{code}",
                name=f"任务{code}",
                phone=f"13820000{len(code):03d}",
            )
            resume = m.Resume.objects.create(
                candidate=candidate,
                apply_id=f"RUN-{code}",
                position_name="算法工程师",
                volunteer_rank=1,
            )
            workflow = m.CandidateWorkflow.objects.create(
                candidate=candidate,
                status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
                current_resume=resume,
                current_rank=1,
                dispatch_strategy=m.AssignmentAttempt.SOURCE_AI,
                started_at=timezone.now(),
            )
            m.ProcessingRunScopeItem.objects.create(
                run=run,
                candidate=candidate,
                status=scope_status,
                result_type=result_type,
                reason_code=reason_code,
            )
            if recommendation:
                m.AgentDispatchDecision.objects.create(
                    workflow=workflow,
                    resume=resume,
                    processing_run=run,
                    recommendation=recommendation,
                )
            return candidate

        success = create_result_candidate("success", "success")
        attention = create_result_candidate(
            "attention",
            "needs_attention",
            result_type="needs_attention",
            reason_code="ai_rate_limited",
        )
        failed = create_result_candidate(
            "failed",
            "failed",
            result_type="failed",
            reason_code="llm_timeout",
        )
        review = create_result_candidate("review", "success", "review")
        dispatch = create_result_candidate("dispatch", "success", "dispatch")
        archive = create_result_candidate("archive", "success", "archive")
        skipped = create_result_candidate(
            "skipped", "skipped_manual_change", result_type=""
        )
        cancelled = create_result_candidate(
            "cancelled", "cancelled", result_type="cancelled"
        )

        expected_ids = {
            "success": {success.id, review.id, dispatch.id, archive.id},
            "completed": {success.id, review.id, dispatch.id, archive.id},
            "needs_attention": {attention.id},
            "failed": {failed.id},
            "review": {review.id},
            "dispatch": {dispatch.id},
            "archive": {archive.id},
            "skipped": {skipped.id},
            "cancelled": {cancelled.id},
        }
        for result, ids in expected_ids.items():
            response = self.client.get(
                "/api/candidates/",
                {"processing_run_id": run.id, "processing_result": result},
            )
            self.assertEqual(response.status_code, 200)
            self.assertEqual(
                {item["id"] for item in response.data["results"]}, ids, result
            )

        response = self.client.get(
            "/api/candidates/",
            {
                "processing_run_id": run.id,
                "result_type": "needs_attention",
                "reason_code": "ai_rate_limited",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            {item["id"] for item in response.data["results"]},
            {attention.id},
        )

        other_run = m.ProcessingRun.objects.create(step="step2", mode="rule")
        exact_candidate = m.Candidate.objects.create(
            identity_hash="candidate-exact-job-not-found",
            name="精确缺岗候选人",
            phone="13820000991",
        )
        mixed_candidate = m.Candidate.objects.create(
            identity_hash="candidate-mixed-processing-results",
            name="混合历史候选人",
            phone="13820000992",
        )
        m.ProcessingRunScopeItem.objects.create(
            run=other_run,
            candidate=exact_candidate,
            status="success",
            result_type="completed",
            reason_code="job_not_found",
        )
        m.ProcessingRunScopeItem.objects.create(
            run=run,
            candidate=mixed_candidate,
            status="success",
            result_type="completed",
            reason_code="education_not_eligible",
        )
        m.ProcessingRunScopeItem.objects.create(
            run=other_run,
            candidate=mixed_candidate,
            status="needs_attention",
            result_type="needs_attention",
            reason_code="job_not_found",
        )

        exact_response = self.client.get(
            "/api/candidates/",
            {"result_type": "completed", "reason_code": "job_not_found"},
        )

        self.assertEqual(exact_response.status_code, 200)
        self.assertEqual(
            {item["id"] for item in exact_response.data["results"]},
            {exact_candidate.id, mixed_candidate.id},
        )

    def test_candidate_reason_filter_matches_latest_displayed_processing_item(self):
        historical_run = m.ProcessingRun.objects.create(step="step2", mode="rule")
        current_run = m.ProcessingRun.objects.create(step="step2", mode="ai")

        def create_candidate(code, old_reason, current_reason):
            candidate = m.Candidate.objects.create(
                identity_hash=f"candidate-current-reason-{code}",
                name=f"原因筛选{code}",
                phone=f"13829990{len(code):03d}",
            )
            m.ProcessingRunScopeItem.objects.create(
                run=historical_run,
                candidate=candidate,
                status="success",
                result_type="completed",
                reason_code=old_reason,
            )
            m.ProcessingRunScopeItem.objects.create(
                run=current_run,
                candidate=candidate,
                status="success",
                result_type="completed",
                reason_code=current_reason,
            )
            return candidate

        current_job_missing = create_candidate(
            "job", "education_not_eligible", "job_not_found"
        )
        current_education_failed = create_candidate(
            "education", "job_not_found", "education_not_eligible"
        )
        current_ai_review = create_candidate(
            "review", "job_not_found", "ai_review"
        )

        response = self.client.get(
            "/api/candidates/", {"reason_code": "job_not_found"}
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [item["id"] for item in response.data["results"]],
            [current_job_missing.id],
        )
        self.assertEqual(response.data["results"][0]["reason_code"], "job_not_found")

        historical_response = self.client.get(
            "/api/candidates/",
            {
                "processing_run_id": historical_run.id,
                "reason_code": "job_not_found",
            },
        )
        self.assertEqual(historical_response.status_code, 200)
        self.assertEqual(
            {item["id"] for item in historical_response.data["results"]},
            {current_education_failed.id, current_ai_review.id},
        )
        self.assertTrue(
            all(
                item["reason_code"] == "job_not_found"
                for item in historical_response.data["results"]
            )
        )

    def test_candidate_header_filters_match_current_resume_and_school_tag(self):
        keep = m.Candidate.objects.create(
            identity_hash="candidate-keep",
            name="张三",
            phone="13800000001",
            first_degree_school="南京大学",
            highest_degree_school="南京大学",
            highest_major="计算机",
            highest_degree_platform="平台A",
        )
        drop = m.Candidate.objects.create(
            identity_hash="candidate-drop",
            name="李四",
            phone="13900000002",
            first_degree_school="普通大学",
            highest_degree_school="普通大学",
            highest_major="市场营销",
            highest_degree_platform="平台B",
        )
        keep_resume = m.Resume.objects.create(
            candidate=keep,
            apply_id="A1001",
            entity="GW",
            position_name="后端工程师",
            volunteer_rank=1,
            job_category="技术类",
        )
        drop_resume = m.Resume.objects.create(
            candidate=drop,
            apply_id="B1001",
            entity="YLS",
            position_name="产品经理",
            volunteer_rank=2,
            job_category="产品类",
        )
        m.CandidateWorkflow.objects.create(
            candidate=keep,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=keep_resume,
            current_rank=1,
        )
        m.CandidateWorkflow.objects.create(
            candidate=drop,
            status=m.CandidateWorkflow.STATUS_PENDING,
            current_resume=drop_resume,
            current_rank=2,
        )

        response = self.client.get(
            "/api/candidates/",
            {
                "name": "张",
                "phone": "138",
                "highest_major": "计算机",
                "current_rank": 1,
                "current_entity": "GW",
                "current_position_name": "后端",
                "current_job_category": "技术",
                "school_tag": "平台A",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual([item["id"] for item in response.data["results"]], [keep.id])
        self.assertEqual(response.data["results"][0]["highest_major"], "计算机")

    def test_candidate_selector_filters_and_options_use_current_display_values(self):
        school_tag = m.SchoolTag.objects.create(code="TARGET", name="目标院校")
        department = m.Department.objects.create(name="研发中心", level=2)
        job = m.Job.objects.create(
            department=department,
            public_name="后端工程师",
            position_name="后端工程师",
            category="技术类",
            headcount=1,
        )
        keep = m.Candidate.objects.create(
            identity_hash="candidate-selector-keep",
            name="选择器候选人",
            phone="13830000901",
            highest_major="计算机科学与技术",
            highest_degree_tag=school_tag,
        )
        keep_resume = m.Resume.objects.create(
            candidate=keep,
            apply_id="SELECT001",
            entity="GW",
            position_name="后端工程师",
            volunteer_rank=2,
            job_category="技术类",
            job=job,
        )
        m.CandidateWorkflow.objects.create(
            candidate=keep,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=keep_resume,
            current_rank=2,
        )
        drop = m.Candidate.objects.create(
            identity_hash="candidate-selector-drop",
            name="其他选择器候选人",
            phone="13830000902",
            highest_major="市场营销",
        )
        drop_resume = m.Resume.objects.create(
            candidate=drop,
            apply_id="SELECT002",
            entity="YLS",
            position_name="产品经理",
            volunteer_rank=1,
            job_category="产品类",
        )
        m.CandidateWorkflow.objects.create(
            candidate=drop,
            status=m.CandidateWorkflow.STATUS_PENDING,
            current_resume=drop_resume,
            current_rank=1,
        )

        options_response = self.client.get("/api/candidates/filter-options/")
        filter_response = self.client.get(
            "/api/candidates/",
            {
                "highest_major_in": "计算机科学与技术",
                "current_rank_in": "2",
                "current_entity_in": "GW",
                "current_position_name_in": "后端工程师",
                "job_department_name_in": "研发中心",
                "current_job_category_in": "技术类",
                "school_tag_in": "目标院校",
            },
        )

        self.assertEqual(options_response.status_code, 200)
        option_values = {
            key: [item["value"] for item in items]
            for key, items in options_response.data.items()
        }
        self.assertEqual(option_values["highest_major"], ["市场营销", "计算机科学与技术"])
        self.assertEqual(option_values["current_rank"], ["1", "2"])
        self.assertIn("GW", option_values["current_entity"])
        self.assertIn("后端工程师", option_values["current_position_name"])
        self.assertIn("研发中心", option_values["job_department_name"])
        self.assertIn("技术类", option_values["current_job_category"])
        self.assertIn("目标院校", option_values["school_tag"])
        self.assertEqual(filter_response.status_code, 200)
        self.assertEqual([item["id"] for item in filter_response.data["results"]], [keep.id])

    def test_candidate_school_tag_filter_matches_manual_non_target_tag(self):
        non_target_tag = m.SchoolTag.objects.create(
            code="NON_TARGET",
            name="非目标院校",
            is_default=False,
            is_active=True,
        )
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-manual-non-target",
            name="未知院校候选人",
            phone="13830000100",
            first_degree_school="未收录大学",
            highest_degree_school="未收录大学",
        )
        m.Resume.objects.create(
            candidate=candidate,
            apply_id="UNKNOWN-SCHOOL",
            position_name="后端工程师",
            volunteer_rank=1,
        )

        classify_school.run()
        candidate.refresh_from_db()
        response = self.client.get("/api/candidates/", {"school_tag": "非目标院校"})

        self.assertEqual(candidate.first_degree_tag, non_target_tag)
        self.assertEqual(candidate.highest_degree_tag, non_target_tag)
        self.assertEqual(response.status_code, 200)
        self.assertEqual([item["id"] for item in response.data["results"]], [candidate.id])
        self.assertEqual(response.data["results"][0]["school_tag"], "非目标院校")

    def test_candidate_list_exposes_workflow_merge_fields_and_assignment_reason(self):
        department = m.Department.objects.create(name="研发中心", level=2)
        contact = m.Contact.objects.create(
            name="二级接口人",
            employee_no="L2200",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
            is_active=True,
        )
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-merge-assignment",
            name="分配候选人",
            phone="13830000001",
        )
        resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="MERGE001",
            position_name="后端工程师",
            volunteer_rank=1,
            job_category="技术类",
            category_reason="岗位名命中",
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=resume,
            current_rank=1,
        )
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_PENDING_DISPATCH,
            department=department,
            contact=contact,
            match_reason="院校准入；分配至研发中心/二级接口人",
        )

        response = self.client.get("/api/candidates/", {"current_apply_id": "MERGE001"})

        self.assertEqual(response.status_code, 200)
        row = response.data["results"][0]
        self.assertEqual(row["current_apply_id"], "MERGE001")
        self.assertEqual(row["job_department_name"], "研发中心")
        self.assertEqual(row["workflow_status"], m.CandidateWorkflow.STATUS_IN_PROGRESS)
        self.assertEqual(row["reason_type"], "assignment")
        self.assertEqual(row["reason_text"], "院校准入；分配至研发中心/二级接口人")
        self.assertEqual(row["attempts"][0]["match_reason"], "院校准入；分配至研发中心/二级接口人")

    def test_candidate_list_exposes_preview_resume_with_file_fallback(self):
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-preview-fallback",
            name="预览候选人",
            phone="13830000002",
        )
        current_resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="NOFILE",
            position_name="当前无文件岗位",
            volunteer_rank=1,
        )
        file_resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="HASFILE",
            position_name="有文件岗位",
            volunteer_rank=2,
            resume_file="预览候选人（HASFILE）.pdf",
        )
        m.CandidateWorkflow.objects.create(
            candidate=candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=current_resume,
            current_rank=1,
        )

        response = self.client.get("/api/candidates/", {"name": "预览候选人"})

        self.assertEqual(response.status_code, 200)
        row = response.data["results"][0]
        self.assertEqual(row["current_resume"]["id"], current_resume.id)
        self.assertEqual(row["preview_resume"]["id"], file_resume.id)
        self.assertEqual(row["preview_resume"]["resume_file"], "预览候选人（HASFILE）.pdf")

    def test_candidate_list_filters_by_blocked_workflow_merge_fields(self):
        department = m.Department.objects.create(name="研发中心", level=2)
        keep = m.Candidate.objects.create(
            identity_hash="candidate-merge-blocked",
            name="阻塞候选人",
            phone="13830000002",
        )
        keep_resume = m.Resume.objects.create(
            candidate=keep,
            apply_id="BLOCK001",
            position_name="后端工程师",
            volunteer_rank=1,
        )
        job = m.Job.objects.create(
            department=department,
            public_name="后端工程师",
            position_name="后端工程师",
            category="技术类",
            headcount=1,
        )
        keep_resume.job = job
        keep_resume.save(update_fields=["job"])
        m.CandidateWorkflow.objects.create(
            candidate=keep,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=keep_resume,
            current_rank=1,
            block_reason=m.CandidateWorkflow.BLOCK_CONTACT_NOT_FOUND,
            block_detail="二级部门研发中心缺少启用接口人",
        )
        drop = m.Candidate.objects.create(
            identity_hash="candidate-merge-drop",
            name="其他候选人",
            phone="13830000003",
        )
        drop_resume = m.Resume.objects.create(
            candidate=drop,
            apply_id="DROP001",
            position_name="产品经理",
            volunteer_rank=1,
        )
        m.CandidateWorkflow.objects.create(
            candidate=drop,
            status=m.CandidateWorkflow.STATUS_ARCHIVED,
            current_resume=drop_resume,
            current_rank=1,
            archive_reason=m.CandidateWorkflow.ARCHIVE_JOB_NOT_MATCHED,
            archive_detail="未匹配岗位",
        )

        response = self.client.get(
            "/api/candidates/",
            {
                "current_apply_id": "BLOCK",
                "job_department_name": "研发",
                "workflow_status": m.CandidateWorkflow.STATUS_IN_PROGRESS,
                "reason_type": "block",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 1)
        row = response.data["results"][0]
        self.assertEqual(row["id"], keep.id)
        self.assertEqual(row["reason_type"], "block")
        self.assertIn("研发中心", row["reason_text"])

    def test_candidate_list_filters_by_multiple_workflow_statuses(self):
        pending = m.Candidate.objects.create(
            identity_hash="candidate-workflow-pending",
            name="待处理候选人",
            phone="13830000004",
        )
        archived = m.Candidate.objects.create(
            identity_hash="candidate-workflow-archived",
            name="归档候选人",
            phone="13830000005",
        )
        passed = m.Candidate.objects.create(
            identity_hash="candidate-workflow-passed",
            name="通过候选人",
            phone="13830000006",
        )
        m.CandidateWorkflow.objects.create(
            candidate=archived,
            status=m.CandidateWorkflow.STATUS_ARCHIVED,
        )
        m.CandidateWorkflow.objects.create(
            candidate=passed,
            status=m.CandidateWorkflow.STATUS_PASSED,
        )

        response = self.client.get(
            "/api/candidates/",
            {
                "workflow_status": ",".join(
                    [
                        m.CandidateWorkflow.STATUS_PENDING,
                        m.CandidateWorkflow.STATUS_ARCHIVED,
                    ]
                )
            },
        )

        self.assertEqual(response.status_code, 200)
        result_ids = {item["id"] for item in response.data["results"]}
        self.assertEqual(result_ids, {pending.id, archived.id})

    def test_candidate_list_uses_current_resume_department_when_history_attempt_exists(self):
        old_department = m.Department.objects.create(name="旧部门", level=2)
        new_department = m.Department.objects.create(name="新部门", level=2)
        old_contact = m.Contact.objects.create(
            name="旧接口人",
            employee_no="L2300",
            department=old_department,
            contact_level=m.Contact.LEVEL_SECONDARY,
            is_active=True,
        )
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-current-department",
            name="当前志愿候选人",
            phone="13830000004",
        )
        old_resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="OLD001",
            position_name="测试工程师",
            volunteer_rank=1,
        )
        current_resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="CUR001",
            position_name="后端工程师",
            volunteer_rank=2,
        )
        current_job = m.Job.objects.create(
            department=new_department,
            public_name="后端工程师",
            position_name="后端工程师",
            category="技术类",
            headcount=1,
        )
        current_resume.job = current_job
        current_resume.save(update_fields=["job"])
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=current_resume,
            current_rank=2,
            block_reason=m.CandidateWorkflow.BLOCK_CONTACT_NOT_FOUND,
            block_detail="当前志愿缺少新部门接口人",
        )
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=old_resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_REJECTED,
            department=old_department,
            contact=old_contact,
            match_reason="历史分配至旧部门",
        )

        response = self.client.get("/api/candidates/", {"current_apply_id": "CUR001"})

        self.assertEqual(response.status_code, 200)
        row = response.data["results"][0]
        self.assertEqual(row["current_apply_id"], "CUR001")
        self.assertEqual(row["job_department_name"], "新部门")
        self.assertEqual(row["reason_type"], "block")
        self.assertIn("新部门", row["reason_text"])
        self.assertEqual(row["system_status"], "archived")

    def test_candidate_list_filters_follow_current_summary_not_history_attempts(self):
        old_department = m.Department.objects.create(name="旧部门", level=2)
        new_department = m.Department.objects.create(name="新部门", level=2)
        old_contact = m.Contact.objects.create(
            name="旧接口人",
            employee_no="L2301",
            department=old_department,
            contact_level=m.Contact.LEVEL_SECONDARY,
            is_active=True,
        )
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-current-filter-summary",
            name="当前筛选候选人",
            phone="13830000005",
        )
        old_resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="OLD002",
            position_name="测试工程师",
            volunteer_rank=1,
            category_reason="历史分类原因",
        )
        current_resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="CUR002",
            position_name="后端工程师",
            volunteer_rank=2,
        )
        current_job = m.Job.objects.create(
            department=new_department,
            public_name="后端工程师",
            position_name="后端工程师",
            category="技术类",
            headcount=1,
        )
        current_resume.job = current_job
        current_resume.save(update_fields=["job"])
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=current_resume,
            current_rank=2,
            block_reason=m.CandidateWorkflow.BLOCK_CONTACT_NOT_FOUND,
            block_detail="当前志愿缺少新部门接口人",
        )
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=old_resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_REJECTED,
            department=old_department,
            contact=old_contact,
            match_reason="历史分配至旧部门",
        )

        old_department_response = self.client.get(
            "/api/candidates/", {"job_department_name": "旧部门"}
        )
        assignment_response = self.client.get(
            "/api/candidates/", {"reason_type": "assignment"}
        )
        classification_response = self.client.get(
            "/api/candidates/", {"reason_type": "classification"}
        )
        current_response = self.client.get(
            "/api/candidates/",
            {"job_department_name": "新部门", "reason_type": "block"},
        )

        self.assertEqual(old_department_response.status_code, 200)
        self.assertEqual(old_department_response.data["count"], 0)
        self.assertEqual(assignment_response.data["count"], 0)
        self.assertEqual(classification_response.data["count"], 0)
        self.assertEqual(current_response.data["count"], 1)
        self.assertEqual(current_response.data["results"][0]["id"], candidate.id)

    def test_candidate_list_current_filters_fallback_to_first_resume_when_workflow_has_no_current_resume(self):
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-current-filter-fallback",
            name="当前字段回退候选人",
            phone="13830000006",
        )
        m.CandidateWorkflow.objects.create(
            candidate=candidate,
            status=m.CandidateWorkflow.STATUS_PENDING,
        )
        m.Resume.objects.create(
            candidate=candidate,
            apply_id="FALL001",
            entity="GW",
            position_name="后端工程师",
            volunteer_rank=1,
            job_category="技术类",
        )
        m.Resume.objects.create(
            candidate=candidate,
            apply_id="FALL002",
            entity="YLS",
            position_name="产品经理",
            volunteer_rank=2,
            job_category="产品类",
        )

        response = self.client.get(
            "/api/candidates/",
            {
                "current_apply_id": "FALL001",
                "current_entity": "GW",
                "current_position_name": "后端",
                "current_job_category": "技术",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 1)
        row = response.data["results"][0]
        self.assertEqual(row["id"], candidate.id)
        self.assertEqual(row["current_apply_id"], "FALL001")
        self.assertEqual(row["current_resume"]["position_name"], "后端工程师")

    def test_candidate_list_filters_by_system_resume_status(self):
        tag = m.SchoolTag.objects.create(code="TARGET", name="目标院校")
        raw = m.Candidate.objects.create(
            identity_hash="candidate-system-raw",
            name="原始候选人",
            phone="13810000000",
        )
        m.Resume.objects.create(candidate=raw, apply_id="RAW001", position_name="后端")
        allocated = m.Candidate.objects.create(
            identity_hash="candidate-system-allocated",
            name="已分配候选人",
            phone="13810000001",
            first_degree_tag=tag,
            highest_degree_tag=tag,
        )
        allocated_resume = m.Resume.objects.create(
            candidate=allocated,
            apply_id="ALLOC001",
            position_name="后端",
            volunteer_rank=1,
            job_category="技术类",
        )
        department = m.Department.objects.create(name="研发中心", level=2)
        contact = m.Contact.objects.create(
            name="二级接口人",
            employee_no="L2100",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=allocated,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=allocated_resume,
            current_rank=1,
        )
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=allocated_resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_PENDING_DISPATCH,
            department=department,
            contact=contact,
        )

        response = self.client.get(
            "/api/candidates/", {"system_status": "pending_dispatch"}
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["id"], allocated.id)
        self.assertEqual(
            response.data["results"][0]["system_status"], "pending_dispatch"
        )
        self.assertEqual(
            response.data["results"][0]["system_status_label"], "待下发"
        )

        for unsupported in ["classified", "allocated", "unknown"]:
            invalid = self.client.get(
                "/api/candidates/", {"system_status": unsupported}
            )
            self.assertEqual(invalid.status_code, 400)

        workflow.status = m.CandidateWorkflow.STATUS_PASSED
        workflow.save(update_fields=["status", "updated_at"])
        attempt = workflow.attempts.get()
        attempt.status = m.AssignmentAttempt.STATUS_PASSED
        attempt.save(update_fields=["status", "updated_at"])
        passed_response = self.client.get(
            "/api/candidates/", {"system_status": "screening_passed"}
        )
        self.assertEqual(
            passed_response.data["results"][0]["system_status_label"], "通过"
        )

        workflow.status = m.CandidateWorkflow.STATUS_IN_PROGRESS
        workflow.save(update_fields=["status", "updated_at"])
        attempt.status = m.AssignmentAttempt.STATUS_REJECTED
        attempt.save(update_fields=["status", "updated_at"])
        rejected_response = self.client.get(
            "/api/candidates/", {"system_status": "screening_rejected"}
        )
        self.assertEqual(
            rejected_response.data["results"][0]["system_status_label"], "不通过"
        )

        attempt.status = m.AssignmentAttempt.STATUS_CANCELLED
        attempt.save(update_fields=["status", "updated_at"])
        archived_response = self.client.get(
            "/api/candidates/", {"system_status": "archived"}
        )
        self.assertEqual(archived_response.status_code, 200)
        self.assertEqual(archived_response.data["results"][0]["id"], allocated.id)

    def test_queued_scope_item_is_raw_but_completed_failure_is_archived(self):
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-queued-status",
            name="排队候选人",
            phone="13810000009",
        )
        m.Resume.objects.create(
            candidate=candidate, apply_id="QUEUE001", position_name="后端"
        )
        run = m.ProcessingRun.objects.create(step="step2", mode="rule")
        item = m.ProcessingRunScopeItem.objects.create(
            run=run, candidate=candidate, status="queued"
        )

        queued = self.client.get(
            "/api/candidates/", {"system_status": "raw"}
        )
        self.assertIn(candidate.id, {row["id"] for row in queued.data["results"]})

        item.status = "failed"
        item.result_type = m.ProcessingRunScopeItem.RESULT_FAILED
        item.reason_code = "ai_connection_error"
        item.result_message = "模型连接失败"
        item.save(
            update_fields=["status", "result_type", "reason_code", "result_message"]
        )
        archived = self.client.get(
            "/api/candidates/", {"system_status": "archived"}
        )
        row = next(row for row in archived.data["results"] if row["id"] == candidate.id)
        self.assertEqual(row["reason_code"], "ai_connection_error")
        self.assertEqual(row["reason_text"], "模型连接失败")

    def test_workflow_list_filters_by_status_search_and_current_position(self):
        keep_candidate = m.Candidate.objects.create(
            identity_hash="candidate-workflow-keep",
            name="归档候选人",
            phone="13820000000",
        )
        keep_resume = m.Resume.objects.create(
            candidate=keep_candidate,
            apply_id="WF001",
            position_name="后端工程师",
            volunteer_rank=1,
        )
        keep = m.CandidateWorkflow.objects.create(
            candidate=keep_candidate,
            status=m.CandidateWorkflow.STATUS_ARCHIVED,
            current_resume=keep_resume,
            current_rank=1,
            archive_reason=m.CandidateWorkflow.ARCHIVE_JOB_NOT_MATCHED,
            archive_detail="未匹配岗位",
        )
        drop_candidate = m.Candidate.objects.create(
            identity_hash="candidate-workflow-drop",
            name="进行中候选人",
            phone="13920000000",
        )
        drop_resume = m.Resume.objects.create(
            candidate=drop_candidate,
            apply_id="WF002",
            position_name="产品经理",
            volunteer_rank=1,
        )
        m.CandidateWorkflow.objects.create(
            candidate=drop_candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=drop_resume,
            current_rank=1,
        )
        archived_drop_candidate = m.Candidate.objects.create(
            identity_hash="candidate-workflow-archived-drop",
            name="归档候选人二",
            phone="13720000000",
        )
        archived_drop_resume = m.Resume.objects.create(
            candidate=archived_drop_candidate,
            apply_id="WF003",
            position_name="产品经理",
            volunteer_rank=1,
        )
        m.CandidateWorkflow.objects.create(
            candidate=archived_drop_candidate,
            status=m.CandidateWorkflow.STATUS_ARCHIVED,
            current_resume=archived_drop_resume,
            current_rank=1,
            archive_reason=m.CandidateWorkflow.ARCHIVE_JOB_NOT_MATCHED,
            archive_detail="未匹配岗位",
        )

        response = self.client.get(
            "/api/workflows/",
            {
                "status": m.CandidateWorkflow.STATUS_ARCHIVED,
                "search": "归档",
                "current_position_name": "后端",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["id"], keep.id)
        self.assertEqual(response.data["results"][0]["current_position_name"], "后端工程师")

    def test_workflow_list_falls_back_to_latest_attempt_resume(self):
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-workflow-fallback",
            name="历史候选人",
            phone="13820000003",
        )
        resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="WF004",
            position_name="算法工程师",
            volunteer_rank=2,
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            status=m.CandidateWorkflow.STATUS_ARCHIVED,
            archive_reason=m.CandidateWorkflow.ARCHIVE_JOB_NOT_MATCHED,
        )
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_CANCELLED,
        )

        response = self.client.get("/api/workflows/", {"status": "archived"})

        self.assertEqual(response.status_code, 200)
        row = response.data["results"][0]
        self.assertEqual(row["current_resume"], resume.id)
        self.assertEqual(row["current_rank"], 2)
        self.assertEqual(row["current_apply_id"], "WF004")
        self.assertEqual(row["current_position_name"], "算法工程师")

    def test_candidate_search_matches_name_full_pinyin_and_initials(self):
        keep = m.Candidate.objects.create(
            identity_hash="candidate-pinyin-keep",
            name="张三",
            phone="13800000001",
        )
        m.Candidate.objects.create(
            identity_hash="candidate-pinyin-drop",
            name="李四",
            phone="13900000002",
        )

        full_response = self.client.get("/api/candidates/", {"search": "zhangsan"})
        initials_response = self.client.get("/api/candidates/", {"search": "zs"})

        self.assertEqual(full_response.status_code, 200)
        self.assertEqual(initials_response.status_code, 200)
        self.assertEqual(
            [item["id"] for item in full_response.data["results"]],
            [keep.id],
        )
        self.assertEqual(
            [item["id"] for item in initials_response.data["results"]],
            [keep.id],
        )

    def test_job_header_filters_cover_visible_columns(self):
        department = m.Department.objects.create(name="研发中心", level=2)
        m.Job.objects.create(
            entity="GW",
            department=department,
            public_name="后端开发",
            position_name="后端工程师",
            category="技术类",
            job_family="研发",
            location="深圳",
            education="本科",
            responsibilities="负责后端服务开发和性能优化。",
            headcount=3,
            is_public=True,
        )
        m.Job.objects.create(
            entity="YLS",
            public_name="产品运营",
            position_name="产品经理",
            category="产品类",
            job_family="产品",
            location="上海",
            education="硕士",
            responsibilities="负责产品规划和项目推进。",
            headcount=1,
            is_public=False,
        )

        response = self.client.get(
            "/api/jobs/",
            {
                "entity": "GW",
                "department_name": "研发",
                "position_name": "工程师",
                "job_family": "研发",
                "location": "深圳",
                "education": "本科",
                "responsibilities": "性能优化",
                "headcount": 3,
                "is_public": "true",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["public_name"], "后端开发")

    def test_job_selector_options_and_exact_multi_value_filters(self):
        department = m.Department.objects.create(name="研发中心", level=2)
        keep = m.Job.objects.create(
            entity="GW",
            department=department,
            public_name="后端开发",
            position_name="后端工程师",
            category="技术类",
            job_family="研发",
            location="深圳",
            education="本科",
            headcount=3,
            is_public=True,
        )
        m.Job.objects.create(
            entity="YLS",
            public_name="产品运营",
            position_name="产品经理",
            category="产品类",
            job_family="产品",
            location="上海",
            education="硕士",
            headcount=1,
            is_public=False,
        )

        options_response = self.client.get("/api/jobs/filter-options/")
        filter_response = self.client.get(
            "/api/jobs/",
            {
                "entity_in": "GW,YLS",
                "public_name_in": "后端开发",
                "position_name_in": "后端工程师",
                "category_in": "技术类",
                "job_family_in": "研发",
                "department_name_in": "研发中心",
                "location_in": "深圳",
                "education_in": "本科",
            },
        )

        self.assertEqual(options_response.status_code, 200)
        self.assertEqual(filter_response.status_code, 200)
        category_options = {
            item["value"]: item for item in options_response.data["category"]
        }
        public_name_options = {
            item["value"]: item
            for item in options_response.data["public_name"]
        }
        position_name_options = {
            item["value"]: item
            for item in options_response.data["position_name"]
        }
        department_options = {
            item["value"]: item
            for item in options_response.data["department_name"]
        }
        self.assertIn("jishulei", category_options["技术类"]["search_text"])
        self.assertIn("jsl", category_options["技术类"]["search_text"])
        self.assertIn("houduankaifa", public_name_options["后端开发"]["search_text"])
        self.assertIn("hdkf", public_name_options["后端开发"]["search_text"])
        self.assertIn(
            "houduangongchengshi",
            position_name_options["后端工程师"]["search_text"],
        )
        self.assertIn("yanfazhongxin", department_options["研发中心"]["search_text"])
        self.assertEqual(
            [item["id"] for item in filter_response.data["results"]], [keep.id]
        )

    def test_job_hierarchy_serialization_filters_and_legacy_secondary_alias(self):
        primary = m.Department.objects.create(name="技术中心", level=1)
        secondary = m.Department.objects.create(
            name="平台部", level=2, parent=primary
        )
        tertiary = m.Department.objects.create(
            name="平台研发组", level=3, parent=secondary
        )
        keep = m.Job.objects.create(
            entity="GW",
            department=tertiary,
            public_name="后端开发",
            position_name="后端工程师",
            category="技术类",
            responsibilities="负责后端研发。",
        )
        m.Job.objects.create(
            entity="GW",
            department=secondary,
            public_name="产品开发",
            position_name="产品经理",
            category="产品类",
            responsibilities="负责产品规划。",
        )

        response = self.client.get(f"/api/jobs/{keep.id}/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["department"], tertiary.id)
        self.assertEqual(response.data["department_name"], "平台部")
        self.assertEqual(response.data["primary_department_id"], primary.id)
        self.assertEqual(response.data["primary_department_name"], "技术中心")
        self.assertEqual(response.data["secondary_department_id"], secondary.id)
        self.assertEqual(response.data["secondary_department_name"], "平台部")
        self.assertEqual(response.data["tertiary_department_id"], tertiary.id)
        self.assertEqual(response.data["tertiary_department_name"], "平台研发组")

        for params in (
            {"primary_department_name_in": "技术中心", "position_name_in": "后端工程师"},
            {"secondary_department_name_in": "平台部", "position_name_in": "后端工程师"},
            {"tertiary_department_name_in": "平台研发组"},
            {"department_name_in": "平台部", "position_name_in": "后端工程师"},
            {"department_name": "平台", "position_name_in": "后端工程师"},
        ):
            filtered = self.client.get("/api/jobs/", params)
            self.assertEqual(filtered.status_code, 200)
            self.assertEqual(
                [item["id"] for item in filtered.data["results"]],
                [keep.id],
            )

        options = self.client.get("/api/jobs/filter-options/")
        self.assertEqual(options.status_code, 200)
        self.assertIn(
            "技术中心",
            [item["value"] for item in options.data["primary_department_name"]],
        )
        self.assertIn(
            "平台部",
            [item["value"] for item in options.data["secondary_department_name"]],
        )
        self.assertIn(
            "平台研发组",
            [item["value"] for item in options.data["tertiary_department_name"]],
        )
        self.assertEqual(
            options.data["department_name"],
            options.data["secondary_department_name"],
        )

    def test_job_name_filters_support_full_pinyin_and_initials(self):
        keep = m.Job.objects.create(
            entity="GW",
            public_name="后端开发",
            position_name="后端工程师",
            category="技术类",
            is_active=True,
        )
        m.Job.objects.create(
            entity="YLS",
            public_name="产品运营",
            position_name="产品经理",
            category="产品类",
            is_active=True,
        )

        full_response = self.client.get(
            "/api/jobs/", {"public_name": "houduankaifa"}
        )
        initials_response = self.client.get(
            "/api/jobs/", {"position_name": "hdgcs"}
        )

        self.assertEqual(full_response.status_code, 200)
        self.assertEqual(initials_response.status_code, 200)
        self.assertEqual(
            [item["id"] for item in full_response.data["results"]], [keep.id]
        )
        self.assertEqual(
            [item["id"] for item in initials_response.data["results"]], [keep.id]
        )

    def test_job_create_and_update_maintains_demand_majors(self):
        department = m.Department.objects.create(name="研发中心", level=2)

        create_response = self.client.post(
            "/api/jobs/",
            {
                "entity": "GW",
                "department": department.id,
                "public_name": "后端开发",
                "position_name": "后端工程师",
                "category": "技术类",
                "responsibilities": "负责后端服务开发和性能优化。",
                "major_names": ["计算机", "软件工程"],
                "headcount": 3,
            },
            format="json",
        )

        self.assertEqual(create_response.status_code, 201)
        self.assertEqual(
            create_response.data["responsibilities"],
            "负责后端服务开发和性能优化。",
        )
        self.assertEqual(create_response.data["majors"], ["计算机", "软件工程"])
        job_id = create_response.data["id"]
        self.assertEqual(
            list(
                m.JobMajor.objects.filter(job_id=job_id).values_list(
                    "major", flat=True
                )
            ),
            ["计算机", "软件工程"],
        )

        update_response = self.client.patch(
            f"/api/jobs/{job_id}/",
            {"major_names": ["数学"]},
            format="json",
        )

        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.data["majors"], ["数学"])
        self.assertEqual(
            list(
                m.JobMajor.objects.filter(job_id=job_id).values_list(
                    "major", flat=True
                )
            ),
            ["数学"],
        )

    def test_job_rejects_non_secondary_department(self):
        root_department = m.Department.objects.create(name="集团总部", level=1)

        response = self.client.post(
            "/api/jobs/",
            {
                "entity": "GW",
                "department": root_department.id,
                "public_name": "后端开发",
                "position_name": "后端工程师",
                "category": "技术类",
                "responsibilities": "负责后端服务开发。",
                "major_names": ["计算机"],
                "headcount": 3,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("department", response.data)

    def test_job_create_and_legacy_update_require_responsibilities(self):
        create_response = self.client.post(
            "/api/jobs/",
            {
                "public_name": "缺少职责岗位",
                "position_name": "缺少职责岗位",
                "responsibilities": "   ",
            },
            format="json",
        )
        legacy_job = m.Job.objects.create(
            public_name="历史岗位",
            position_name="历史岗位",
            responsibilities="",
        )
        update_response = self.client.patch(
            f"/api/jobs/{legacy_job.id}/",
            {"headcount": 2},
            format="json",
        )

        self.assertEqual(create_response.status_code, 400)
        self.assertIn("responsibilities", create_response.data)
        self.assertEqual(update_response.status_code, 400)
        self.assertIn("responsibilities", update_response.data)

    def test_delete_job_deactivates_instead_of_removing_record(self):
        department = m.Department.objects.create(name="研发中心", level=2)
        job = m.Job.objects.create(
            entity="GW",
            department=department,
            public_name="后端开发",
            position_name="后端工程师",
            category="技术类",
            headcount=3,
            is_active=True,
        )

        response = self.client.delete(f"/api/jobs/{job.id}/")

        self.assertEqual(response.status_code, 204)
        job.refresh_from_db()
        self.assertFalse(job.is_active)

    def test_job_list_defaults_to_active_and_can_filter_inactive(self):
        m.Job.objects.create(
            entity="GW",
            public_name="启用岗位",
            position_name="后端工程师",
            category="技术类",
            is_active=True,
        )
        inactive = m.Job.objects.create(
            entity="GW",
            public_name="停用岗位",
            position_name="旧岗位",
            category="技术类",
            is_active=False,
        )

        active_response = self.client.get("/api/jobs/")
        inactive_response = self.client.get("/api/jobs/", {"is_active": "false"})

        self.assertEqual(active_response.status_code, 200)
        self.assertEqual(inactive_response.status_code, 200)
        self.assertEqual(active_response.data["count"], 1)
        self.assertEqual(inactive_response.data["count"], 1)
        self.assertEqual(inactive_response.data["results"][0]["id"], inactive.id)

    def test_school_header_filters_cover_visible_columns(self):
        m.School.objects.create(name="南京大学", platform="平台A", province="江苏")
        m.School.objects.create(name="北京大学", platform="平台B", province="北京")

        response = self.client.get("/api/schools/", {"province": "江苏"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["name"], "南京大学")

    def test_school_name_pinyin_search_and_platform_selector(self):
        keep = m.School.objects.create(
            name="南京大学", platform="双一流", province="江苏"
        )
        m.School.objects.create(name="北京大学", platform="985", province="北京")

        full_response = self.client.get("/api/schools/", {"name": "nanjingdaxue"})
        initials_response = self.client.get("/api/schools/", {"name": "njdx"})
        selector_response = self.client.get(
            "/api/schools/", {"platform_in": "双一流"}
        )
        options_response = self.client.get("/api/schools/filter-options/")

        self.assertEqual(full_response.status_code, 200)
        self.assertEqual(initials_response.status_code, 200)
        self.assertEqual(selector_response.status_code, 200)
        self.assertEqual(options_response.status_code, 200)
        self.assertEqual(
            [item["id"] for item in full_response.data["results"]], [keep.id]
        )
        self.assertEqual(
            [item["id"] for item in initials_response.data["results"]], [keep.id]
        )
        self.assertEqual(
            [item["id"] for item in selector_response.data["results"]], [keep.id]
        )
        options = {item["value"]: item for item in options_response.data["platform"]}
        self.assertIn("shuangyiliu", options["双一流"]["search_text"])
        keep.refresh_from_db()
        self.assertEqual(keep.name_pinyin, "nanjingdaxue")
        self.assertEqual(keep.name_pinyin_initials, "njdx")

    def test_contact_header_filters_cover_visible_columns(self):
        department = m.Department.objects.create(name="研发二部", level=2, entity="GW")
        m.Contact.objects.create(
            name="王五",
            employee_no="E1001",
            email="wangwu@example.com",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
            can_delegate=True,
            is_active=True,
        )
        m.Contact.objects.create(
            name="赵六",
            employee_no="E2001",
            email="zhaoliu@example.com",
            department=m.Department.objects.create(name="产品二部", level=2, entity="YLS"),
            contact_level=m.Contact.LEVEL_TERTIARY,
            can_delegate=False,
            is_active=False,
        )

        response = self.client.get(
            "/api/contacts/",
            {
                "department_level": 2,
                "can_delegate": "true",
                "entity": "GW",
                "email": "wangwu@",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["name"], "王五")

    def test_contact_department_selector_options_and_multi_filter(self):
        research = m.Department.objects.create(name="研发二部", level=2)
        product = m.Department.objects.create(name="产品二部", level=2)
        keep = m.Contact.objects.create(
            name="王五",
            employee_no="E1001",
            department=research,
            contact_level=m.Contact.LEVEL_SECONDARY,
            is_active=True,
        )
        m.Contact.objects.create(
            name="赵六",
            employee_no="E2001",
            department=product,
            contact_level=m.Contact.LEVEL_TERTIARY,
            is_active=True,
        )

        options_response = self.client.get("/api/contacts/filter-options/")
        filter_response = self.client.get(
            "/api/contacts/", {"department_in": f"{research.id},999999"}
        )

        self.assertEqual(options_response.status_code, 200)
        self.assertEqual(filter_response.status_code, 200)
        options = {
            item["value"]: item for item in options_response.data["department"]
        }
        self.assertIn(research.id, options)
        self.assertIn("yanfaerbu", options[research.id]["search_text"])
        self.assertEqual(
            [item["id"] for item in filter_response.data["results"]], [keep.id]
        )

    def test_contact_name_and_bound_user_support_full_pinyin_and_initials(self):
        department = m.Department.objects.create(name="研发二部", level=2)
        contact = m.Contact.objects.create(
            name="王五",
            employee_no="PINYIN001",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        user = User.objects.create_user(
            username="PINYIN001",
            password="pass",
            role=User.ROLE_SECONDARY_CONTACT,
            contact=contact,
        )

        full_response = self.client.get("/api/contacts/", {"name": "wangwu"})
        initials_response = self.client.get("/api/contacts/", {"name": "ww"})
        user_response = self.client.get("/api/users/", {"contact_name": "wangwu"})

        self.assertEqual([item["id"] for item in full_response.data["results"]], [contact.id])
        self.assertEqual([item["id"] for item in initials_response.data["results"]], [contact.id])
        self.assertEqual([item["id"] for item in user_response.data["results"]], [user.id])
        contact.refresh_from_db()
        self.assertEqual(contact.name_pinyin, "wangwu")
        self.assertEqual(contact.name_pinyin_initials, "ww")

    def test_candidate_contact_filter_supports_contact_name_pinyin(self):
        department = m.Department.objects.create(name="技术二部", level=2)
        contact = m.Contact.objects.create(
            name="王五",
            employee_no="PINYIN002",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-contact-pinyin",
            name="候选人甲",
            phone="13812340000",
        )
        resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="PINYIN-APPLY",
            position_name="后端工程师",
            volunteer_rank=1,
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=resume,
            current_rank=1,
        )
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_PENDING_DISPATCH,
            department=department,
            contact=contact,
        )

        response = self.client.get("/api/candidates/", {"contact_name": "ww"})

        self.assertEqual([item["id"] for item in response.data["results"]], [candidate.id])

    def test_candidate_filter_options_respect_contact_rbac_scope(self):
        department = m.Department.objects.create(name="范围二部", level=2)
        own_contact = m.Contact.objects.create(
            name="本人接口人",
            employee_no="SCOPE001",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        other_contact = m.Contact.objects.create(
            name="其他接口人",
            employee_no="SCOPE002",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        contact_user = User.objects.create_user(
            username="SCOPE001",
            password="pass",
            role=User.ROLE_SECONDARY_CONTACT,
            contact=own_contact,
        )
        contact_user.groups.add(Group.objects.get(name="二级接口人"))
        for index, (contact, position) in enumerate(
            [(own_contact, "可见岗位"), (other_contact, "不可见岗位")], start=1
        ):
            candidate = m.Candidate.objects.create(
                identity_hash=f"candidate-option-scope-{index}",
                name=f"范围候选人{index}",
                phone=f"1385555000{index}",
            )
            resume = m.Resume.objects.create(
                candidate=candidate,
                apply_id=f"SCOPE-APPLY-{index}",
                position_name=position,
                volunteer_rank=1,
            )
            workflow = m.CandidateWorkflow.objects.create(
                candidate=candidate,
                status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
                current_resume=resume,
                current_rank=1,
            )
            m.AssignmentAttempt.objects.create(
                workflow=workflow,
                resume=resume,
                attempt_no=1,
                source=m.AssignmentAttempt.SOURCE_RULE,
                status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
                department=department,
                contact=contact,
            )

        self.client.force_authenticate(contact_user)
        response = self.client.get("/api/candidates/filter-options/")

        values = [item["value"] for item in response.data["current_position_name"]]
        self.assertEqual(values, ["可见岗位"])

    def test_contact_and_school_pinyin_data_migration_populates_existing_rows(self):
        contact = m.Contact.objects.create(name="赵六", employee_no="MIGRATION001")
        school = m.School.objects.create(name="北京大学")
        m.Contact.objects.filter(pk=contact.pk).update(
            name_pinyin="", name_pinyin_initials=""
        )
        m.School.objects.filter(pk=school.pk).update(
            name_pinyin="", name_pinyin_initials=""
        )
        migration = importlib.import_module(
            "apps.core.migrations.0023_contact_school_name_pinyin"
        )

        migration.populate_name_pinyin(django_apps, None)

        contact.refresh_from_db()
        school.refresh_from_db()
        self.assertEqual((contact.name_pinyin, contact.name_pinyin_initials), ("zhaoliu", "zl"))
        self.assertEqual((school.name_pinyin, school.name_pinyin_initials), ("beijingdaxue", "bjdx"))


class JobExportApiTests(TestCase):
    def setUp(self):
        ensure_rbac_defaults()
        self.client = APIClient()
        self.admin = User.objects.create_user(
            username="admin-job-export", password="pass", role=User.ROLE_ADMIN
        )
        self.admin.groups.add(Group.objects.get(name="管理员"))
        self.client.force_authenticate(self.admin)
        self.primary = m.Department.objects.create(name="技术中心", level=1)
        self.secondary = m.Department.objects.create(
            name="平台部", level=2, parent=self.primary
        )
        self.tertiary = m.Department.objects.create(
            name="平台研发组", level=3, parent=self.secondary
        )

    def test_export_returns_all_filtered_active_jobs_as_reimportable_xlsx(self):
        jobs = []
        for index in range(22):
            job = m.Job.objects.create(
                entity="GW",
                department=self.tertiary,
                category="技术类",
                public_name=f"后端开发{index}",
                is_public=index % 2 == 0,
                position_name=f"后端工程师{index}",
                job_family="研发族",
                location="深圳",
                education="本科",
                responsibilities=(
                    "=HYPERLINK(\"https://example.test\")"
                    if index == 0
                    else f"负责后端研发{index}"
                ),
                headcount=index + 1,
            )
            jobs.append(job)
        m.JobMajor.objects.create(job=jobs[0], major="计算机")
        m.JobMajor.objects.create(job=jobs[0], major="软件工程")
        m.Job.objects.create(
            entity="GW",
            department=self.tertiary,
            public_name="停用岗位",
            position_name="停用岗位",
            responsibilities="停用岗位职责",
            is_active=False,
        )
        m.Job.objects.create(
            entity="YLS",
            department=self.tertiary,
            public_name="其它主体岗位",
            position_name="其它主体岗位",
            responsibilities="其它主体职责",
        )

        response = self.client.get(
            "/api/jobs/export/",
            {
                "entity_in": "GW",
                "tertiary_department_name_in": "平台研发组",
                "page": 2,
                "page_size": 1,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["X-Export-Count"], "22")
        self.assertIn(quote("职位清单.xlsx"), response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["职位清单"]
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            [
                "招聘主体",
                "一层部门",
                "二层部门",
                "三级部门",
                "岗位类别",
                "对外发布名称",
                "是否对外发布",
                "职位名称",
                "岗位族",
                "工作地点",
                "学历",
                "工作职责",
                "需求专业",
                "HC",
            ],
        )
        self.assertEqual(sheet.max_row, 23)
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.auto_filter.ref, "A1:N23")
        self.assertEqual(
            [sheet.cell(row=2, column=index).value for index in range(1, 15)],
            [
                "GW",
                "技术中心",
                "平台部",
                "平台研发组",
                "技术类",
                "后端开发0",
                "是",
                "后端工程师0",
                "研发族",
                "深圳",
                "本科",
                "'=HYPERLINK(\"https://example.test\")",
                "计算机、软件工程",
                1,
            ],
        )

    def test_export_rejects_empty_result_and_requires_job_view_permission(self):
        empty = self.client.get("/api/jobs/export/", {"entity_in": "不存在"})
        self.assertEqual(empty.status_code, 400)
        self.assertIn("没有可下载", empty.data["detail"])

        viewer = User.objects.create_user(username="job-export-no-permission")
        self.client.force_authenticate(viewer)
        forbidden = self.client.get("/api/jobs/export/")
        self.assertEqual(forbidden.status_code, 403)


class CandidateExportApiTests(TestCase):
    def setUp(self):
        ensure_rbac_defaults()
        self.client = APIClient()
        self.hr = User.objects.create_user(
            username="hr-export", password="pass", role=User.ROLE_HR
        )
        self.hr.groups.add(Group.objects.get(name="HR"))
        self.client.force_authenticate(self.hr)

    @staticmethod
    def _export_workbook(response):
        with zipfile.ZipFile(BytesIO(response.content)) as archive:
            return load_workbook(BytesIO(archive.read("简历库清单.xlsx")))

    def test_export_fields_returns_stable_public_catalog_and_defaults(self):
        response = self.client.get("/api/candidates/export-fields/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["version"], 2)
        self.assertEqual(
            [group["key"] for group in response.data["groups"]],
            ["candidate", "current_resume", "job", "allocation", "status_reason"],
        )
        fields = [
            field
            for group in response.data["groups"]
            for field in group["fields"]
        ]
        keys = [field["key"] for field in fields]
        self.assertNotIn("processing_result", keys)
        self.assertEqual(len(keys), len(set(keys)))
        self.assertEqual(
            {field["key"] for field in fields if field["default_selected"]},
            {
                "candidate_name",
                "candidate_phone",
                "current_apply_id",
                "current_position_name",
                "volunteer_rank",
                "allocation_secondary_department",
                "secondary_contact",
                "tertiary_contact",
                "allocation_source",
                "resume_status",
            },
        )
        for private_key in {
            "id",
            "identity_hash",
            "raw_text",
            "special_route_config_snapshot",
            "created_by_username_snapshot",
        }:
            self.assertNotIn(private_key, keys)

    def test_export_rejects_empty_and_unknown_fields_and_uses_catalog_order(self):
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-export-fields",
            name="字段候选人",
            phone="13800000009",
        )
        m.Resume.objects.create(candidate=candidate, apply_id="FIELD-1")

        empty = self.client.get(
            f"/api/candidates/export/?ids={candidate.id}&fields="
        )
        unknown = self.client.get(
            "/api/candidates/export/",
            {"ids": candidate.id, "fields": "candidate_name,database_id"},
        )
        ordered = self.client.get(
            "/api/candidates/export/",
            {
                "ids": candidate.id,
                "fields": "responsibilities,current_apply_id,candidate_name",
            },
        )
        defaulted = self.client.get(
            "/api/candidates/export/", {"ids": candidate.id}
        )

        self.assertEqual(empty.status_code, 400)
        self.assertEqual(unknown.status_code, 400)
        self.assertIn("database_id", unknown.data["detail"])
        self.assertEqual(ordered.status_code, 200)
        sheet = self._export_workbook(ordered)["简历库"]
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            ["姓名", "当前应聘ID", "工作职责"],
        )
        default_sheet = self._export_workbook(defaulted)["简历库"]
        self.assertEqual(
            [cell.value for cell in default_sheet[1]],
            [
                "姓名",
                "手机号",
                "当前应聘ID",
                "当前岗位",
                "当前志愿",
                "分配来源",
                "二级部门",
                "二级接口人",
                "三级接口人",
                "简历状态",
            ],
        )

    def test_candidate_export_uses_current_volunteer_and_protects_excel_text(self):
        secondary = m.Department.objects.create(name="当前二级部", level=2)
        contact = m.Contact.objects.create(
            name="当前接口人",
            employee_no="EXPORT-CURRENT-L2",
            department=secondary,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        job = m.Job.objects.create(
            entity="GW",
            department=secondary,
            public_name="当前岗位对外名",
            position_name="当前岗位职位名",
            responsibilities="=HYPERLINK(\"bad\")",
        )
        m.JobMajor.objects.create(job=job, major="计算机")
        m.JobMajor.objects.create(job=job, major="软件工程")
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-current-volunteer-export",
            name="=2+2",
            phone="+8613800000010",
        )
        first = m.Resume.objects.create(
            candidate=candidate,
            apply_id="CURRENT-1",
            volunteer_rank=1,
            position_name="第一志愿",
            resume_file="first.pdf",
        )
        current = m.Resume.objects.create(
            candidate=candidate,
            apply_id="CURRENT-2",
            volunteer_rank=2,
            position_name="当前志愿",
            resume_file="current.pdf",
            job=job,
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            current_resume=current,
            current_rank=2,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
        )
        fixed_utc = datetime.fromisoformat("2026-07-01T00:00:00+00:00")
        m.Candidate.objects.filter(pk=candidate.pk).update(imported_at=fixed_utc)
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=current,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
            department=secondary,
            contact=contact,
            department_name_snapshot="当前二级部快照",
            dispatched_at=fixed_utc,
        )

        response = self.client.get(
            "/api/candidates/export/",
            {
                "ids": candidate.id,
                "fields": (
                    "responsibilities,required_majors,all_resume_filenames,"
                    "all_apply_ids,allocation_secondary_department,volunteer_rank,"
                    "current_position_name,current_apply_id,candidate_phone,candidate_name,"
                    "candidate_imported_at,dispatched_at"
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        sheet = self._export_workbook(response)["简历库"]
        headers = [cell.value for cell in sheet[1]]
        row = dict(zip(headers, [cell.value for cell in sheet[2]]))
        self.assertEqual(row["姓名"], "'=2+2")
        self.assertEqual(row["手机号"], "'+8613800000010")
        self.assertEqual(row["当前应聘ID"], current.apply_id)
        self.assertEqual(row["当前岗位"], "当前志愿")
        self.assertEqual(row["当前志愿"], 2)
        self.assertEqual(row["全部应聘ID"], f"{first.apply_id}、{current.apply_id}")
        self.assertEqual(row["全部简历文件名"], "first.pdf、current.pdf")
        self.assertEqual(row["需求专业"], "计算机、软件工程")
        self.assertEqual(row["工作职责"], "'=HYPERLINK(\"bad\")")
        self.assertEqual(row["二级部门"], "当前二级部快照")
        self.assertEqual(row["候选人导入时间"], "2026-07-01 08:00:00")
        self.assertEqual(row["下发时间"], "2026-07-01 08:00:00")
        responsibility_column = headers.index("工作职责") + 1
        self.assertTrue(sheet.cell(row=2, column=responsibility_column).alignment.wrap_text)

    def test_contact_export_only_exposes_visible_attempt_resume_and_phone(self):
        own_department = m.Department.objects.create(name="本人二级部", level=2)
        other_department = m.Department.objects.create(name="其他二级部", level=2)
        own_contact = m.Contact.objects.create(
            name="本人接口人",
            employee_no="EXPORT-OWN-L2",
            department=own_department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        other_contact = m.Contact.objects.create(
            name="其他接口人",
            employee_no="EXPORT-OTHER-L2",
            department=other_department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-contact-export-scope",
            name="范围候选人",
            phone="13812345678",
        )
        visible_resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="VISIBLE-APPLY",
            volunteer_rank=1,
            resume_file="visible.pdf",
        )
        hidden_resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="HIDDEN-APPLY",
            volunteer_rank=2,
            resume_file="hidden.pdf",
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            current_resume=hidden_resume,
            current_rank=2,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
        )
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=visible_resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
            department=own_department,
            contact=own_contact,
        )
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=hidden_resume,
            attempt_no=2,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
            department=other_department,
            contact=other_contact,
        )
        user = User.objects.create_user(
            username="EXPORT-OWN-L2",
            password="pass",
            role=User.ROLE_SECONDARY_CONTACT,
            contact=own_contact,
        )
        user.groups.add(Group.objects.get(name="二级接口人"))

        with TemporaryDirectory() as media_root:
            resume_dir = Path(media_root) / "resumes"
            resume_dir.mkdir()
            (resume_dir / "visible.pdf").write_bytes(b"visible")
            (resume_dir / "hidden.pdf").write_bytes(b"hidden")
            self.client.force_authenticate(user)
            fields_response = self.client.get("/api/candidates/export-fields/")
            with self.settings(MEDIA_ROOT=media_root):
                response = self.client.get(
                    "/api/candidates/export/",
                    {
                        "ids": candidate.id,
                        "fields": (
                            "candidate_phone,current_apply_id,resume_filename,"
                            "all_apply_ids,all_resume_filenames"
                        ),
                    },
                )

        self.assertEqual(fields_response.status_code, 200)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["X-Export-Candidate-Count"], "1")
        sheet = self._export_workbook(response)["简历库"]
        headers = [cell.value for cell in sheet[1]]
        row = dict(zip(headers, [cell.value for cell in sheet[2]]))
        self.assertEqual(row["手机号"], candidate.phone)
        self.assertEqual(row["当前应聘ID"], visible_resume.apply_id)
        self.assertEqual(row["简历文件名"], "visible.pdf")
        self.assertEqual(row["全部应聘ID"], visible_resume.apply_id)
        self.assertEqual(row["全部简历文件名"], "visible.pdf")
        with zipfile.ZipFile(BytesIO(response.content)) as archive:
            self.assertIn("简历文件/visible.pdf", archive.namelist())
            self.assertNotIn("简历文件/hidden.pdf", archive.namelist())

    def test_same_named_resume_files_are_disambiguated_with_apply_id(self):
        with TemporaryDirectory() as media_root:
            resume_dir = Path(media_root) / "resumes"
            resume_dir.mkdir()
            (resume_dir / "同名.pdf").write_bytes(b"same")
            candidate = m.Candidate.objects.create(
                identity_hash="candidate-same-file-export",
                name="同名候选人",
                phone="13800000011",
            )
            for apply_id in ("SAME-1", "SAME-2"):
                m.Resume.objects.create(
                    candidate=candidate,
                    apply_id=apply_id,
                    resume_file="同名.pdf",
                )
            with self.settings(MEDIA_ROOT=media_root):
                response = self.client.get(
                    "/api/candidates/export/", {"ids": candidate.id}
                )

        self.assertEqual(response.status_code, 200)
        with zipfile.ZipFile(BytesIO(response.content)) as archive:
            self.assertIn("简历文件/同名（SAME-1）.pdf", archive.namelist())
            self.assertIn("简历文件/同名（SAME-2）.pdf", archive.namelist())

    def test_candidate_export_returns_zip_for_single_available_resume(self):
        with TemporaryDirectory() as media_root:
            resume_dir = Path(media_root) / "resumes"
            resume_dir.mkdir()
            (resume_dir / "张三（A1001）.txt").write_text("resume body", encoding="utf-8")
            candidate = m.Candidate.objects.create(
                identity_hash="candidate-export",
                name="张三",
                phone="13800000000",
            )
            m.Resume.objects.create(
                candidate=candidate,
                apply_id="A1001",
                position_name="后端工程师",
                resume_file="张三（A1001）.txt",
            )

            with self.settings(MEDIA_ROOT=media_root):
                response = self.client.get(
                    "/api/candidates/export/", {"ids": str(candidate.id)}
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        self.assertEqual(response["X-Export-Count"], "1")
        self.assertEqual(response["X-Export-Missing"], "0")
        self.assertEqual(response["X-Export-Candidate-Count"], "1")
        self.assertIn("attachment", response["Content-Disposition"])
        with zipfile.ZipFile(BytesIO(response.content)) as zf:
            self.assertEqual(
                set(zf.namelist()),
                {
                    "简历库清单.xlsx",
                    "简历文件/",
                    "简历文件/张三（A1001）.txt",
                },
            )
            self.assertEqual(
                zf.read("简历文件/张三（A1001）.txt").decode("utf-8"),
                "resume body",
            )

    def test_candidate_export_returns_zip_when_any_resume_is_missing(self):
        with TemporaryDirectory() as media_root:
            resume_dir = Path(media_root) / "resumes"
            resume_dir.mkdir()
            (resume_dir / "张三（A1001）.txt").write_text("resume body", encoding="utf-8")
            candidate = m.Candidate.objects.create(
                identity_hash="candidate-export-missing",
                name="张三",
                phone="13800000000",
            )
            m.Resume.objects.create(
                candidate=candidate,
                apply_id="A1001",
                position_name="后端工程师",
                resume_file="张三（A1001）.txt",
            )
            m.Resume.objects.create(
                candidate=candidate,
                apply_id="A1002",
                position_name="算法工程师",
                resume_file="",
            )

            with self.settings(MEDIA_ROOT=media_root):
                response = self.client.get(
                    "/api/candidates/export/", {"ids": str(candidate.id)}
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        self.assertEqual(response["X-Export-Count"], "1")
        self.assertEqual(response["X-Export-Missing"], "1")
        self.assertEqual(response["X-Export-Candidate-Count"], "1")
        with zipfile.ZipFile(BytesIO(response.content)) as zf:
            self.assertIn("简历库清单.xlsx", zf.namelist())
            self.assertIn("简历文件/", zf.namelist())
            self.assertIn("简历文件/张三（A1001）.txt", zf.namelist())
            self.assertIn("缺失简历文件清单.txt", zf.namelist())

    def test_candidate_export_keeps_resume_directory_when_all_files_are_missing(self):
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-export-all-missing",
            name="无文件候选人",
            phone="13800000012",
        )
        m.Resume.objects.create(
            candidate=candidate,
            apply_id="MISSING-ONLY",
            resume_file="missing.pdf",
        )

        response = self.client.get(
            "/api/candidates/export/", {"ids": candidate.id}
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["X-Export-Count"], "0")
        self.assertEqual(response["X-Export-Missing"], "1")
        with zipfile.ZipFile(BytesIO(response.content)) as archive:
            self.assertIn("简历库清单.xlsx", archive.namelist())
            self.assertIn("简历文件/", archive.namelist())
            self.assertIn("缺失简历文件清单.txt", archive.namelist())

    def test_attempt_export_returns_zip_for_single_available_resume(self):
        department = m.Department.objects.create(name="研发中心", level=2)
        contact = m.Contact.objects.create(
            name="二级接口人",
            employee_no="S9001",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        with TemporaryDirectory() as media_root:
            resume_dir = Path(media_root) / "resumes"
            resume_dir.mkdir()
            (resume_dir / "李四（B1001）.txt").write_text("attempt resume", encoding="utf-8")
            candidate = m.Candidate.objects.create(
                identity_hash="attempt-export",
                name="李四",
                phone="13900000000",
            )
            resume = m.Resume.objects.create(
                candidate=candidate,
                apply_id="B1001",
                position_name="产品经理",
                resume_file="李四（B1001）.txt",
            )
            workflow = m.CandidateWorkflow.objects.create(
                candidate=candidate,
                status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
                current_resume=resume,
                current_rank=1,
            )
            attempt = m.AssignmentAttempt.objects.create(
                workflow=workflow,
                resume=resume,
                attempt_no=1,
                source=m.AssignmentAttempt.SOURCE_RULE,
                status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
                department=department,
                contact=contact,
            )

            with self.settings(MEDIA_ROOT=media_root):
                response = self.client.get(
                    "/api/workflow-attempts/export/", {"ids": str(attempt.id)}
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        self.assertEqual(response["X-Export-Count"], "1")
        self.assertEqual(response["X-Export-Missing"], "0")
        self.assertEqual(response["X-Export-Candidate-Count"], "1")
        self.assertIn("attachment", response["Content-Disposition"])
        with zipfile.ZipFile(BytesIO(response.content)) as zf:
            self.assertEqual(
                zf.read("简历文件/李四（B1001）.txt").decode("utf-8"),
                "attempt resume",
            )

    def test_attempt_export_uses_latest_selected_attempt_per_candidate(self):
        first_department = m.Department.objects.create(name="第一尝试部门", level=2)
        latest_department = m.Department.objects.create(name="最新尝试部门", level=2)
        candidate = m.Candidate.objects.create(
            identity_hash="attempt-export-latest-selected",
            name="多尝试候选人",
            phone="13900000001",
        )
        first_resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="ATTEMPT-FIRST",
            volunteer_rank=1,
            resume_file="first-attempt.pdf",
        )
        latest_resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="ATTEMPT-LATEST",
            volunteer_rank=2,
            resume_file="latest-attempt.pdf",
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            current_resume=latest_resume,
            current_rank=2,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
        )
        first_attempt = m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=first_resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_REJECTED,
            department=first_department,
        )
        latest_attempt = m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=latest_resume,
            attempt_no=2,
            source=m.AssignmentAttempt.SOURCE_MANUAL,
            status=m.AssignmentAttempt.STATUS_PENDING_DISPATCH,
            department=latest_department,
        )

        response = self.client.get(
            "/api/workflow-attempts/export/",
            {
                "ids": f"{first_attempt.id},{latest_attempt.id}",
                "fields": (
                    "current_apply_id,all_apply_ids,allocation_source,"
                    "allocation_secondary_department"
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["X-Export-Candidate-Count"], "1")
        sheet = self._export_workbook(response)["简历库"]
        self.assertEqual(sheet.max_row, 2)
        headers = [cell.value for cell in sheet[1]]
        row = dict(zip(headers, [cell.value for cell in sheet[2]]))
        self.assertEqual(row["当前应聘ID"], latest_resume.apply_id)
        self.assertEqual(
            row["全部应聘ID"],
            f"{first_resume.apply_id}、{latest_resume.apply_id}",
        )
        self.assertEqual(row["分配来源"], "手动强制分配")
        self.assertEqual(row["二级部门"], latest_department.name)


class ResumeResultReportApiTests(TestCase):
    def setUp(self):
        ensure_rbac_defaults()
        self.client = APIClient()
        self.hr = User.objects.create_user(
            username="hr-result-report", password="pass", role=User.ROLE_HR
        )
        self.hr.groups.add(Group.objects.get(name="HR"))
        self.client.force_authenticate(self.hr)
        self.department = m.Department.objects.create(name="当前研发部", level=2)
        self.contact = m.Contact.objects.create(
            name="当前二级接口人",
            employee_no="REPORT-L2",
            department=self.department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        self.start_at = timezone.make_aware(datetime(2026, 7, 1, 0, 0, 0))

    def _resume(self, suffix, *, name="张三", offset_hours=0):
        candidate = m.Candidate.objects.create(
            identity_hash=f"report-{suffix}",
            name=name,
            phone=f"1380000{int(suffix):04d}",
            highest_education=m.Candidate.EDUCATION_MASTER,
        )
        resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id=f"REPORT-{suffix}",
            entity="GW",
            position_name="后端工程师",
            volunteer_rank=1,
        )
        m.Resume.objects.filter(pk=resume.pk).update(
            imported_at=self.start_at + timedelta(hours=offset_hours)
        )
        resume.refresh_from_db()
        return resume

    def _attempt(self, resume, status, attempt_no=1):
        workflow = m.CandidateWorkflow.objects.create(
            candidate=resume.candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=resume,
            current_rank=1,
        )
        return m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=resume,
            attempt_no=attempt_no,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=status,
            department=self.department,
            contact=self.contact,
            department_name_snapshot="历史研发部",
            contact_name_snapshot="历史二级接口人",
        )

    def test_result_report_validates_required_and_ordered_dates(self):
        missing = self.client.get("/api/resumes/result-report/")
        invalid = self.client.get(
            "/api/resumes/result-report/",
            {"imported_after": "2026-07-xx", "imported_before": "2026-07-02"},
        )
        reversed_range = self.client.get(
            "/api/resumes/result-report/",
            {"imported_after": "2026-07-03", "imported_before": "2026-07-02"},
        )

        self.assertEqual(missing.status_code, 400)
        self.assertEqual(invalid.status_code, 400)
        self.assertEqual(reversed_range.status_code, 400)

        invalid_department = self.client.get(
            "/api/resumes/result-report/",
            {
                "imported_after": "2026-07-01",
                "imported_before": "2026-07-02",
                "department_id": "not-an-id",
            },
        )
        self.assertEqual(invalid_department.status_code, 400)

    def test_result_report_department_filter_uses_latest_non_cancelled_attempt(self):
        latest_department = m.Department.objects.create(name="最新归属部门", level=2)
        latest_contact = m.Contact.objects.create(
            name="最新接口人",
            employee_no="REPORT-LATEST-L2",
            department=latest_department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        moved = self._resume("0101")
        unchanged = self._resume("0102", offset_hours=1)
        moved_workflow = m.CandidateWorkflow.objects.create(
            candidate=moved.candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=moved,
            current_rank=1,
        )
        m.AssignmentAttempt.objects.create(
            workflow=moved_workflow,
            resume=moved,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_REJECTED,
            department=self.department,
            contact=self.contact,
            department_name_snapshot="旧归属部门快照",
        )
        m.AssignmentAttempt.objects.create(
            workflow=moved_workflow,
            resume=moved,
            attempt_no=2,
            source=m.AssignmentAttempt.SOURCE_MANUAL,
            status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
            department=latest_department,
            contact=latest_contact,
            department_name_snapshot="最新归属部门快照",
        )
        m.AssignmentAttempt.objects.create(
            workflow=moved_workflow,
            resume=moved,
            attempt_no=3,
            source=m.AssignmentAttempt.SOURCE_MANUAL,
            status=m.AssignmentAttempt.STATUS_CANCELLED,
            department=self.department,
            contact=self.contact,
            department_name_snapshot="取消尝试部门快照",
        )
        self._attempt(unchanged, m.AssignmentAttempt.STATUS_DISPATCHED_L2)

        response = self.client.get(
            "/api/resumes/result-report/",
            {
                "imported_after": "2026-07-01",
                "imported_before": "2026-07-02",
                "department_id": latest_department.id,
            },
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook["简历明细"].iter_rows(values_only=True))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][2], moved.apply_id)
        self.assertEqual(rows[1][7], "最新归属部门快照")

    def test_result_report_department_filter_uses_current_volunteer_attempt(self):
        historical_department = m.Department.objects.create(
            name="历史志愿归属部门", level=2
        )
        historical_contact = m.Contact.objects.create(
            name="历史志愿接口人",
            employee_no="REPORT-HISTORICAL-L2",
            department=historical_department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        current = self._resume("0201")
        historical = m.Resume.objects.create(
            candidate=current.candidate,
            apply_id="REPORT-0201-HISTORY",
            entity="GW",
            position_name="历史志愿岗位",
            volunteer_rank=2,
        )
        m.Resume.objects.filter(pk=historical.pk).update(
            imported_at=self.start_at + timedelta(hours=1)
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=current.candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=current,
            current_rank=1,
        )
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=historical,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
            department=historical_department,
            contact=historical_contact,
        )
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=current,
            attempt_no=2,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_PENDING_DISPATCH,
            department=self.department,
            contact=self.contact,
        )

        response = self.client.get(
            "/api/resumes/result-report/",
            {
                "imported_after": "2026-07-01",
                "imported_before": "2026-07-02",
                "department_id": historical_department.id,
            },
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook["简历明细"].iter_rows(values_only=True))
        self.assertEqual(len(rows), 1)

    def test_result_report_uses_resume_status_snapshot_and_formula_protection(self):
        unassigned = self._resume("0001", name="=2+2", offset_hours=0)
        pending = self._resume("0002", offset_hours=12)
        passed = self._resume("0003", offset_hours=47)
        excluded = self._resume("0004", offset_hours=48)
        pending_attempt = self._attempt(
            pending, m.AssignmentAttempt.STATUS_ASSIGNED_L3
        )
        pending_attempt.sub_department_name_snapshot = "历史三级部"
        pending_attempt.sub_contact_name_snapshot = "历史三级接口人"
        pending_attempt.save(
            update_fields=["sub_department_name_snapshot", "sub_contact_name_snapshot"]
        )
        m.AssignmentAttempt.objects.create(
            workflow=pending_attempt.workflow,
            resume=pending,
            attempt_no=2,
            source=m.AssignmentAttempt.SOURCE_MANUAL,
            status=m.AssignmentAttempt.STATUS_CANCELLED,
            department=self.department,
            contact=self.contact,
            department_name_snapshot="不应统计的取消部门",
        )
        passed_attempt = self._attempt(passed, m.AssignmentAttempt.STATUS_PASSED)
        passed_attempt.feedback_result = m.AssignmentAttempt.FEEDBACK_PASSED
        passed_attempt.feedback_at = self.start_at + timedelta(hours=48)
        passed_attempt.save(update_fields=["feedback_result", "feedback_at"])
        self._attempt(excluded, m.AssignmentAttempt.STATUS_PENDING_DISPATCH)

        response = self.client.get(
            "/api/resumes/result-report/",
            {"imported_after": "2026-07-01", "imported_before": "2026-07-02"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            quote("简历结果报表_20260701_20260702.xlsx"),
            response["Content-Disposition"],
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["部门汇总", "简历明细"])
        summary_rows = list(workbook["部门汇总"].iter_rows(values_only=True))
        summary_by_department = {row[0]: row for row in summary_rows[1:]}
        self.assertEqual(summary_by_department["历史研发部"][1:4], (2, 2, 0))
        self.assertEqual(summary_by_department["历史研发部"][8:10], (1, 1))
        self.assertEqual(summary_by_department["未分配"][1:4], (1, 0, 1))
        self.assertEqual(summary_by_department["合计"][1], 3)
        detail_rows = list(workbook["简历明细"].iter_rows(values_only=True))
        self.assertEqual(len(detail_rows), 4)
        self.assertNotIn("手机号", detail_rows[0])
        self.assertEqual(detail_rows[1][1], "'=2+2")
        self.assertEqual(detail_rows[2][7], "历史研发部")
        self.assertEqual(detail_rows[2][9], "历史三级部")
        self.assertEqual(detail_rows[3][13], "通过")
        self.assertEqual(unassigned.apply_id, detail_rows[1][2])

    def test_result_report_requires_resume_view_permission(self):
        contact = m.Contact.objects.create(
            name="二级", employee_no="REPORT-NO", contact_level=m.Contact.LEVEL_SECONDARY
        )
        user = User.objects.create_user(
            username="report-secondary",
            password="pass",
            role=User.ROLE_SECONDARY_CONTACT,
            contact=contact,
        )
        user.groups.add(Group.objects.get(name="二级接口人"))
        self.client.force_authenticate(user)

        response = self.client.get(
            "/api/resumes/result-report/",
            {"imported_after": "2026-07-01", "imported_before": "2026-07-02"},
        )

        self.assertEqual(response.status_code, 403)


class ResumePreviewApiTests(TestCase):
    def setUp(self):
        ensure_rbac_defaults()
        self.client = APIClient()
        self.hr = User.objects.create_user(
            username="hr-preview", password="pass", role=User.ROLE_HR
        )
        self.hr.groups.add(Group.objects.get(name="HR"))
        self.client.force_authenticate(self.hr)

    def test_resume_preview_returns_inline_file_content(self):
        with TemporaryDirectory() as media_root:
            resume_dir = Path(media_root) / "resumes"
            resume_dir.mkdir()
            (resume_dir / "张三（A1001）.txt").write_text("resume body", encoding="utf-8")
            candidate = m.Candidate.objects.create(
                identity_hash="candidate-preview",
                name="张三",
                phone="13800000000",
            )
            resume = m.Resume.objects.create(
                candidate=candidate,
                apply_id="A1001",
                position_name="后端工程师",
                resume_file="张三（A1001）.txt",
            )

            with self.settings(MEDIA_ROOT=media_root):
                response = self.client.get(f"/api/resumes/{resume.id}/preview/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.decode("utf-8"), "resume body")
        self.assertIn("inline", response["Content-Disposition"])
        self.assertEqual(response["X-Resume-Filename"], quote("张三（A1001）.txt"))

    def test_resume_preview_returns_pdf_type_and_inline_filename(self):
        with TemporaryDirectory() as media_root:
            resume_dir = Path(media_root) / "resumes"
            resume_dir.mkdir()
            (resume_dir / "张三（A1001）.pdf").write_bytes(b"%PDF-1.4\n%test\n")
            candidate = m.Candidate.objects.create(
                identity_hash="candidate-preview-pdf",
                name="张三",
                phone="13800000000",
            )
            resume = m.Resume.objects.create(
                candidate=candidate,
                apply_id="A1001",
                position_name="后端工程师",
                resume_file="张三（A1001）.pdf",
            )

            with self.settings(MEDIA_ROOT=media_root):
                response = self.client.get(f"/api/resumes/{resume.id}/preview/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("inline", response["Content-Disposition"])
        self.assertIn(
            f"filename*=UTF-8''{quote('张三（A1001）.pdf')}",
            response["Content-Disposition"],
        )
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_resume_preview_returns_404_when_file_is_missing(self):
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-preview-missing",
            name="李四",
            phone="13900000000",
        )
        resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="A1002",
            position_name="产品经理",
        )

        response = self.client.get(f"/api/resumes/{resume.id}/preview/")

        self.assertEqual(response.status_code, 404)

    def test_attempt_resume_preview_uses_attempt_scope(self):
        department = m.Department.objects.create(name="技术二部", level=2)
        contact = m.Contact.objects.create(
            name="技术二级接口人",
            employee_no="S-A",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        user = User.objects.create_user(
            username="secondary-preview",
            password="pass",
            role=User.ROLE_SECONDARY_CONTACT,
            contact=contact,
        )
        user.groups.add(Group.objects.get(name="二级接口人"))

        with TemporaryDirectory() as media_root:
            resume_dir = Path(media_root) / "resumes"
            resume_dir.mkdir()
            (resume_dir / "王五（A1003）.txt").write_text("attempt body", encoding="utf-8")
            candidate = m.Candidate.objects.create(
                identity_hash="candidate-attempt-preview",
                name="王五",
                phone="13700000000",
            )
            resume = m.Resume.objects.create(
                candidate=candidate,
                apply_id="A1003",
                position_name="测试工程师",
                resume_file="王五（A1003）.txt",
            )
            workflow = m.CandidateWorkflow.objects.create(
                candidate=candidate,
                status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
                current_resume=resume,
                current_rank=1,
            )
            attempt = m.AssignmentAttempt.objects.create(
                workflow=workflow,
                resume=resume,
                attempt_no=1,
                source=m.AssignmentAttempt.SOURCE_RULE,
                status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
                department=department,
                contact=contact,
            )

            self.client.force_authenticate(user)
            with self.settings(MEDIA_ROOT=media_root):
                response = self.client.get(
                    f"/api/workflow-attempts/{attempt.id}/resume-preview/"
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.decode("utf-8"), "attempt body")
        self.assertIn("inline", response["Content-Disposition"])

    def test_attempt_resume_preview_hides_unscoped_attempt(self):
        own_department = m.Department.objects.create(name="技术二部", level=2)
        other_department = m.Department.objects.create(name="产品二部", level=2)
        own_contact = m.Contact.objects.create(
            name="技术二级接口人",
            employee_no="S-A",
            department=own_department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        other_contact = m.Contact.objects.create(
            name="产品二级接口人",
            employee_no="S-B",
            department=other_department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        user = User.objects.create_user(
            username="secondary-unscoped-preview",
            password="pass",
            role=User.ROLE_SECONDARY_CONTACT,
            contact=own_contact,
        )
        user.groups.add(Group.objects.get(name="二级接口人"))
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-attempt-unscoped",
            name="赵六",
            phone="13600000000",
        )
        resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="A1004",
            position_name="产品经理",
            resume_file="赵六（A1004）.txt",
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=resume,
            current_rank=1,
        )
        attempt = m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
            department=other_department,
            contact=other_contact,
        )

        self.client.force_authenticate(user)
        response = self.client.get(
            f"/api/workflow-attempts/{attempt.id}/resume-preview/"
        )

        self.assertEqual(response.status_code, 404)


class ImportApiTests(TestCase):
    def setUp(self):
        ensure_rbac_defaults()
        self.client = APIClient()
        self.hr = User.objects.create_user(
            username="hr-import", password="pass", role=User.ROLE_HR
        )
        self.hr.groups.add(Group.objects.get(name="HR"))
        self.client.force_authenticate(self.hr)

    def test_downloads_all_standard_import_templates(self):
        for template_type in ("resume_list", "jobs", "schools", "contacts"):
            with self.subTest(template_type=template_type):
                schema = get_import_table_schema(template_type)
                response = self.client.get(
                    f"/api/import/templates/{template_type}/"
                )

                self.assertEqual(response.status_code, 200)
                self.assertIn(quote(schema.filename), response["Content-Disposition"])
                workbook = load_workbook(BytesIO(response.content))
                sheet = workbook[schema.sheet_name]
                self.assertEqual(
                    [cell.value for cell in sheet[1]],
                    list(schema.headers),
                )
                self.assertEqual(sheet.freeze_panes, "A2")
                self.assertIn("填写说明", workbook.sheetnames)

    def test_unknown_import_template_returns_404(self):
        response = self.client.get("/api/import/templates/unknown/")

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.data["detail"], "未知导入模板类型")

    def test_import_template_requires_import_permission(self):
        user = User.objects.create_user(
            username="template-viewer",
            password="pass",
            role=User.ROLE_SECONDARY_CONTACT,
        )
        self.client.force_authenticate(user)

        response = self.client.get("/api/import/templates/jobs/")

        self.assertEqual(response.status_code, 403)

    def test_job_import_rejects_nonstandard_headers_with_explicit_differences(self):
        schema = get_import_table_schema("jobs")
        row = dict.fromkeys(schema.headers, "")
        row.update(
            {
                "招聘主体": "GW",
                "二层部门": "平台部",
                "对外发布名称": "后端开发",
                "职位名称": "后端工程师",
                "工作职责": "负责后端研发。",
                "HC": 2,
            }
        )
        row["主体"] = row.pop("招聘主体")
        row["二级组织"] = row.pop("二层部门")
        row["需求数量"] = row.pop("HC")
        output = BytesIO()
        pd.DataFrame([row]).to_excel(output, index=False)

        response = self.client.post(
            "/api/import/",
            {
                "jobs": SimpleUploadedFile(
                    "非标准岗位表.xlsx",
                    output.getvalue(),
                    content_type=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                )
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("缺少字段【招聘主体、二层部门、HC】", response.data["detail"])
        self.assertIn("未知字段【主体、二级组织、需求数量】", response.data["detail"])
        self.assertFalse(m.Job.objects.exists())

    @patch("apps.api.views.snapshot.take_snapshot")
    def test_resume_header_validation_runs_before_snapshot(self, take_snapshot):
        output = BytesIO()
        pd.DataFrame([{"姓名": "张三", "应聘ID": "A1001"}]).to_excel(
            output,
            index=False,
        )

        response = self.client.post(
            "/api/import/",
            {
                "processing_mode": "rule",
                "resume_list": SimpleUploadedFile(
                    "非标准简历表.xlsx",
                    output.getvalue(),
                    content_type=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                ),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("简历信息列表表头不符合标准模板", response.data["detail"])
        take_snapshot.assert_not_called()

    @patch("apps.api.views.snapshot.take_snapshot")
    @patch("apps.api.views.import_files")
    def test_resume_upload_requires_processing_mode(
        self,
        mock_import_files,
        mock_take_snapshot,
    ):
        response = self.client.post(
            "/api/import/",
            {"resume_package": SimpleUploadedFile("简历包.zip", b"zip")},
            format="multipart",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("分配模式必须是 rule 或 ai", response.data["detail"])
        mock_take_snapshot.assert_not_called()
        mock_import_files.assert_not_called()

    @patch("apps.api.views.snapshot.take_snapshot")
    @patch("apps.api.views.import_files")
    def test_resume_upload_rejects_unavailable_ai_mode(
        self,
        mock_import_files,
        mock_take_snapshot,
    ):
        response = self.client.post(
            "/api/import/",
            {
                "processing_mode": "ai",
                "resume_package": SimpleUploadedFile("简历包.zip", b"zip"),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("模型连接尚未测试成功", response.data["detail"])
        mock_take_snapshot.assert_not_called()
        mock_import_files.assert_not_called()

    @patch("apps.api.views.import_files")
    def test_job_import_returns_skipped_responsibility_warning(self, mock_import_files):
        mock_import_files.return_value = {
            "jobs": 1,
            "jobs_skipped": 2,
            "_warnings": [
                {
                    "code": "job_responsibility_missing",
                    "count": 2,
                    "rows": [3, 5],
                    "message": "工作职责为空的岗位已跳过",
                }
            ],
        }

        response = self.client.post(
            "/api/import/",
            {
                "jobs": SimpleUploadedFile(
                    "岗位.xlsx",
                    standard_import_template_bytes("jobs"),
                    content_type=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                )
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["counts"]["jobs"], 1)
        self.assertEqual(response.data["counts"]["jobs_skipped"], 2)
        self.assertEqual(
            response.data["detail"], "导入完成，已跳过 2 条缺少工作职责的岗位"
        )
        self.assertEqual(
            response.data["warnings"][0],
            {
                "code": "job_responsibility_missing",
                "count": 2,
                "rows": [3, 5],
                "message": "工作职责为空的岗位已跳过",
            },
        )

    @patch("apps.api.views.execute_runs_sequence_task.delay")
    @patch("apps.api.views.runner.create_run")
    @patch("apps.api.views.snapshot.take_snapshot")
    @patch("apps.api.views.import_files")
    def test_resume_upload_starts_one_run_with_selected_mode(
        self,
        mock_import_files,
        mock_take_snapshot,
        mock_create_run,
        mock_execute,
    ):
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-upload-modes", name="上传候选人", phone="13800000001"
        )
        ai_run = m.ProcessingRun.objects.create(step="resume_process", mode="ai")
        mock_import_files.return_value = {
            "candidates_created": 1,
            "candidates_updated": 0,
            "resumes_created": 1,
            "resumes_updated": 0,
            "_candidate_ids": [candidate.id],
        }
        mock_create_run.return_value = ai_run
        mock_execute.return_value = SimpleNamespace(id="upload-modes")

        with patch(
            "apps.api.views.ai_config.validate_allocation_mode", return_value="ai"
        ):
            response = self.client.post(
                "/api/import/",
                {
                    "processing_mode": "ai",
                    "resume_package": SimpleUploadedFile("简历包.zip", b"zip"),
                },
                format="multipart",
            )

        self.assertEqual(response.status_code, 202)
        mock_create_run.assert_called_once_with(
            "resume_process",
            mode="ai",
            scope={"candidate_ids": [candidate.id], "source": "resume_import"},
            created_by=self.hr,
        )
        mock_execute.assert_called_once_with([ai_run.id])
        self.assertEqual([item["mode"] for item in response.data["processing_runs"]], ["ai"])

    def test_replace_contacts_import_keeps_existing_resume_pool(self):
        candidate = m.Candidate.objects.create(
            identity_hash="candidate-import",
            name="张三",
            phone="13800000000",
        )
        m.Resume.objects.create(
            candidate=candidate,
            apply_id="A1001",
            position_name="后端工程师",
        )
        old_department = m.Department.objects.create(name="旧部门", level=2)
        old_contact = m.Contact.objects.create(
            name="旧接口人",
            employee_no="OLD001",
            department=old_department,
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
        )
        m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=m.Resume.objects.get(apply_id="A1001"),
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_PENDING_DISPATCH,
            department=old_department,
            contact=old_contact,
        )
        buf = BytesIO()
        pd.DataFrame(
            [
                {
                    "工号": "NEW001",
                    "邮箱": "new001@example.com",
                    "姓名": "新接口人",
                    "一层部门": "技术中心",
                    "二层部门": "后端组",
                    "接口人层级": "二级接口人",
                }
            ],
            columns=get_import_table_schema("contacts").headers,
        ).to_excel(buf, index=False)
        buf.seek(0)

        response = self.client.post(
            "/api/import/",
            {
                "mode": "replace",
                "contacts": SimpleUploadedFile(
                    "部门接口人信息.xlsx",
                    buf.getvalue(),
                    content_type=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                ),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(m.Candidate.objects.count(), 1)
        self.assertEqual(m.Resume.objects.count(), 1)
        self.assertEqual(m.AssignmentAttempt.objects.count(), 1)
        self.assertFalse(m.Contact.objects.get(employee_no="OLD001").is_active)
        self.assertTrue(m.Contact.objects.get(employee_no="NEW001").is_active)

    def test_replace_contacts_import_deactivates_old_contacts_and_users(self):
        old_department = m.Department.objects.create(name="旧部门", level=2)
        old_contact = m.Contact.objects.create(
            name="旧接口人",
            employee_no="OLD002",
            department=old_department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        old_user = User.objects.create_user(
            username="OLD002",
            password="pass",
            role=User.ROLE_SECONDARY_CONTACT,
            contact=old_contact,
        )
        old_user.groups.add(Group.objects.get(name="二级接口人"))
        buf = BytesIO()
        pd.DataFrame(
            [
                {
                    "工号": "NEW002",
                    "邮箱": "new002@example.com",
                    "姓名": "新接口人",
                    "一层部门": "技术中心",
                    "二层部门": "后端组",
                    "接口人层级": "二级接口人",
                }
            ],
            columns=get_import_table_schema("contacts").headers,
        ).to_excel(buf, index=False)
        buf.seek(0)

        response = self.client.post(
            "/api/import/",
            {
                "mode": "replace",
                "contacts": SimpleUploadedFile(
                    "部门接口人信息.xlsx",
                    buf.getvalue(),
                    content_type=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                ),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        old_contact.refresh_from_db()
        self.assertFalse(old_contact.is_active)
        old_user.refresh_from_db()
        self.assertFalse(old_user.is_active)
        self.assertEqual(old_user.contact_id, old_contact.id)

    @patch("apps.api.views.import_files")
    @patch("apps.api.views.snapshot.take_snapshot")
    def test_large_resume_package_reaches_import_service(
        self, mock_take_snapshot, mock_import_files
    ):
        mock_import_files.return_value = {
            "candidates_created": 0,
            "candidates_updated": 0,
            "resumes_created": 0,
            "resumes_updated": 0,
        }
        large_package = SimpleUploadedFile(
            "简历包.zip",
            b"x" * (4 * 1024 * 1024),
            content_type="application/zip",
        )

        response = self.client.post(
            "/api/import/",
            {
                "mode": "incremental",
                "processing_mode": "rule",
                "resume_package": large_package,
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        mock_take_snapshot.assert_called_once_with(label="上传简历前")
        mock_import_files.assert_called_once()


class MasterDataCrudApiTests(TestCase):
    def setUp(self):
        ensure_rbac_defaults()
        self.client = APIClient()
        self.admin = User.objects.create_user(
            username="admin-master-data", password="pass", role=User.ROLE_ADMIN
        )
        self.admin.groups.add(Group.objects.get(name="管理员"))
        self.client.force_authenticate(self.admin)

    def test_user_management_requires_unique_email(self):
        missing = self.client.post(
            "/api/users/",
            {"username": "NO-MAIL"},
            format="json",
        )
        created = self.client.post(
            "/api/users/",
            {
                "username": "MAIL001",
                "email": "MAIL001@EXAMPLE.COM",
            },
            format="json",
        )
        duplicate = self.client.post(
            "/api/users/",
            {
                "username": "MAIL002",
                "email": "mail001@example.com",
            },
            format="json",
        )

        self.assertEqual(missing.status_code, 400)
        self.assertIn("email", missing.data)
        self.assertEqual(created.status_code, 201)
        self.assertEqual(created.data["email"], "mail001@example.com")
        self.assertNotIn("password", created.data)
        self.assertFalse(User.objects.get(username="MAIL001").has_usable_password())
        self.assertEqual(duplicate.status_code, 400)
        self.assertIn("email", duplicate.data)

    def test_job_create_and_update_accepts_valid_tertiary_but_rejects_primary_or_orphan(self):
        primary = m.Department.objects.create(name="技术中心", level=1)
        secondary = m.Department.objects.create(
            name="平台部", level=2, parent=primary
        )
        tertiary = m.Department.objects.create(
            name="研发组", level=3, parent=secondary
        )
        orphan = m.Department.objects.create(
            name="错误三级", level=3, parent=primary
        )

        created = self.client.post(
            "/api/jobs/",
            {
                "entity": "GW",
                "department": secondary.id,
                "public_name": "后端开发",
                "position_name": "后端工程师",
                "category": "技术类",
                "responsibilities": "负责后端研发。",
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201)

        updated = self.client.patch(
            f"/api/jobs/{created.data['id']}/",
            {"department": tertiary.id},
            format="json",
        )
        self.assertEqual(updated.status_code, 200)
        self.assertEqual(updated.data["department"], tertiary.id)
        self.assertEqual(updated.data["department_name"], "平台部")
        self.assertEqual(updated.data["tertiary_department_name"], "研发组")

        primary_response = self.client.patch(
            f"/api/jobs/{created.data['id']}/",
            {"department": primary.id},
            format="json",
        )
        orphan_response = self.client.patch(
            f"/api/jobs/{created.data['id']}/",
            {"department": orphan.id},
            format="json",
        )
        self.assertEqual(primary_response.status_code, 400)
        self.assertIn("department", primary_response.data)
        self.assertEqual(orphan_response.status_code, 400)
        self.assertIn("department", orphan_response.data)

    def test_user_management_rejects_password_on_create_and_update(self):
        create_response = self.client.post(
            "/api/users/",
            {
                "username": "PASSWORD001",
                "email": "password001@example.com",
                "password": "forbidden",
            },
            format="json",
        )
        user = User.objects.create_user(
            username="PASSWORD002",
            email="password002@example.com",
        )
        update_response = self.client.patch(
            f"/api/users/{user.id}/",
            {"password": "forbidden"},
            format="json",
        )

        self.assertEqual(create_response.status_code, 400)
        self.assertIn("password", create_response.data)
        self.assertEqual(update_response.status_code, 400)
        self.assertIn("password", update_response.data)

    def test_contact_cannot_claim_protected_administrator_identity(self):
        department = m.Department.objects.create(name="受保护账号测试部门", level=2)

        response = self.client.post(
            "/api/contacts/",
            {
                "name": "伪造接口人",
                "employee_no": PROTECTED_ADMIN_USERNAME,
                "email": PROTECTED_ADMIN_EMAIL,
                "department": department.id,
                "is_active": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(
            m.Contact.objects.filter(employee_no=PROTECTED_ADMIN_USERNAME).exists()
        )
        protected_admin = User.objects.get(username=PROTECTED_ADMIN_USERNAME)
        self.assertIsNone(protected_admin.contact_id)
        self.assertTrue(protected_admin.is_superuser)

    def test_create_and_update_school_keeps_platform_in_sync_with_tag(self):
        first_tag = m.SchoolTag.objects.create(code="FIRST", name="第一标签")
        second_tag = m.SchoolTag.objects.create(code="SECOND", name="第二标签")

        create_response = self.client.post(
            "/api/schools/",
            {
                "name": "测试大学",
                "province": "湖北",
                "school_tag": first_tag.id,
            },
            format="json",
        )

        self.assertEqual(create_response.status_code, 201)
        school = m.School.objects.get(name="测试大学")
        self.assertEqual(school.school_tag_id, first_tag.id)
        self.assertEqual(school.platform, "第一标签")

        update_response = self.client.patch(
            f"/api/schools/{school.id}/",
            {"province": "湖南", "school_tag": second_tag.id},
            format="json",
        )

        self.assertEqual(update_response.status_code, 200)
        school.refresh_from_db()
        self.assertEqual(school.province, "湖南")
        self.assertEqual(school.school_tag_id, second_tag.id)
        self.assertEqual(school.platform, "第二标签")

        ignored_platform_response = self.client.patch(
            f"/api/schools/{school.id}/",
            {"platform": "手工篡改值"},
            format="json",
        )

        self.assertEqual(ignored_platform_response.status_code, 200)
        school.refresh_from_db()
        self.assertEqual(school.platform, "第二标签")

    def test_create_and_update_contact_syncs_bound_login_account(self):
        secondary = m.Department.objects.create(name="二级部门", level=2)
        tertiary = m.Department.objects.create(
            name="三级部门", level=3, parent=secondary
        )

        create_response = self.client.post(
            "/api/contacts/",
            {
                "name": "新接口人",
                "employee_no": "NEW100",
                "email": "new100@example.com",
                "department": tertiary.id,
                "contact_level": m.Contact.LEVEL_SECONDARY,
                "can_delegate": True,
                "is_active": True,
            },
            format="json",
        )

        self.assertEqual(create_response.status_code, 201)
        contact = m.Contact.objects.get(employee_no="NEW100")
        self.assertEqual(contact.email, "new100@example.com")
        self.assertEqual(contact.contact_level, m.Contact.LEVEL_TERTIARY)
        self.assertFalse(contact.can_delegate)
        user = User.objects.get(username="NEW100")
        self.assertEqual(user.contact_id, contact.id)
        self.assertEqual(user.email, "new100@example.com")
        self.assertEqual(user.role, User.ROLE_TERTIARY_CONTACT)
        self.assertFalse(user.has_usable_password())
        self.assertIn("三级接口人", user.groups.values_list("name", flat=True))

        extra_group = Group.objects.create(name="保留的额外角色")
        user.groups.add(extra_group)

        update_response = self.client.patch(
            f"/api/contacts/{contact.id}/",
            {
                "name": "修改后接口人",
                "employee_no": "NEW101",
                "email": "new101@example.com",
                "department": secondary.id,
                "can_delegate": True,
                "is_active": False,
            },
            format="json",
        )

        self.assertEqual(update_response.status_code, 200)
        contact.refresh_from_db()
        user.refresh_from_db()
        self.assertEqual(contact.employee_no, "NEW101")
        self.assertEqual(contact.email, "new101@example.com")
        self.assertEqual(contact.contact_level, m.Contact.LEVEL_SECONDARY)
        self.assertTrue(contact.can_delegate)
        self.assertEqual(user.username, "NEW101")
        self.assertEqual(user.email, "new101@example.com")
        self.assertEqual(user.role, User.ROLE_SECONDARY_CONTACT)
        self.assertFalse(user.is_active)
        self.assertFalse(user.has_usable_password())
        self.assertIn("二级接口人", user.groups.values_list("name", flat=True))
        self.assertNotIn("三级接口人", user.groups.values_list("name", flat=True))
        self.assertIn("保留的额外角色", user.groups.values_list("name", flat=True))

        reactivate_response = self.client.patch(
            f"/api/contacts/{contact.id}/",
            {"is_active": True},
            format="json",
        )

        self.assertEqual(reactivate_response.status_code, 200)
        contact.refresh_from_db()
        user.refresh_from_db()
        self.assertTrue(contact.is_active)
        self.assertTrue(user.is_active)

    def test_create_contact_requires_email(self):
        department = m.Department.objects.create(name="缺邮箱部门", level=2)

        response = self.client.post(
            "/api/contacts/",
            {
                "name": "缺邮箱接口人",
                "employee_no": "NOEMAIL001",
                "department": department.id,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("email", response.data)
        self.assertFalse(
            m.Contact.objects.filter(employee_no="NOEMAIL001").exists()
        )

    def test_view_only_user_cannot_create_school_or_contact(self):
        viewer = User.objects.create_user(username="master-data-viewer", password="pass")
        viewer.user_permissions.add(
            Permission.objects.get(codename=permission_codename("school.view")),
            Permission.objects.get(codename=permission_codename("department.view")),
        )
        self.client.force_authenticate(viewer)
        department = m.Department.objects.create(name="只读部门", level=2)

        school_response = self.client.post(
            "/api/schools/", {"name": "无权新增大学"}, format="json"
        )
        contact_response = self.client.post(
            "/api/contacts/",
            {
                "name": "无权新增接口人",
                "employee_no": "VIEW001",
                "department": department.id,
            },
            format="json",
        )

        self.assertEqual(school_response.status_code, 403)
        self.assertEqual(contact_response.status_code, 403)
        self.assertFalse(m.School.objects.filter(name="无权新增大学").exists())
        self.assertFalse(m.Contact.objects.filter(employee_no="VIEW001").exists())


class ContactDeleteApiTests(TestCase):
    def setUp(self):
        ensure_rbac_defaults()
        self.client = APIClient()
        self.admin = User.objects.create_user(
            username="admin-contact-delete", password="pass", role=User.ROLE_ADMIN
        )
        self.admin.groups.add(Group.objects.get(name="管理员"))
        self.client.force_authenticate(self.admin)

    def test_delete_contact_removes_contact_and_bound_user(self):
        department = m.Department.objects.create(name="技术二部", level=2)
        contact = m.Contact.objects.create(
            name="待删除接口人",
            employee_no="DEL001",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        user = User.objects.create_user(
            username="DEL001",
            password="pass",
            role=User.ROLE_SECONDARY_CONTACT,
            contact=contact,
        )
        user.groups.add(Group.objects.get(name="二级接口人"))

        response = self.client.delete(f"/api/contacts/{contact.id}/")

        self.assertEqual(response.status_code, 204)
        self.assertFalse(m.Contact.objects.filter(id=contact.id).exists())
        self.assertFalse(User.objects.filter(id=user.id).exists())

    def test_delete_contact_keeps_history_and_clears_contact_references(self):
        department = m.Department.objects.create(name="技术二部", level=2)
        contact = m.Contact.objects.create(
            name="有历史接口人",
            employee_no="DEL002",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        candidate = m.Candidate.objects.create(
            identity_hash="contact-delete-history",
            name="张三",
            phone="13800000000",
        )
        resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="A1001",
            position_name="后端工程师",
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=resume,
            current_rank=1,
        )
        attempt = m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
            department=department,
            contact=contact,
            contact_name_snapshot=contact.name,
            contact_employee_no_snapshot=contact.employee_no,
            created_by=self.admin,
        )
        handoff = m.AssignmentHandoff.objects.create(
            attempt=attempt,
            action=m.AssignmentHandoff.ACTION_HR_DISPATCH,
            to_department=department,
            to_contact=contact,
            to_department_name_snapshot=department.name,
            to_contact_name_snapshot=contact.name,
            to_contact_employee_no_snapshot=contact.employee_no,
            created_by_username_snapshot=self.admin.username,
            created_by=self.admin,
        )
        decision = m.AgentDispatchDecision.objects.create(
            workflow=workflow,
            resume=resume,
            recommendation=m.AgentDispatchDecision.RECOMMEND_DISPATCH,
            recommended_contact=contact,
            recommended_contact_name_snapshot=contact.name,
            recommended_contact_employee_no_snapshot=contact.employee_no,
        )

        response = self.client.delete(f"/api/contacts/{contact.id}/")

        self.assertEqual(response.status_code, 204)
        self.assertFalse(m.Contact.objects.filter(id=contact.id).exists())
        attempt.refresh_from_db()
        handoff.refresh_from_db()
        decision.refresh_from_db()
        self.assertIsNone(attempt.contact_id)
        self.assertEqual(attempt.contact_name_snapshot, "有历史接口人")
        self.assertEqual(attempt.contact_employee_no_snapshot, "DEL002")
        self.assertIsNone(handoff.to_contact_id)
        self.assertEqual(handoff.to_contact_name_snapshot, "有历史接口人")
        self.assertEqual(handoff.to_contact_employee_no_snapshot, "DEL002")
        self.assertEqual(handoff.to_department_name_snapshot, "技术二部")
        self.assertEqual(handoff.created_by_username_snapshot, "admin-contact-delete")
        self.assertIsNone(decision.recommended_contact_id)
        self.assertEqual(decision.recommended_contact_name_snapshot, "有历史接口人")
        self.assertEqual(decision.recommended_contact_employee_no_snapshot, "DEL002")

    def test_contacts_list_defaults_to_active_contacts(self):
        department = m.Department.objects.create(name="技术二部", level=2)
        active = m.Contact.objects.create(
            name="启用接口人",
            employee_no="LIST001",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
            is_active=True,
        )
        m.Contact.objects.create(
            name="停用接口人",
            employee_no="LIST002",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
            is_active=False,
        )

        response = self.client.get("/api/contacts/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual([item["id"] for item in response.data["results"]], [active.id])

    def test_contacts_list_can_filter_inactive_contacts(self):
        department = m.Department.objects.create(name="技术二部", level=2)
        m.Contact.objects.create(
            name="启用接口人",
            employee_no="LIST003",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
            is_active=True,
        )
        inactive = m.Contact.objects.create(
            name="停用接口人",
            employee_no="LIST004",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
            is_active=False,
        )

        response = self.client.get("/api/contacts/", {"is_active": "false"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [item["id"] for item in response.data["results"]], [inactive.id]
        )


class UserDeleteApiTests(TestCase):
    def setUp(self):
        ensure_rbac_defaults()
        self.client = APIClient()
        self.admin = User.objects.create_user(
            username="admin-user-delete", password="pass", role=User.ROLE_ADMIN
        )
        self.admin.groups.add(Group.objects.get(name="管理员"))
        self.client.force_authenticate(self.admin)

    def test_delete_user_removes_record_and_token(self):
        user = User.objects.create_user(
            username="E9001", password="ignored", role=User.ROLE_SECONDARY_CONTACT
        )
        token = Token.objects.create(user=user)

        response = self.client.delete(f"/api/users/{user.id}/")

        self.assertEqual(response.status_code, 204)
        self.assertFalse(User.objects.filter(id=user.id).exists())
        self.assertFalse(Token.objects.filter(key=token.key).exists())

    def test_protected_administrator_cannot_be_edited_disabled_or_deleted(self):
        protected_admin = User.objects.get(username=PROTECTED_ADMIN_USERNAME)

        update_response = self.client.patch(
            f"/api/users/{protected_admin.id}/",
            {
                "username": "changed",
                "email": "changed@example.com",
                "is_active": False,
            },
            format="json",
        )
        delete_response = self.client.delete(f"/api/users/{protected_admin.id}/")

        self.assertEqual(update_response.status_code, 403)
        self.assertEqual(delete_response.status_code, 403)
        protected_admin.refresh_from_db()
        self.assertEqual(protected_admin.username, PROTECTED_ADMIN_USERNAME)
        self.assertEqual(protected_admin.email, PROTECTED_ADMIN_EMAIL)
        self.assertTrue(protected_admin.is_active)
        self.assertTrue(protected_admin.is_staff)
        self.assertTrue(protected_admin.is_superuser)
        self.assertFalse(protected_admin.has_usable_password())
        self.assertEqual(
            list(protected_admin.groups.values_list("name", flat=True)),
            ["管理员"],
        )

        detail_response = self.client.get(f"/api/users/{protected_admin.id}/")
        self.assertEqual(detail_response.status_code, 200)
        self.assertTrue(detail_response.data["is_protected"])

    def test_deleted_user_stays_deleted_and_password_login_route_is_absent(self):
        user = User.objects.create_user(
            username="E9002", password="ignored", role=User.ROLE_SECONDARY_CONTACT
        )

        self.client.delete(f"/api/users/{user.id}/")
        response = self.client.post(
            "/api/auth/login/", {"username": "E9002", "password": "ignored"}, format="json"
        )

        self.assertFalse(User.objects.filter(username="E9002").exists())
        self.assertEqual(response.status_code, 404)

    def test_delete_user_removes_bound_contact_and_clears_history_references(self):
        department = m.Department.objects.create(name="技术二部", level=2)
        contact = m.Contact.objects.create(
            name="绑定接口人",
            employee_no="E9003",
            department=department,
            contact_level=m.Contact.LEVEL_SECONDARY,
        )
        user = User.objects.create_user(
            username="E9003",
            password="ignored",
            role=User.ROLE_SECONDARY_CONTACT,
            contact=contact,
        )
        user.groups.add(Group.objects.get(name="二级接口人"))
        candidate = m.Candidate.objects.create(
            identity_hash="user-delete-history",
            name="李四",
            phone="13900000000",
        )
        resume = m.Resume.objects.create(
            candidate=candidate,
            apply_id="A1002",
            position_name="算法工程师",
        )
        workflow = m.CandidateWorkflow.objects.create(
            candidate=candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=resume,
            current_rank=1,
        )
        attempt = m.AssignmentAttempt.objects.create(
            workflow=workflow,
            resume=resume,
            attempt_no=1,
            source=m.AssignmentAttempt.SOURCE_RULE,
            status=m.AssignmentAttempt.STATUS_DISPATCHED_L2,
            department=department,
            contact=contact,
            contact_name_snapshot=contact.name,
            contact_employee_no_snapshot=contact.employee_no,
            created_by_username_snapshot=user.username,
            created_by=user,
        )
        handoff = m.AssignmentHandoff.objects.create(
            attempt=attempt,
            action=m.AssignmentHandoff.ACTION_HR_DISPATCH,
            to_department=department,
            to_contact=contact,
            to_department_name_snapshot=department.name,
            to_contact_name_snapshot=contact.name,
            to_contact_employee_no_snapshot=contact.employee_no,
            created_by_username_snapshot=user.username,
            created_by=user,
        )

        response = self.client.delete(f"/api/users/{user.id}/")

        self.assertEqual(response.status_code, 204)
        self.assertFalse(User.objects.filter(id=user.id).exists())
        self.assertFalse(m.Contact.objects.filter(id=contact.id).exists())
        attempt.refresh_from_db()
        handoff.refresh_from_db()
        self.assertIsNone(attempt.contact_id)
        self.assertIsNone(attempt.created_by_id)
        self.assertEqual(attempt.contact_name_snapshot, "绑定接口人")
        self.assertEqual(attempt.contact_employee_no_snapshot, "E9003")
        self.assertEqual(attempt.created_by_username_snapshot, "E9003")
        self.assertIsNone(handoff.to_contact_id)
        self.assertIsNone(handoff.created_by_id)
        self.assertEqual(handoff.to_contact_name_snapshot, "绑定接口人")
        self.assertEqual(handoff.to_contact_employee_no_snapshot, "E9003")
        self.assertEqual(handoff.created_by_username_snapshot, "E9003")
