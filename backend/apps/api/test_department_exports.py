"""部门收件箱导出与结果报表的聚焦回归测试。"""

from datetime import datetime, timedelta, timezone as datetime_timezone
from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook

from apps.core import models as m

from .result_report import build_result_report, current_effective_attempt
from .resume_export import (
    CandidateExportRecord,
    ExportFieldError,
    build_resume_export_workbook,
    export_fields_payload,
    parse_export_fields,
)


class DepartmentExportTests(TestCase):
    def setUp(self):
        self.primary = m.Department.objects.create(name="研发体系", level=1)
        self.secondary = m.Department.objects.create(
            name="平台研发部", level=2, parent=self.primary
        )
        self.tertiary = m.Department.objects.create(
            name="云平台组", level=3, parent=self.secondary
        )
        self.candidate = m.Candidate.objects.create(
            identity_hash="department-export-candidate",
            name="导出候选人",
            phone="13800000001",
            highest_education=m.Candidate.EDUCATION_MASTER,
        )
        self.resume = m.Resume.objects.create(
            candidate=self.candidate,
            apply_id="DEPARTMENT-EXPORT-1",
            entity="研发中心",
            position_name="平台工程师",
            volunteer_rank=1,
        )
        self.workflow = m.CandidateWorkflow.objects.create(
            candidate=self.candidate,
            status=m.CandidateWorkflow.STATUS_IN_PROGRESS,
            current_resume=self.resume,
            current_rank=1,
        )
        self.attempt = m.AssignmentAttempt.objects.create(
            workflow=self.workflow,
            resume=self.resume,
            attempt_no=1,
            status=m.AssignmentAttempt.STATUS_REJECTED,
            initial_department=self.secondary,
            current_department=self.tertiary,
            feedback_result=m.AssignmentAttempt.FEEDBACK_REJECTED,
            feedback_reason_code=(
                m.AssignmentAttempt.REJECTION_REASON_MAJOR_BACKGROUND_MISMATCH
            ),
            feedback_note="缺少岗位要求的专业基础",
        )
        self.started_at = datetime(2026, 8, 1, tzinfo=datetime_timezone.utc)
        m.AssignmentAttempt.objects.filter(pk=self.attempt.pk).update(
            created_at=self.started_at
        )
        self.attempt.refresh_from_db()
        self._event(
            m.AssignmentHandlingEvent.EVENT_ATTEMPT_CREATED,
            self.started_at,
        )
        self._event(
            m.AssignmentHandlingEvent.EVENT_DEPARTMENT_DISPATCHED,
            self.started_at + timedelta(hours=2),
            to_department=self.secondary,
        )
        self._event(
            m.AssignmentHandlingEvent.EVENT_DEPARTMENT_TRANSFERRED,
            self.started_at + timedelta(hours=2),
            from_department=self.secondary,
            to_department=self.tertiary,
            is_system_auto=True,
        )
        self._event(
            m.AssignmentHandlingEvent.EVENT_FEEDBACK_REJECTED,
            self.started_at + timedelta(hours=6),
            from_department=self.tertiary,
        )

    def _event(self, event_type, occurred_at, **kwargs):
        event = m.AssignmentHandlingEvent.objects.create(
            attempt=self.attempt,
            event_type=event_type,
            **kwargs,
        )
        m.AssignmentHandlingEvent.objects.filter(pk=event.pk).update(
            occurred_at=occurred_at
        )
        event.refresh_from_db()
        return event

    def test_catalog_breaks_old_contact_fields_and_increments_version(self):
        payload = export_fields_payload()
        keys = {
            field["key"]
            for group in payload["groups"]
            for field in group["fields"]
        }

        self.assertEqual(payload["version"], 4)
        self.assertTrue(
            {
                "initial_department",
                "current_primary_department",
                "current_department",
                "feedback_reason",
                "hr_dispatch_duration_hours",
                "current_department_duration_hours",
                "total_feedback_duration_hours",
            }.issubset(keys)
        )
        self.assertTrue(
            {
                "secondary_contact",
                "tertiary_contact",
                "allocation_secondary_department",
            }.isdisjoint(keys)
        )
        with self.assertRaises(ExportFieldError):
            parse_export_fields({"fields": "secondary_contact"})

    def test_resume_export_uses_department_and_event_timings(self):
        fields = [
            "candidate_name",
            "initial_department",
            "current_primary_department",
            "current_department",
            "feedback_reason_code",
            "feedback_reason",
            "first_dispatched_at",
            "current_department_entered_at",
            "feedback_at",
            "hr_dispatch_duration_hours",
            "current_department_duration_hours",
            "total_feedback_duration_hours",
        ]
        record = CandidateExportRecord(
            candidate=self.candidate,
            current_resume=self.resume,
            attempt=self.attempt,
            file_resumes=[self.resume],
        )

        workbook = load_workbook(
            BytesIO(build_resume_export_workbook([record], fields))
        )
        sheet = workbook["简历库"]
        values = dict(
            zip(
                [cell.value for cell in sheet[1]],
                [cell.value for cell in sheet[2]],
            )
        )

        self.assertEqual(values["首次部门"], self.secondary.name)
        self.assertEqual(values["当前接收一级部门"], self.primary.name)
        self.assertEqual(values["当前接收部门"], self.tertiary.name)
        self.assertEqual(values["不通过原因"], "专业背景不匹配")
        self.assertEqual(values["HR 下发时长（小时）"], 2)
        self.assertEqual(values["当前部门处理时长（小时）"], 4)
        self.assertEqual(values["总反馈时长（小时）"], 4)

    def test_result_report_groups_l3_node_by_secondary_and_reason_code(self):
        self.resume.report_attempts = [self.attempt]

        workbook = load_workbook(BytesIO(build_result_report([self.resume])))

        self.assertEqual(
            workbook.sheetnames,
            ["部门汇总", "简历明细", "不通过原因汇总"],
        )
        summary = list(workbook["部门汇总"].iter_rows(values_only=True))
        self.assertEqual(summary[1][0:2], (self.primary.name, self.secondary.name))
        self.assertEqual(summary[1][2:4], (1, 1))
        details = list(workbook["简历明细"].iter_rows(values_only=True))
        detail = dict(zip(details[0], details[1]))
        self.assertEqual(detail["当前接收二级部门"], self.secondary.name)
        self.assertEqual(detail["当前接收节点"], self.tertiary.name)
        reasons = list(
            workbook["不通过原因汇总"].iter_rows(values_only=True)
        )
        self.assertEqual(
            reasons[1],
            (
                self.primary.name,
                self.secondary.name,
                m.AssignmentAttempt.REJECTION_REASON_MAJOR_BACKGROUND_MISMATCH,
                "专业背景不匹配",
                1,
            ),
        )

    def test_result_report_uses_current_resume_latest_non_cancelled_attempt(self):
        historical_resume = m.Resume.objects.create(
            candidate=self.candidate,
            apply_id="DEPARTMENT-EXPORT-HISTORICAL",
            position_name="历史志愿岗位",
            volunteer_rank=2,
        )
        historical_attempt = m.AssignmentAttempt.objects.create(
            workflow=self.workflow,
            resume=historical_resume,
            attempt_no=2,
            status=m.AssignmentAttempt.STATUS_DISPATCHED,
            initial_department=self.secondary,
            current_department=self.secondary,
        )
        cancelled = m.AssignmentAttempt.objects.create(
            workflow=self.workflow,
            resume=self.resume,
            attempt_no=3,
            status=m.AssignmentAttempt.STATUS_CANCELLED,
            initial_department=self.secondary,
            current_department=self.secondary,
        )
        self.resume.report_attempts = [self.attempt, cancelled]
        historical_resume.report_attempts = [historical_attempt]

        effective = current_effective_attempt(historical_resume)

        self.assertEqual(effective.id, self.attempt.id)

        self.workflow.current_resume = None
        self.workflow.save(update_fields=["current_resume"])
        self.candidate.refresh_from_db()
        historical_resume.candidate = self.candidate
        effective_without_pointer = current_effective_attempt(historical_resume)

        self.assertEqual(effective_without_pointer.id, self.attempt.id)
