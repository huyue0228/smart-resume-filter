"""生成“字段清单 + 简历文件”的候选人级组合导出包。"""

from dataclasses import dataclass
from io import BytesIO
import os
from typing import Optional
from zoneinfo import ZoneInfo
import zipfile

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.core import candidate_summary
from apps.core import models as m
from apps.core import system_status
from apps.ingestion.sources import RESUME_SUBDIR


EXPORT_FIELDS_VERSION = 2
DEFAULT_FIELD_KEYS = {
    "candidate_name",
    "candidate_phone",
    "current_apply_id",
    "current_position_name",
    "volunteer_rank",
    "allocation_secondary_department",
    "secondary_contact",
    "tertiary_contact",
    "allocation_source",
    "resume_status",
}


def _field(key, label):
    return {"key": key, "label": label, "default_selected": key in DEFAULT_FIELD_KEYS}


EXPORT_FIELD_GROUPS = [
    {
        "key": "candidate",
        "label": "候选人",
        "fields": [
            _field("candidate_name", "姓名"),
            _field("candidate_phone", "手机号"),
            _field("gender", "性别"),
            _field("household_province", "户口所在地"),
            _field("highest_education", "最高学历"),
            _field("highest_major", "最高学历专业"),
            _field("first_degree_school", "第一学历院校"),
            _field("first_degree_tag", "第一学历院校标签"),
            _field("highest_degree_school", "最高学历院校"),
            _field("highest_degree_tag", "最高学历院校标签"),
            _field("candidate_imported_at", "候选人导入时间"),
        ],
    },
    {
        "key": "current_resume",
        "label": "当前投递",
        "fields": [
            _field("current_apply_id", "当前应聘ID"),
            _field("entity", "主体"),
            _field("org", "所属机构"),
            _field("current_position_name", "当前岗位"),
            _field("original_status", "原始应聘状态"),
            _field("apply_date", "应聘日期"),
            _field("volunteer_rank", "当前志愿"),
            _field("assigned_entity", "分配主体"),
            _field("job_category", "岗位类别"),
            _field("category_mode", "分类模式"),
            _field("category_reason", "分类理由"),
            _field("resume_filename", "简历文件名"),
            _field("all_apply_ids", "全部应聘ID"),
            _field("all_resume_filenames", "全部简历文件名"),
        ],
    },
    {
        "key": "job",
        "label": "岗位需求",
        "fields": [
            _field("job_public_name", "对外名称"),
            _field("job_position_name", "职位名称"),
            _field("job_family", "岗位族"),
            _field("job_secondary_department", "岗位二级部门"),
            _field("job_location", "地点"),
            _field("education_requirement", "学历要求"),
            _field("required_majors", "需求专业"),
            _field("responsibilities", "工作职责"),
            _field("headcount", "HC"),
            _field("is_public", "是否发布"),
        ],
    },
    {
        "key": "allocation",
        "label": "分配反馈",
        "fields": [
            _field("allocation_source", "分配来源"),
            _field("attempt_status", "尝试状态"),
            _field("allocation_secondary_department", "二级部门"),
            _field("secondary_contact", "二级接口人"),
            _field("tertiary_department", "三级部门"),
            _field("tertiary_contact", "三级接口人"),
            _field("allocation_reason", "匹配/人工理由"),
            _field("confidence_score", "置信度"),
            _field("feedback_result", "反馈结果"),
            _field("feedback_note", "反馈备注"),
            _field("dispatched_at", "下发时间"),
            _field("assigned_to_sub_at", "转派时间"),
            _field("feedback_at", "反馈时间"),
        ],
    },
    {
        "key": "status_reason",
        "label": "状态与原因",
        "fields": [
            _field("resume_status", "简历状态"),
            _field("reason_code", "原因码"),
            _field("reason_detail", "原因说明"),
            _field("workflow_status", "流程状态"),
            _field("archive_reason", "归档原因"),
            _field("archive_detail", "归档详情"),
        ],
    },
]

FIELD_CATALOG = {
    item["key"]: item
    for group in EXPORT_FIELD_GROUPS
    for item in group["fields"]
}
FIELD_ORDER = list(FIELD_CATALOG)


class ExportFieldError(ValueError):
    pass


class ExportOptionError(ValueError):
    pass


def export_fields_payload():
    return {"version": EXPORT_FIELDS_VERSION, "groups": EXPORT_FIELD_GROUPS}


def parse_export_fields(params):
    """未传时兼容旧调用；显式空值和未知字段均拒绝。"""
    if "fields" not in params:
        return [key for key in FIELD_ORDER if key in DEFAULT_FIELD_KEYS]
    raw = params.get("fields", "")
    requested = [item.strip() for item in str(raw).split(",") if item.strip()]
    if not requested:
        raise ExportFieldError("fields 不能为空，请至少选择一个导出字段")
    unknown = sorted(set(requested) - set(FIELD_CATALOG))
    if unknown:
        raise ExportFieldError(f"存在未知导出字段：{','.join(unknown)}")
    requested_set = set(requested)
    return [key for key in FIELD_ORDER if key in requested_set]


def parse_include_resume_files(params):
    """解析原件导出开关；缺失时兼容旧客户端，继续导出 ZIP。"""
    if "include_resume_files" not in params:
        return True
    raw = str(params.get("include_resume_files", "")).strip().lower()
    if raw in {"true", "1"}:
        return True
    if raw in {"false", "0"}:
        return False
    raise ExportOptionError("include_resume_files 必须是 true 或 false")


@dataclass
class CandidateExportRecord:
    candidate: m.Candidate
    current_resume: Optional[m.Resume]
    attempt: Optional[m.AssignmentAttempt]
    file_resumes: list


def _safe_excel_text(value):
    if not isinstance(value, str):
        return value
    if value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _time_text(value):
    if not value:
        return ""
    if timezone.is_naive(value):
        value = timezone.make_aware(value, ZoneInfo("Asia/Shanghai"))
    return value.astimezone(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d %H:%M:%S")


def _related_or_snapshot(attempt, snapshot_field, related_field):
    if not attempt:
        return ""
    snapshot = getattr(attempt, snapshot_field, "")
    related = getattr(attempt, related_field, None)
    return snapshot or (related.name if related else "")


def _job_secondary_department(job):
    if not job or not job.department:
        return ""
    department = job.department
    if department.level == 3 and department.parent:
        return department.parent.name
    return department.name


def _latest_processing_item(candidate):
    items = list(candidate.processing_scope_items.all())
    return max(items, key=lambda item: (item.created_at, item.id)) if items else None


def _public_ai_text(value):
    return (
        str(value or "")
        .replace("AI 专项强制分配", "AI 自动分配")
        .replace("AI 专项分流", "AI 后台分配")
    )


def _attempt_status(candidate, attempt):
    if attempt:
        if attempt.status == m.AssignmentAttempt.STATUS_PASSED:
            return system_status.SCREENING_PASSED
        if attempt.status == m.AssignmentAttempt.STATUS_REJECTED:
            return system_status.SCREENING_REJECTED
        if attempt.status in {
            m.AssignmentAttempt.STATUS_DISPATCHED_L2,
            m.AssignmentAttempt.STATUS_ASSIGNED_L3,
        }:
            return system_status.PENDING_SCREENING
        if attempt.status == m.AssignmentAttempt.STATUS_PENDING_REVIEW:
            return system_status.PENDING_REVIEW
        if attempt.status == m.AssignmentAttempt.STATUS_PENDING_DISPATCH:
            return system_status.PENDING_DISPATCH
    return system_status.candidate_system_status(candidate)


def _record_values(record):
    candidate = record.candidate
    resume = record.current_resume
    attempt = record.attempt
    job = resume.job if resume and resume.job_id else None
    workflow = candidate_summary.workflow_or_none(candidate)
    processing_item = _latest_processing_item(candidate)
    visible_resumes = {item.id: item for item in record.file_resumes}
    resumes = sorted(
        visible_resumes.values(),
        key=lambda item: (
            item.volunteer_rank if item.volunteer_rank is not None else 999,
            item.apply_date.toordinal() if item.apply_date else 0,
            item.id,
        ),
    )
    education_labels = dict(m.Candidate.HIGHEST_EDUCATION_CHOICES)
    source_labels = dict(m.AssignmentAttempt.SOURCE_CHOICES)
    attempt_status_labels = dict(m.AssignmentAttempt.STATUS_CHOICES)
    feedback_labels = dict(m.AssignmentAttempt.FEEDBACK_CHOICES)
    workflow_labels = dict(m.CandidateWorkflow.STATUS_CHOICES)
    archive_labels = dict(m.CandidateWorkflow.ARCHIVE_REASON_CHOICES)
    allocation_source = (
        attempt.source if attempt else candidate_summary.allocation_source(candidate)
    )
    resume_status = _attempt_status(candidate, attempt)
    allocation_department = _related_or_snapshot(
        attempt, "department_name_snapshot", "department"
    ) or _job_secondary_department(job)
    return {
        "candidate_name": candidate.name,
        "candidate_phone": candidate.phone,
        "gender": candidate.gender,
        "household_province": candidate.household_province,
        "highest_education": education_labels.get(
            candidate.highest_education, candidate.highest_education
        ),
        "highest_major": candidate.highest_major,
        "first_degree_school": candidate.first_degree_school,
        "first_degree_tag": (
            getattr(candidate.first_degree_tag, "name", "")
            or candidate.first_degree_platform
        ),
        "highest_degree_school": candidate.highest_degree_school,
        "highest_degree_tag": (
            getattr(candidate.highest_degree_tag, "name", "")
            or candidate.highest_degree_platform
        ),
        "candidate_imported_at": _time_text(candidate.imported_at),
        "current_apply_id": resume.apply_id if resume else "",
        "entity": resume.entity if resume else "",
        "org": resume.org if resume else "",
        "current_position_name": resume.position_name if resume else "",
        "original_status": resume.status if resume else "",
        "apply_date": resume.apply_date.isoformat() if resume and resume.apply_date else "",
        "volunteer_rank": resume.volunteer_rank if resume else "",
        "assigned_entity": resume.assigned_entity if resume else "",
        "job_category": resume.job_category if resume else "",
        "category_mode": resume.category_mode if resume else "",
        "category_reason": resume.category_reason if resume else "",
        "resume_filename": os.path.basename(resume.resume_file) if resume else "",
        "all_apply_ids": "、".join(item.apply_id for item in resumes if item.apply_id),
        "all_resume_filenames": "、".join(
            os.path.basename(item.resume_file) for item in resumes if item.resume_file
        ),
        "job_public_name": job.public_name if job else "",
        "job_position_name": job.position_name if job else "",
        "job_family": job.job_family if job else "",
        "job_secondary_department": _job_secondary_department(job),
        "job_location": job.location if job else "",
        "education_requirement": job.education if job else "",
        "required_majors": "、".join(
            item.major for item in job.majors.all()
        ) if job else "",
        "responsibilities": job.responsibilities if job else "",
        "headcount": job.headcount if job else "",
        "is_public": "是" if job and job.is_public else ("否" if job else ""),
        "allocation_source": source_labels.get(allocation_source, allocation_source),
        "attempt_status": attempt_status_labels.get(attempt.status, attempt.status)
        if attempt
        else "",
        "allocation_secondary_department": allocation_department,
        "secondary_contact": _related_or_snapshot(
            attempt, "contact_name_snapshot", "contact"
        ),
        "tertiary_department": _related_or_snapshot(
            attempt, "sub_department_name_snapshot", "sub_department"
        ),
        "tertiary_contact": _related_or_snapshot(
            attempt, "sub_contact_name_snapshot", "sub_contact"
        ),
        "allocation_reason": (
            attempt.manual_reason
            or (
                "AI 自动分配"
                if attempt.route_code == "ai_special_route"
                else _public_ai_text(attempt.match_reason)
            )
        )
        if attempt
        else "",
        "confidence_score": attempt.confidence_score if attempt else "",
        "feedback_result": feedback_labels.get(
            attempt.feedback_result, attempt.feedback_result
        ) if attempt else "",
        "feedback_note": attempt.feedback_note if attempt else "",
        "dispatched_at": _time_text(attempt.dispatched_at) if attempt else "",
        "assigned_to_sub_at": _time_text(attempt.assigned_to_sub_at)
        if attempt else "",
        "feedback_at": _time_text(attempt.feedback_at) if attempt else "",
        "resume_status": system_status.system_status_label(resume_status),
        "reason_code": (
            {
                "ai_special_route": "ai_dispatched",
                "ai_special_route_unavailable": "ai_connection_error",
            }.get(processing_item.reason_code, processing_item.reason_code)
            if processing_item
            else ""
        ),
        "reason_detail": _public_ai_text(processing_item.result_message)
        if processing_item else "",
        "workflow_status": workflow_labels.get(workflow.status, workflow.status)
        if workflow else "",
        "archive_reason": (
            archive_labels.get(workflow.archive_reason, workflow.archive_reason)
            if workflow
            else ""
        ),
        "archive_detail": workflow.archive_detail if workflow else "",
    }


def build_resume_export_workbook(records, field_keys):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "简历库"
    sheet.append([FIELD_CATALOG[key]["label"] for key in field_keys])
    for record in records:
        values = _record_values(record)
        sheet.append([_safe_excel_text(values[key]) for key in field_keys])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for index, key in enumerate(field_keys, start=1):
        column_letter = sheet.cell(row=1, column=index).column_letter
        values = [sheet.cell(row=row, column=index).value for row in range(1, sheet.max_row + 1)]
        sheet.column_dimensions[column_letter].width = min(
            60 if key == "responsibilities" else 36,
            max(10, max(len(str(value or "")) for value in values) + 2),
        )
        if key == "responsibilities":
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=index).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _resume_file_info(resume):
    filename = os.path.basename(resume.resume_file or "")
    path = os.path.join(settings.MEDIA_ROOT, RESUME_SUBDIR, filename) if filename else ""
    return (filename, path) if path and os.path.isfile(path) else ("", "")


def _deduplicated_records(records):
    deduplicated = {}
    for record in records:
        deduplicated[record.candidate.id] = record
    return list(deduplicated.values())


def build_resume_export_excel(records, field_keys):
    records = _deduplicated_records(records)
    return build_resume_export_workbook(records, field_keys), len(records)


def build_resume_export_zip(records, field_keys):
    records = _deduplicated_records(records)
    file_resumes = []
    seen_resume_ids = set()
    for record in records:
        for resume in record.file_resumes:
            if resume.id not in seen_resume_ids:
                seen_resume_ids.add(resume.id)
                file_resumes.append(resume)

    available = []
    missing = []
    for resume in file_resumes:
        filename, path = _resume_file_info(resume)
        if path:
            available.append((resume, filename, path))
        else:
            missing.append(f"{resume.candidate.name}（{resume.apply_id}）")

    output = BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "简历库清单.xlsx",
            build_resume_export_workbook(records, field_keys),
        )
        archive.writestr("简历文件/", b"")
        used_names = set()
        filename_counts = {}
        for _, filename, _ in available:
            filename_counts[filename] = filename_counts.get(filename, 0) + 1
        for resume, filename, path in available:
            if filename_counts[filename] > 1:
                stem, extension = os.path.splitext(filename)
                archive_name = f"简历文件/{stem}（{resume.apply_id}）{extension}"
            else:
                archive_name = f"简历文件/{filename}"
            if archive_name in used_names:
                stem, extension = os.path.splitext(filename)
                suffix = 2
                while archive_name in used_names:
                    archive_name = (
                        f"简历文件/{stem}（{resume.apply_id}-{suffix}）{extension}"
                    )
                    suffix += 1
            used_names.add(archive_name)
            archive.write(path, arcname=archive_name)
        if missing:
            archive.writestr(
                "缺失简历文件清单.txt",
                "以下应聘记录暂无简历文件（未上传简历包或未匹配）：\n"
                + "\n".join(missing),
            )
    return output.getvalue(), len(available), len(missing), len(records)
