"""按应聘 ID 生成简历处理结果 Excel 报表。"""

from collections import OrderedDict
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from apps.core import candidate_summary
from apps.core import models as m
from apps.core import system_status

from .resume_export import (
    attempt_processing_values,
    primary_department,
    receiving_secondary_department,
)


SUMMARY_HEADERS = [
    "当前接收一级部门",
    "当前接收二级部门",
    "导入简历数",
    "分配简历数",
    "待处理",
    "已归档",
    "待重新分配",
    "待复核",
    "待下发",
    "待业务反馈",
    "通过",
    "不通过",
]
DETAIL_HEADERS = [
    "导入时间",
    "姓名",
    "应聘ID",
    "招聘主体",
    "岗位",
    "志愿",
    "最高学历",
    "首次部门",
    "当前接收一级部门",
    "当前接收二级部门",
    "当前接收节点",
    "分配来源",
    "简历状态",
    "反馈结果",
    "不通过原因码",
    "不通过原因",
    "反馈备注",
    "首次下发时间",
    "当前部门进入时间",
    "反馈时间",
    "HR 下发时长（小时）",
    "当前部门处理时长（小时）",
    "总反馈时长（小时）",
]
REJECTION_SUMMARY_HEADERS = [
    "当前接收一级部门",
    "当前接收二级部门",
    "不通过原因码",
    "不通过原因",
    "数量",
]
STATUS_ORDER = [
    system_status.RAW,
    system_status.ARCHIVED,
    system_status.PENDING_REALLOCATION,
    system_status.PENDING_REVIEW,
    system_status.PENDING_DISPATCH,
    system_status.PENDING_SCREENING,
    system_status.SCREENING_PASSED,
    system_status.SCREENING_REJECTED,
]


def safe_excel_text(value):
    """阻止外部文本被 Excel 当成公式执行。"""
    if not isinstance(value, str):
        return value
    stripped = value.lstrip()
    if stripped.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def current_effective_attempt(report_resume):
    workflow = candidate_summary.workflow_or_none(report_resume.candidate)
    current = candidate_summary.current_resume(report_resume.candidate)
    if not workflow or not current:
        return None
    return candidate_summary.latest_effective_attempt(
        workflow,
        resume_id=current.id,
    )


def resume_report_status(resume, attempt):
    return system_status.candidate_system_status(resume.candidate)


def _summary_bucket():
    return {
        "imported": 0,
        "allocated": 0,
        **{status: 0 for status in STATUS_ORDER},
    }


def build_result_report(resumes):
    rows = []
    summaries = OrderedDict()
    rejection_summaries = OrderedDict()
    total = _summary_bucket()
    education_labels = dict(m.Candidate.HIGHEST_EDUCATION_CHOICES)
    source_labels = dict(m.AssignmentAttempt.SOURCE_CHOICES)
    feedback_labels = dict(m.AssignmentAttempt.FEEDBACK_CHOICES)
    reason_labels = dict(m.AssignmentAttempt.REJECTION_REASON_CHOICES)
    as_of = timezone.now()

    for resume in resumes:
        attempt = current_effective_attempt(resume)
        status_code = resume_report_status(resume, attempt)
        initial_department = attempt.initial_department if attempt else None
        current_department = attempt.current_department if attempt else None
        current_primary = primary_department(current_department)
        current_secondary = receiving_secondary_department(current_department)
        if current_primary:
            primary_name = current_primary.name
        elif current_department:
            primary_name = "未归属一级部门"
        else:
            primary_name = "未分配"
        secondary_name = (
            current_secondary.name if current_secondary else "未分配"
        )
        group_key = (primary_name, secondary_name)
        bucket = summaries.setdefault(group_key, _summary_bucket())
        for target in (bucket, total):
            target["imported"] += 1
            target[status_code] += 1
            if attempt:
                target["allocated"] += 1

        imported_at = timezone.localtime(resume.imported_at).strftime(
            "%Y-%m-%d %H:%M:%S"
        )
        feedback_result = ""
        feedback_reason = ""
        if attempt:
            feedback_result = feedback_labels.get(
                attempt.feedback_result, attempt.feedback_result
            )
            feedback_reason = (
                attempt.feedback_reason_label_snapshot
                or reason_labels.get(
                    attempt.feedback_reason_code, attempt.feedback_reason_code
                )
            )
        if (
            attempt
            and attempt.status == m.AssignmentAttempt.STATUS_REJECTED
            and attempt.feedback_reason_code
        ):
            rejection_key = (
                primary_name,
                secondary_name,
                attempt.feedback_reason_code,
            )
            rejection_summaries[rejection_key] = (
                rejection_summaries.get(rejection_key, 0) + 1
            )
        timings = attempt_processing_values(attempt, as_of=as_of)
        rows.append(
            [
                imported_at,
                resume.candidate.name,
                resume.apply_id,
                resume.entity,
                resume.position_name,
                resume.volunteer_rank or "",
                education_labels.get(
                    resume.candidate.highest_education,
                    resume.candidate.highest_education,
                ),
                initial_department.name if initial_department else "",
                current_primary.name if current_primary else "",
                current_secondary.name if current_secondary else "",
                current_department.name if current_department else "",
                source_labels.get(attempt.source, attempt.source) if attempt else "",
                system_status.system_status_label(status_code),
                feedback_result,
                attempt.feedback_reason_code if attempt else "",
                feedback_reason,
                attempt.feedback_note if attempt else "",
                timings["first_dispatched_at"],
                timings["current_department_entered_at"],
                timings["feedback_at"],
                timings["hr_dispatch_duration_hours"],
                timings["current_department_duration_hours"],
                timings["total_feedback_duration_hours"],
            ]
        )

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "部门汇总"
    detail_sheet = workbook.create_sheet("简历明细")
    rejection_sheet = workbook.create_sheet("不通过原因汇总")
    summary_sheet.append(SUMMARY_HEADERS)
    unassigned_key = ("未分配", "未分配")
    ordered_names = sorted(name for name in summaries if name != unassigned_key)
    if unassigned_key in summaries:
        ordered_names.append(unassigned_key)
    for primary_name, secondary_name in ordered_names:
        bucket = summaries[(primary_name, secondary_name)]
        summary_sheet.append(
            [
                safe_excel_text(primary_name),
                safe_excel_text(secondary_name),
                bucket["imported"],
                bucket["allocated"],
            ]
            + [bucket[status] for status in STATUS_ORDER]
        )
    summary_sheet.append(
        ["合计", "", total["imported"], total["allocated"]]
        + [total[status] for status in STATUS_ORDER]
    )

    detail_sheet.append(DETAIL_HEADERS)
    for row in rows:
        detail_sheet.append([safe_excel_text(value) for value in row])
    rejection_sheet.append(REJECTION_SUMMARY_HEADERS)
    for key in sorted(rejection_summaries):
        primary_name, secondary_name, reason_code = key
        rejection_sheet.append(
            [
                safe_excel_text(primary_name),
                safe_excel_text(secondary_name),
                safe_excel_text(reason_code),
                safe_excel_text(reason_labels.get(reason_code, reason_code)),
                rejection_summaries[key],
            ]
        )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in (summary_sheet, detail_sheet, rejection_sheet):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column_cells in sheet.columns:
            width = min(
                36,
                max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2),
            )
            sheet.column_dimensions[column_cells[0].column_letter].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
