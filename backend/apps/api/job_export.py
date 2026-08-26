"""生成可直接回导的职位清单 Excel。"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.core.departments import resolve_department_hierarchy
from apps.ingestion.tabular_imports import get_import_table_schema


JOB_EXPORT_HEADERS = list(get_import_table_schema("jobs").headers)


def safe_excel_value(value):
    """阻止岗位主数据被 Excel 当作公式执行。"""
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _job_row(job):
    hierarchy = resolve_department_hierarchy(job.department)
    return [
        job.entity,
        hierarchy.primary.name if hierarchy.primary else "",
        hierarchy.secondary.name if hierarchy.secondary else "",
        job.category,
        job.public_name,
        "是" if job.is_public else "否",
        job.position_name,
        job.job_family,
        job.location,
        job.education,
        job.responsibilities,
        "、".join(item.major for item in job.majors.all()),
        job.headcount,
    ]


def build_job_export_workbook(jobs):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "职位清单"
    sheet.append(JOB_EXPORT_HEADERS)
    for job in jobs:
        sheet.append([safe_excel_value(value) for value in _job_row(job)])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    preferred_widths = {
        "工作职责": 60,
        "需求专业": 36,
        "对外发布名称": 24,
        "职位名称": 24,
    }
    for index, header in enumerate(JOB_EXPORT_HEADERS, start=1):
        values = [
            sheet.cell(row=row, column=index).value
            for row in range(1, sheet.max_row + 1)
        ]
        natural_width = max(10, max(len(str(value or "")) for value in values) + 2)
        maximum_width = preferred_widths.get(header, 28)
        column_letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column_letter].width = min(maximum_width, natural_width)
        if header == "工作职责":
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=index).alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
